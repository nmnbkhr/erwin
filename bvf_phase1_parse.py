#!/usr/bin/env python3
"""Phase 1: Parse the BVF XLSM File"""

import json
import os
from openpyxl import load_workbook

OUTPUT_DIR = "/mnt/e/erwin/bvf_output"
XLSM_FILE = "/mnt/e/erwin/Banking Business Value Framework Data Mappings 1.2.xlsm"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def parse_capability_to_data_sheet(wb):
    """Parse 'Banking BVF Capability to Data' sheet."""
    ws = wb["Banking BVF Capability to Data"]

    # Read subject areas from row 3 (cols 4+)
    subject_areas = {}
    for col in range(4, 120):
        val = ws.cell(row=3, column=col+1).value  # openpyxl is 1-based
        if val:
            subject_areas[col] = str(val).strip()

    # Read data requirement names from row 4 (cols 4+)
    data_req_names = {}
    for col in range(4, 120):
        val = ws.cell(row=4, column=col+1).value
        if val:
            data_req_names[col] = str(val).strip()

    # Forward-fill subject areas for data requirements
    last_sa = None
    for col in sorted(subject_areas.keys()):
        last_sa = subject_areas[col]

    data_req_subject_areas = {}
    last_sa = None
    for col in sorted(data_req_names.keys()):
        if col in subject_areas:
            last_sa = subject_areas[col]
        data_req_subject_areas[col] = last_sa

    print(f"  Found {len(data_req_names)} data requirements across columns")
    print(f"  Subject areas: {sorted(set(data_req_subject_areas.values()))}")

    # Read capability rows (rows 5-116)
    capabilities = []
    current_area = None
    current_sub_area = None

    for row in range(5, 120):
        # Col 0 = BVF Area (forward-fill)
        area_val = ws.cell(row=row+1, column=1).value
        if area_val:
            current_area = str(area_val).strip()

        # Col 1 = Sub-area (forward-fill)
        sub_val = ws.cell(row=row+1, column=2).value
        if sub_val:
            current_sub_area = str(sub_val).strip()

        # Col 3 = Capability name
        cap_name = ws.cell(row=row+1, column=4).value
        if not cap_name:
            continue
        cap_name = str(cap_name).strip()
        if not cap_name:
            continue

        # Cols 4-116: Binary mappings
        data_reqs = []
        output_data = []
        for col in sorted(data_req_names.keys()):
            cell_val = ws.cell(row=row+1, column=col+1).value
            if cell_val is not None:
                if str(cell_val).strip().upper() == "OUTPUT":
                    output_data.append(data_req_names[col])
                elif str(cell_val).strip() == "1" or cell_val == 1:
                    data_reqs.append(data_req_names[col])

        # Col 117 = total count
        total_val = ws.cell(row=row+1, column=118).value

        capabilities.append({
            "name": cap_name,
            "bvf_area": current_area,
            "sub_area": current_sub_area,
            "data_requirements": data_reqs,
            "output_data": output_data,
            "data_req_count": len(data_reqs) + len(output_data),
            "stated_count": int(total_val) if total_val else 0
        })

    print(f"  Parsed {len(capabilities)} capabilities")
    return capabilities, data_req_names, data_req_subject_areas


def parse_data_to_capability_sheet(wb):
    """Parse 'Banking BVF Data to Capability' sheet."""
    ws = wb["Banking BVF Data to Capability"]

    # Row 4 (cols 4-115): Capability names as headers
    cap_headers = {}
    for col in range(4, 120):
        val = ws.cell(row=4, column=col+1).value
        if val:
            cap_headers[col] = str(val).strip()

    print(f"  Found {len(cap_headers)} capability columns in Data-to-Capability sheet")

    # Rows 5+: Data requirement rows
    data_requirements = []
    for row in range(5, 130):
        dr_name = ws.cell(row=row+1, column=1).value
        if not dr_name:
            continue
        dr_name = str(dr_name).strip()
        if not dr_name:
            continue

        fsdm_sa = ws.cell(row=row+1, column=2).value
        sort_order = ws.cell(row=row+1, column=3).value
        line_sort = ws.cell(row=row+1, column=4).value

        caps_using = []
        output_of = []
        for col in sorted(cap_headers.keys()):
            cell_val = ws.cell(row=row+1, column=col+1).value
            if cell_val is not None:
                if str(cell_val).strip().upper() == "OUTPUT":
                    output_of.append(cap_headers[col])
                elif str(cell_val).strip() == "1" or cell_val == 1:
                    caps_using.append(cap_headers[col])

        data_requirements.append({
            "name": dr_name,
            "fsdm_subject_area": str(fsdm_sa).strip() if fsdm_sa else "",
            "sort_order": int(sort_order) if sort_order else 0,
            "line_sort_order": int(line_sort) if line_sort else 0,
            "capability_count": len(caps_using) + len(output_of),
            "capabilities": caps_using,
            "output_of": output_of
        })

    print(f"  Parsed {len(data_requirements)} data requirements")
    return data_requirements


def parse_reuse_matrix(wb):
    """Parse 'Banking BVF Data Reuse Matrix' sheet."""
    ws = wb["Banking BVF Data Reuse Matrix"]

    # Read capability names from row headers and column headers
    # Row 3 or 4: column headers (capability names)
    col_caps = {}
    for col in range(2, 130):
        val = ws.cell(row=4, column=col+1).value
        if val and str(val).strip():
            col_caps[col] = str(val).strip()

    if not col_caps:
        # Try row 3
        for col in range(2, 130):
            val = ws.cell(row=3, column=col+1).value
            if val and str(val).strip():
                col_caps[col] = str(val).strip()

    print(f"  Found {len(col_caps)} columns in reuse matrix")

    # Read matrix rows
    reuse_matrix = {}
    for row in range(4, 130):
        row_cap = ws.cell(row=row+1, column=1).value
        if not row_cap:
            row_cap = ws.cell(row=row+1, column=2).value
        if not row_cap:
            continue
        row_cap = str(row_cap).strip()
        if not row_cap:
            continue

        reuse_matrix[row_cap] = {}
        for col, col_cap in col_caps.items():
            val = ws.cell(row=row+1, column=col+1).value
            if val is not None:
                try:
                    coeff = float(val)
                    if coeff > 0 and row_cap != col_cap:
                        reuse_matrix[row_cap][col_cap] = round(coeff, 4)
                except (ValueError, TypeError):
                    pass

    # Remove empty entries
    reuse_matrix = {k: v for k, v in reuse_matrix.items() if v}
    print(f"  Parsed reuse matrix for {len(reuse_matrix)} capabilities")
    return reuse_matrix


def main():
    print("=" * 70)
    print("Phase 1: Parse BVF XLSM File")
    print("=" * 70)

    print(f"\nLoading {XLSM_FILE}...")
    wb = load_workbook(XLSM_FILE, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    # 1.1 Parse Capability to Data
    print("\n--- Parsing 'Banking BVF Capability to Data' ---")
    capabilities, data_req_names, data_req_sas = parse_capability_to_data_sheet(wb)

    # 1.2 Parse Data to Capability
    print("\n--- Parsing 'Banking BVF Data to Capability' ---")
    data_requirements = parse_data_to_capability_sheet(wb)

    # 1.3 Parse Reuse Matrix
    print("\n--- Parsing 'Banking BVF Data Reuse Matrix' ---")
    reuse_matrix = parse_reuse_matrix(wb)

    wb.close()

    # Enrich data requirements with subject areas from cap-to-data sheet
    dr_sa_map = {}
    for col, name in data_req_names.items():
        if col in data_req_sas and data_req_sas[col]:
            dr_sa_map[name] = data_req_sas[col]

    for dr in data_requirements:
        if not dr["fsdm_subject_area"] and dr["name"] in dr_sa_map:
            dr["fsdm_subject_area"] = dr_sa_map[dr["name"]]

    # Summary
    print("\n" + "=" * 70)
    print("BVF PARSING SUMMARY")
    print("=" * 70)

    areas = {}
    for cap in capabilities:
        area = cap["bvf_area"] or "Unknown"
        sub = cap["sub_area"] or "Unknown"
        if area not in areas:
            areas[area] = {}
        if sub not in areas[area]:
            areas[area][sub] = []
        areas[area][sub].append(cap["name"])

    for area, subs in areas.items():
        total = sum(len(v) for v in subs.values())
        print(f"\n  {area} ({total} capabilities)")
        for sub, caps in subs.items():
            print(f"    {sub}: {len(caps)} capabilities")

    sa_counts = {}
    for dr in data_requirements:
        sa = dr["fsdm_subject_area"] or "Unknown"
        sa_counts[sa] = sa_counts.get(sa, 0) + 1
    print(f"\n  Data Requirements by FSDM Subject Area:")
    for sa, count in sorted(sa_counts.items()):
        print(f"    {sa}: {count}")

    total_mappings = sum(cap["data_req_count"] for cap in capabilities)
    print(f"\n  Total capability-data mappings: {total_mappings}")

    # Save outputs
    print("\n--- Saving outputs ---")

    out1 = os.path.join(OUTPUT_DIR, "bvf_parsed_capabilities.json")
    with open(out1, "w") as f:
        json.dump(capabilities, f, indent=2)
    print(f"  Saved: {out1} ({len(capabilities)} capabilities)")

    out2 = os.path.join(OUTPUT_DIR, "bvf_parsed_data_requirements.json")
    with open(out2, "w") as f:
        json.dump(data_requirements, f, indent=2)
    print(f"  Saved: {out2} ({len(data_requirements)} data requirements)")

    out3 = os.path.join(OUTPUT_DIR, "bvf_reuse_matrix.json")
    with open(out3, "w") as f:
        json.dump(reuse_matrix, f, indent=2)
    print(f"  Saved: {out3} ({len(reuse_matrix)} capability entries)")

    print("\nPhase 1 complete!")


if __name__ == "__main__":
    main()

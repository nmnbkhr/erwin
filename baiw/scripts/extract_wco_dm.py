#!/usr/bin/env python3
"""
Extract real WCO DM v4.2.0 data from Excel files and merge with TAIW JSON.
Preserves all TAIW-specific fields while replacing WCO DM data layer.
"""

import json
import os
import openpyxl
from collections import defaultdict

BASE = "/home/adnoman/projects/wco_data_model/wco_dm_app_v4.2.0"
TAIW_DIR = "/mnt/e/erwin/baiw/src/data/taiw"

LIB_PATH = os.path.join(BASE, "1. Library", "1.1 WCO DM LIB_V4.2.0.xlsx")
DEC_PATH = os.path.join(BASE, "3. Information Packages", "3.1 Data Model", "Section 3.3. WCO DM BIP DEC_V4.2.0.xlsx")
LPCO_PATH = os.path.join(BASE, "3. Information Packages", "3.1 Data Model", "Section 3.4. WCO DM BIP LPCO_V4.2.0.xlsx")
AIP_PATH = os.path.join(BASE, "3. Information Packages", "3.1 Data Model", "Section 3.2. WCO DM AIP_V4.2.0.xlsx")
MDIP_PATH = os.path.join(BASE, "3. Information Packages", "3.1 Data Model", "Section 3.1. WCO DM MDIP_V4.2.0.xlsx")


# ── Domain inference from UML path and superclass ──
DOMAIN_MAP = {
    "Declaration": "Declaration",
    "DeclarationGovernment": "Declaration",
    "GoodsShipment": "Consignment",
    "Consignment": "Consignment",
    "GovernmentAgencyGoods": "Goods",
    "Commodity": "Goods",
    "GoodsItem": "Goods",
    "Packaging": "Goods",
    "DangerousGoods": "Goods",
    "TransportEquipment": "Transport",
    "TransportMeans": "Transport",
    "ArrivalTransportMeans": "Transport",
    "DepartureTransportMeans": "Transport",
    "BorderTransportMeans": "Transport",
    "Carrier": "Transport",
    "Crew": "Transport",
    "Passenger": "Transport",
    "Agent": "Party",
    "Buyer": "Party",
    "Seller": "Party",
    "Consignee": "Party",
    "Consignor": "Party",
    "Importer": "Party",
    "Exporter": "Party",
    "Declarant": "Party",
    "Notifier": "Party",
    "Broker": "Party",
    "Manufacturer": "Party",
    "Producer": "Party",
    "Packer": "Party",
    "Surety": "Party",
    "RepresentativePerson": "Party",
    "PersonOnBoard": "Party",
    "TradingParty": "Party",
    "PropertyOwner": "Party",
    "PropertyOperator": "Party",
    "RecognizedSecurityOrganization": "Party",
    "Address": "Location",
    "Location": "Location",
    "LoadingLocation": "Location",
    "UnloadingLocation": "Location",
    "AcceptanceLocation": "Location",
    "ExitLocation": "Location",
    "EntryLocation": "Location",
    "BorderingOffice": "Location",
    "AppealOffice": "Location",
    "CustomsOffice": "Location",
    "OriginTradeCountry": "Origin",
    "ExportTradeCountry": "Origin",
    "DestinationTradeCountry": "Origin",
    "Origin": "Origin",
    "Certificate": "Document",
    "AdditionalDocument": "Document",
    "PreviousDocument": "Document",
    "TransportDocument": "Document",
    "WrittenOff": "Document",
    "UCR": "Document",
    "Submitter": "Document",
    "Amendment": "Document",
    "AdditionalInformation": "Document",
    "Invoice": "Financial",
    "DutyTaxFee": "Financial",
    "Payment": "Financial",
    "CustomsValuation": "Financial",
    "Charges": "Financial",
    "Guarantee": "Financial",
    "Bond": "Financial",
    "Classification": "Classification",
    "Tariff": "Classification",
    "AdditionalProcedure": "Classification",
    "RequestedProcedure": "Classification",
    "Warehouse": "Warehouse",
    "TemporaryStorage": "Warehouse",
    "FreeZone": "Warehouse",
    "Seal": "Risk & Control",
    "Examination": "Risk & Control",
    "RiskAssessment": "Risk & Control",
    "Control": "Risk & Control",
    "SelectionCriteria": "Risk & Control",
    "Observation": "Risk & Control",
    "ECommerce": "E-Commerce",
    "PostalItem": "E-Commerce",
    "Response": "Response",
    "Error": "Response",
    "Status": "Response",
    "Notification": "Response",
    "Decision": "Response",
    # UML path segments (deeper hierarchy)
    "GoodsItem": "Goods",
    "ConsignmentItem": "Goods",
    "GoodsLocation": "Location",
    "LoadingLocation": "Location",
    "UnloadingLocation": "Location",
    "TransitDestinationLocation": "Location",
    "TranshipmentLocation": "Location",
    "ArrivalConveyanceFacility": "Location",
    "DepartureLocation": "Location",
    "BorderTransportMeans": "Transport",
    "ArrivalTransportMeans": "Transport",
    "DepartureTransportMeans": "Transport",
    "TransitTransportMeans": "Transport",
    "Reservation": "Transport",
    "Itinerary": "Transport",
    "Bunker": "Transport",
    "PersonOnBoard": "Transport",
    "Operator": "Party",
    "Principal": "Party",
    "Authentication": "Document",
    "Authenticator": "Document",
    "BinaryFile": "Document",
    "TransportContractDocument": "Document",
    "DutyTaxFee": "Financial",
    "Payment": "Financial",
    "ObligationGuarantee": "Financial",
    "TransitOperation": "Financial",
    "TradeTerms": "Financial",
    "Waste": "Goods",
    "HouseConsignment": "Consignment",
    "DeclarationOffice": "Declaration",
    "DeclarationGovernment": "Declaration",
    "CustomsOffice": "Declaration",
    "DestinationCountry": "Origin",
    "OriginCountry": "Origin",
    "ExportCountry": "Origin",
    "Destination": "Origin",
    "TransitCountry": "Origin",
}

SUPERCLASS_DOMAIN = {
    "DOC": "Document",
    "GOV": "Declaration",
    "LOC": "Location",
    "PAR": "Party",
    "TRA": "Transport",
    "NONE": None,
}

def infer_domain(uml_path, class_name, superclass=None):
    """Infer domain from UML path, class name, or superclass marker."""
    # For elements: look at the UML path segments for the BEST domain match
    if uml_path:
        parts = [p.strip().split(".")[0] for p in uml_path.split("/")]
        # Try each segment from deepest to shallowest (skip the root "Declaration"/"LPCO"/etc)
        for p in reversed(parts):
            if p in DOMAIN_MAP:
                return DOMAIN_MAP[p]
        # Try second segment (the main class under Declaration/LPCO)
        if len(parts) >= 2 and parts[1] in DOMAIN_MAP:
            return DOMAIN_MAP[parts[1]]

    # Try direct class name mapping
    if class_name in DOMAIN_MAP:
        return DOMAIN_MAP[class_name]

    # Try superclass
    if superclass and superclass in SUPERCLASS_DOMAIN and SUPERCLASS_DOMAIN[superclass]:
        return SUPERCLASS_DOMAIN[superclass]
    return "Other"


def extract_classes():
    """Extract all 273 classes from LIB DataSet 1.2/1.3/1.4."""
    print("Extracting classes from Library...")
    wb = openpyxl.load_workbook(LIB_PATH, read_only=True, data_only=True)
    ws = wb["DataSet 1.2, 1.3, 1.4"]

    classes = {}
    current_class = None

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        # Row structure: NONE, DOC, GOV, LOC, PAR, TRA, UML Path, ClassID, ClassName,
        #                AttrID, AttrName, Definition, Format, CodeList, CodeID, CodeDesc,
        #                CoreDataType, ...
        none_flag, doc_flag, gov_flag, loc_flag, par_flag, tra_flag = row[0:6]
        uml_path = row[6]
        class_id = row[7]
        class_name = row[8]
        attr_id = row[9]
        attr_name = row[10]
        definition = row[11]
        fmt = row[12]
        code_list = row[13]
        core_data_type = row[16]

        # Determine superclass
        superclass = "NONE"
        if doc_flag: superclass = "DOC"
        elif gov_flag: superclass = "GOV"
        elif loc_flag: superclass = "LOC"
        elif par_flag: superclass = "PAR"
        elif tra_flag: superclass = "TRA"

        if class_id and class_name and not attr_id:
            # This is a class definition row
            domain = infer_domain(str(uml_path or ""), class_name, superclass)
            current_class = {
                "id": str(class_id).strip(),
                "name": str(class_name).strip(),
                "domain": domain,
                "description": str(definition).strip() if definition else "",
                "superclass": superclass,
                "stereotype": "ABIE",
                "attributes": [],
                "elementCount": 0,
                "parentClass": None,
                "childClasses": [],
            }
            classes[current_class["id"]] = current_class
        elif attr_id and current_class:
            # This is an attribute row
            current_class["attributes"].append({
                "id": str(attr_id).strip(),
                "name": str(attr_name).strip() if attr_name else "",
                "definition": str(definition).strip() if definition else "",
                "format": str(fmt).strip() if fmt else "",
                "codeList": str(code_list).strip() if code_list else "",
                "coreDataType": str(core_data_type).strip() if core_data_type else "",
            })
            current_class["elementCount"] += 1

    wb.close()

    # Build parent-child relationships from UML paths
    class_list = list(classes.values())
    class_names = {c["name"]: c["id"] for c in class_list}

    print(f"  Extracted {len(class_list)} classes")
    return class_list


def extract_elements_from_sheet(path, sheet_name, package_name):
    """Extract data elements from a BIP/AIP/MDIP sheet."""
    print(f"  Extracting from {package_name} ({sheet_name})...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    elements = []
    seen_ids = set()

    # Find header row (row 3 typically)
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        uml_path = row[0]
        wco_id = row[1]
        wco_name = row[2]
        wco_def = row[3]

        if not wco_id or not wco_name:
            continue

        wco_id_str = str(wco_id).strip()
        if wco_id_str in seen_ids:
            continue
        seen_ids.add(wco_id_str)

        # Get format and code list (position varies by file)
        if package_name == "MDIP":
            fmt = row[4] if len(row) > 4 else None
            code_list = row[7] if len(row) > 7 else None
        else:
            remark = row[4] if len(row) > 4 else None
            fmt = row[5] if len(row) > 5 else None
            code_list = row[8] if len(row) > 8 else None

        # Infer class and domain from UML path
        class_name = ""
        domain = "Other"
        if uml_path:
            parts = str(uml_path).split("/")
            # Find the class part (usually the second-to-last segment before the attribute)
            for p in reversed(parts):
                p = p.strip().split(".")[0]
                if p and p != package_name and p not in ("Declaration", "LPCO", "AIP", "MDIP"):
                    class_name = p
                    break
            if not class_name and len(parts) > 1:
                class_name = parts[1].strip().split(".")[0] if len(parts) > 1 else parts[0]
            domain = infer_domain(str(uml_path), class_name)

        elements.append({
            "id": wco_id_str,
            "name": str(wco_name).strip(),
            "definition": str(wco_def).strip() if wco_def else "",
            "class": class_name,
            "domain": domain,
            "dataType": str(fmt).strip() if fmt else "",
            "codeList": str(code_list).strip() if code_list else "",
            "format": str(fmt).strip() if fmt else "",
            "informationPackages": [package_name],
            "umlPath": str(uml_path).strip() if uml_path else "",
        })

    wb.close()
    print(f"    Found {len(elements)} unique elements")
    return elements


def extract_all_elements():
    """Extract and merge elements from all 4 information packages."""
    print("Extracting data elements...")

    all_elements = {}

    # Declaration BIP (largest)
    dec_elems = extract_elements_from_sheet(DEC_PATH, "BIP DEC", "Declaration")
    for e in dec_elems:
        all_elements[e["id"]] = e

    # LPCO BIP
    lpco_elems = extract_elements_from_sheet(LPCO_PATH, "BIP LPCO", "LPCO")
    for e in lpco_elems:
        if e["id"] in all_elements:
            if "LPCO" not in all_elements[e["id"]]["informationPackages"]:
                all_elements[e["id"]]["informationPackages"].append("LPCO")
        else:
            all_elements[e["id"]] = e

    # AIP
    aip_elems = extract_elements_from_sheet(AIP_PATH, "AIP (RES and InterGov)", "AIP")
    for e in aip_elems:
        if e["id"] in all_elements:
            if "AIP" not in all_elements[e["id"]]["informationPackages"]:
                all_elements[e["id"]]["informationPackages"].append("AIP")
        else:
            all_elements[e["id"]] = e

    # MDIP
    mdip_elems = extract_elements_from_sheet(MDIP_PATH, "MDIP (MetaData)", "MDIP")
    for e in mdip_elems:
        if e["id"] in all_elements:
            if "MDIP" not in all_elements[e["id"]]["informationPackages"]:
                all_elements[e["id"]]["informationPackages"].append("MDIP")
        else:
            all_elements[e["id"]] = e

    result = list(all_elements.values())
    print(f"  Total unique elements: {len(result)}")
    return result


def extract_code_lists():
    """Extract all 135 code lists from LIB DataSet 1.6."""
    print("Extracting code lists...")
    wb = openpyxl.load_workbook(LIB_PATH, read_only=True, data_only=True)
    ws = wb["DataSet 1.6"]

    code_lists = []
    current_cl = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        seq = row[0]
        cl_name = row[1]
        cl_source = row[2]
        code_id = row[3]
        code_desc = row[4]

        if seq and cl_name:
            # New code list
            if current_cl:
                code_lists.append(current_cl)
            current_cl = {
                "id": f"CL_{int(seq):03d}",
                "name": str(cl_name).strip(),
                "source": str(cl_source).strip() if cl_source else "WCO",
                "values": [],
            }
            if code_id:
                current_cl["values"].append({
                    "code": str(code_id).strip(),
                    "description": str(code_desc).strip() if code_desc else "",
                })
        elif current_cl and code_id:
            current_cl["values"].append({
                "code": str(code_id).strip(),
                "description": str(code_desc).strip() if code_desc else "",
            })

    if current_cl:
        code_lists.append(current_cl)

    wb.close()
    print(f"  Extracted {len(code_lists)} code lists")
    return code_lists


def extract_core_data_types():
    """Extract 10 Core Data Types from LIB DataSet 1.5."""
    print("Extracting Core Data Types...")
    wb = openpyxl.load_workbook(LIB_PATH, read_only=True, data_only=True)
    ws = wb["DataSet 1.5"]

    cdts = []
    current_cdt = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        seq = row[0]
        name = row[1]
        attr_name = row[2]
        code_list = row[3]
        remarks = row[4]

        if seq and name and not current_cdt or (seq and name and name != (current_cdt or {}).get("name")):
            if current_cdt:
                cdts.append(current_cdt)
            current_cdt = {
                "id": f"CDT_{int(seq):02d}",
                "name": str(name).strip(),
                "attributes": [],
                "remarks": str(remarks).strip() if remarks else "",
            }
        if attr_name and current_cdt:
            current_cdt["attributes"].append({
                "name": str(attr_name).strip(),
                "codeList": str(code_list).strip() if code_list else "",
            })

    if current_cdt:
        cdts.append(current_cdt)

    wb.close()
    print(f"  Extracted {len(cdts)} Core Data Types")
    return cdts


def merge_with_taiw_elements(new_elements):
    """Merge new elements with existing TAIW data, preserving TAIW-specific fields."""
    taiw_path = os.path.join(TAIW_DIR, "dataElements.json")
    with open(taiw_path, "r") as f:
        old_elements = json.load(f)

    # Build lookup of old elements by name (case-insensitive)
    old_by_name = {}
    for e in old_elements:
        old_by_name[e["name"].lower()] = e

    # For each new element, check if TAIW had it and preserve TAIW-specific fields
    taiw_fields = ["unEdifactTag", "pakMapping", "capabilityLinks", "reuseScore", "tier"]

    for elem in new_elements:
        name_lower = elem["name"].lower()
        if name_lower in old_by_name:
            old = old_by_name[name_lower]
            for field in taiw_fields:
                if field in old:
                    elem[field] = old[field]

    return new_elements


def merge_with_taiw_code_lists(new_cls):
    """Merge new code lists with existing TAIW data, preserving pakMapping."""
    taiw_path = os.path.join(TAIW_DIR, "codeLists.json")
    with open(taiw_path, "r") as f:
        old_cls = json.load(f)

    old_by_name = {}
    for cl in old_cls:
        old_by_name[cl["name"].lower()] = cl

    taiw_fields = ["pakMapping", "webocMapping"]

    for cl in new_cls:
        name_lower = cl["name"].lower()
        if name_lower in old_by_name:
            old = old_by_name[name_lower]
            for field in taiw_fields:
                if field in old:
                    cl[field] = old[field]

    return new_cls


def build_domains(classes, elements):
    """Build domain summary from extracted classes and elements."""
    domain_colors = {
        "Declaration": "#6366f1",
        "Consignment": "#8b5cf6",
        "Goods": "#f59e0b",
        "Transport": "#14b8a6",
        "Party": "#3b82f6",
        "Location": "#10b981",
        "Financial": "#ef4444",
        "Document": "#f97316",
        "Classification": "#06b6d4",
        "Risk & Control": "#64748b",
        "Origin": "#22c55e",
        "Warehouse": "#a855f7",
        "E-Commerce": "#ec4899",
        "Response": "#78716c",
        "Other": "#94a3b8",
    }
    domain_icons = {
        "Declaration": "FileText",
        "Consignment": "Package",
        "Goods": "ShoppingCart",
        "Transport": "Truck",
        "Party": "Users",
        "Location": "MapPin",
        "Financial": "DollarSign",
        "Document": "File",
        "Classification": "Tag",
        "Risk & Control": "Shield",
        "Origin": "Globe",
        "Warehouse": "Warehouse",
        "E-Commerce": "ShoppingBag",
        "Response": "MessageSquare",
        "Other": "MoreHorizontal",
    }

    # Count elements and classes per domain
    elem_counts = defaultdict(int)
    class_counts = defaultdict(int)

    for e in elements:
        elem_counts[e["domain"]] += 1
    for c in classes:
        class_counts[c["domain"]] += 1

    # Load existing domains for descriptions
    with open(os.path.join(TAIW_DIR, "domains.json"), "r") as f:
        old_domains = json.load(f)
    old_by_name = {d["name"]: d for d in old_domains}

    domains = []
    for domain_name in sorted(set(list(elem_counts.keys()) + list(class_counts.keys()))):
        old = old_by_name.get(domain_name, {})
        domains.append({
            "name": domain_name,
            "slug": domain_name.lower().replace(" & ", "-").replace(" ", "-"),
            "count": elem_counts.get(domain_name, 0),
            "classCount": class_counts.get(domain_name, 0),
            "description": old.get("description", f"WCO DM elements related to {domain_name.lower()} operations"),
            "color": domain_colors.get(domain_name, "#94a3b8"),
            "icon": domain_icons.get(domain_name, "MoreHorizontal"),
        })

    # Sort by element count descending
    domains.sort(key=lambda d: d["count"], reverse=True)
    return domains


def build_relationships(classes):
    """Build class relationships from class hierarchy."""
    relationships = []
    class_by_name = {c["name"]: c for c in classes}

    for cls in classes:
        # Create composition relationships from UML path structure
        # Classes with superclass markers inherit from their parent
        if cls["superclass"] != "NONE":
            parent_name = {
                "DOC": "Document",
                "GOV": "GovernmentAgency",
                "LOC": "Location",
                "PAR": "Party",
                "TRA": "TransportMeans",
            }.get(cls["superclass"])
            if parent_name and parent_name in class_by_name:
                relationships.append({
                    "parent": parent_name,
                    "child": cls["name"],
                    "type": "generalization",
                    "cardinality": "1:1",
                    "parentDomain": class_by_name[parent_name]["domain"],
                    "childDomain": cls["domain"],
                })

    # Add known composition relationships
    compositions = [
        ("Declaration", "GoodsShipment", "1:N"),
        ("Declaration", "Agent", "0:N"),
        ("Declaration", "Amendment", "0:N"),
        ("Declaration", "AdditionalDocument", "0:N"),
        ("Declaration", "AdditionalInformation", "0:N"),
        ("Declaration", "Submitter", "0:1"),
        ("GoodsShipment", "Consignment", "1:N"),
        ("GoodsShipment", "GovernmentAgencyGoods", "0:N"),
        ("GoodsShipment", "CustomsValuation", "0:1"),
        ("GoodsShipment", "Destination", "0:1"),
        ("GoodsShipment", "ExportCountry", "0:1"),
        ("Consignment", "GoodsItem", "0:N"),
        ("Consignment", "TransportEquipment", "0:N"),
        ("Consignment", "Consignee", "0:1"),
        ("Consignment", "Consignor", "0:1"),
        ("GoodsItem", "Commodity", "0:1"),
        ("GoodsItem", "Packaging", "0:N"),
        ("GoodsItem", "DutyTaxFee", "0:N"),
        ("GoodsItem", "Classification", "0:N"),
        ("Commodity", "DangerousGoods", "0:N"),
        ("TransportEquipment", "Seal", "0:N"),
    ]

    for parent, child, card in compositions:
        if parent in class_by_name and child in class_by_name:
            relationships.append({
                "parent": parent,
                "child": child,
                "type": "composition",
                "cardinality": card,
                "parentDomain": class_by_name[parent]["domain"],
                "childDomain": class_by_name[child]["domain"],
            })

    print(f"  Built {len(relationships)} relationships")
    return relationships


def save_json(data, filename):
    """Save JSON file to TAIW directory."""
    path = os.path.join(TAIW_DIR, filename)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"  Saved {filename} ({len(data)} records)")


def update_index(classes, elements, code_lists, domains):
    """Update index.json with new counts."""
    idx_path = os.path.join(TAIW_DIR, "index.json")
    with open(idx_path, "r") as f:
        index = json.load(f)

    index["dataModel"] = "WCO Data Model v4.2.0 (Official Distribution)"
    index["stats"]["dataElements"] = len(elements)
    index["stats"]["classes"] = len(classes)
    index["stats"]["domains"] = len(domains)
    index["stats"]["codeLists"] = len(code_lists)

    with open(idx_path, "w") as f:
        json.dump(index, f, indent=2, ensure_ascii=False)
    print(f"  Updated index.json")


def main():
    print("=" * 60)
    print("WCO DM v4.2.0 Extraction & TAIW Enrichment")
    print("=" * 60)

    # Step 1: Extract classes
    classes = extract_classes()

    # Step 2: Extract data elements
    elements = extract_all_elements()

    # Step 3: Extract code lists
    code_lists = extract_code_lists()

    # Step 4: Extract Core Data Types
    cdts = extract_core_data_types()

    # Step 5: Merge with TAIW-specific fields
    print("\nMerging with existing TAIW data...")
    elements = merge_with_taiw_elements(elements)
    code_lists = merge_with_taiw_code_lists(code_lists)

    # Step 6: Build domains from actual data
    domains = build_domains(classes, elements)

    # Step 7: Build relationships
    relationships = build_relationships(classes)

    # Step 8: Save all files
    print("\nSaving enriched data...")
    save_json(classes, "classes.json")
    save_json(elements, "dataElements.json")
    save_json(code_lists, "codeLists.json")
    save_json(cdts, "coreDataTypes.json")
    save_json(domains, "domains.json")
    save_json(relationships, "relationships.json")

    # Step 9: Update index
    update_index(classes, elements, code_lists, domains)

    # Summary
    print("\n" + "=" * 60)
    print("ENRICHMENT COMPLETE")
    print("=" * 60)
    print(f"  Classes:      {len(classes)} (was 103)")
    print(f"  Elements:     {len(elements)} (was 727)")
    print(f"  Code Lists:   {len(code_lists)} (was 50)")
    print(f"  CDTs:         {len(cdts)} (new)")
    print(f"  Domains:      {len(domains)} (was 14)")
    print(f"  Relationships:{len(relationships)}")
    print("\nPreserved TAIW-specific files:")
    for f in ["capabilities.json", "tacrQuestions.json", "pakistanContext.json",
              "starSchema.json", "gapExtensions.json", "enrichment.json",
              "mappings.json", "dataRequirements.json", "reuseScores.json",
              "dependencies.json", "informationPackages.json"]:
        print(f"  ✓ {f}")


if __name__ == "__main__":
    main()

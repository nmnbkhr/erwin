#!/usr/bin/env python3
"""
prepare_data.py — Convert nmnbkhr/erwin repo outputs to BAIW app JSON

Usage:
  python scripts/prepare_data.py --repo /path/to/erwin --output src/data/
  python scripts/prepare_data.py --repo /path/to/erwin --output src/data/ --validate
  python scripts/prepare_data.py --repo /path/to/erwin --output src/data/ --dry-run

Requires: pip install openpyxl
"""

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from collections import defaultdict

# ---------------------------------------------------------------------------
# Domain name normalisation: source → app display name
# ---------------------------------------------------------------------------
DOMAIN_MAP = {
    # Direct FSDM domain names (from fsdm_entity_catalog.csv / fsdm_domain_map.json)
    'Party Management': 'Party Management',
    'Agreement/Account': 'Agreement & Account',
    'Campaign/Marketing': 'Campaign & Marketing',
    'Product Management': 'Product Management',
    'Investment Management': 'Investment Management',
    'Channel Management': 'Channel Management',
    'Claims Management': 'Claims Management',
    'Financial Transaction': 'Financial Transaction',
    'Risk Management': 'Risk Management',
    'Event Management': 'Event Management',
    'Web Analytics': 'Web Analytics',
    'Asset Management': 'Asset Management',
    'Reference Data': 'Reference Data',
    'Human Resources': 'Human Resources',
    'Location/Geography': 'Location & Geography',
    'Other': 'Other',
    # Aliases
    'Party': 'Party Management',
    'Agreement': 'Agreement & Account',
    'Agreement & Account': 'Agreement & Account',
    'Classification': 'Reference Data',
    'Classification & Reference': 'Reference Data',
    'Financial Instrument': 'Investment Management',
    'Transaction': 'Financial Transaction',
    'Location & Geography': 'Location & Geography',
    'Market Data': 'Asset Management',
    'General Ledger': 'Other',
    'Channel': 'Channel Management',
    'Event & Calendar': 'Event Management',
    'Campaign & Marketing': 'Campaign & Marketing',
    'Document & Content': 'Other',
}

DOMAIN_COLORS = {
    'Party Management': '#4E79A7',
    'Agreement & Account': '#F28E2B',
    'Campaign & Marketing': '#499894',
    'Product Management': '#76B7B2',
    'Investment Management': '#59A14F',
    'Channel Management': '#FABFD2',
    'Claims Management': '#D37295',
    'Financial Transaction': '#EDC948',
    'Risk Management': '#BAB0AC',
    'Event Management': '#B6992D',
    'Web Analytics': '#86BCB6',
    'Asset Management': '#FF9DA7',
    'Reference Data': '#E15759',
    'Human Resources': '#9C755F',
    'Location & Geography': '#B07AA1',
    'Other': '#A0CBE8',
}

THEME_COLORS = {
    'Marketing and Customer Experience': 'blue',
    'Marketing & Customer Experience': 'blue',
    'Risk & Compliance': 'red',
    'Risk and Compliance': 'red',
    'Finance & Performance Management': 'amber',
    'Finance and Performance Management': 'amber',
    'Product Management': 'green',
}


def parse_args():
    parser = argparse.ArgumentParser(description='Convert erwin repo data to BAIW JSON')
    parser.add_argument('--repo', required=True, help='Path to erwin repo root')
    parser.add_argument('--output', default='src/data/', help='Output directory for JSON files')
    parser.add_argument('--validate', action='store_true', help='Validate counts after conversion')
    parser.add_argument('--dry-run', action='store_true', help='Parse and validate without writing')
    return parser.parse_args()


def read_csv(filepath):
    """Read CSV, auto-detect encoding."""
    for enc in ('utf-8', 'utf-8-sig', 'latin-1'):
        try:
            with open(filepath, 'r', encoding=enc) as f:
                reader = csv.DictReader(f)
                rows = list(reader)
                return rows, list(reader.fieldnames or [])
        except (UnicodeDecodeError, Exception):
            continue
    print(f"  WARNING: Could not read {filepath}")
    return [], []


def read_json(filepath):
    """Read JSON file."""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def slugify(text):
    """Convert 'Profitability Modelling' → 'profitability_modelling'"""
    return re.sub(r'[^a-z0-9]+', '_', text.lower()).strip('_')


def domain_slug(name):
    """Convert domain name to URL slug."""
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')


def norm_domain(raw):
    """Normalise domain name to app display name."""
    if not raw:
        return 'Other'
    raw = raw.strip()
    return DOMAIN_MAP.get(raw, 'Other')


def write_json(filepath, data, dry_run=False):
    """Write JSON data to file."""
    if dry_run:
        return
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ============================
# 1. Entities
# ============================
def convert_entities(repo, entity_domain_map=None):
    path = repo / 'fsdm_output' / 'fsdm_entity_catalog.csv'
    if not path.exists():
        print(f"  WARNING: {path} not found, skipping entities")
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} entity rows, columns: {headers}")

    entities = []
    for i, row in enumerate(rows, start=1):
        name = row.get('Entity', '').strip()
        if not name:
            continue
        domain_raw = row.get('Domain', 'Other').strip()
        domain = norm_domain(domain_raw)
        if entity_domain_map and name in entity_domain_map:
            domain = entity_domain_map[name]
        desc = row.get('Description', '').strip()
        col_count = int(row.get('Column_Count', 0) or 0)
        entities.append({
            'id': i,
            'name': name,
            'domain': domain,
            'description': desc or f"{name} entity within the {domain} subject area.",
            'attributeCount': col_count,
        })
    entities.sort(key=lambda e: (e['domain'], e['name']))
    for i, e in enumerate(entities, start=1):
        e['id'] = i
    return entities


# ============================
# 2. Relationships
# ============================
def convert_relationships(repo, entity_id_map):
    path = repo / 'fsdm_output' / 'fsdm_relationships.csv'
    if not path.exists():
        print(f"  WARNING: {path} not found")
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} relationship rows, columns: {headers}")

    relationships = []
    for row in rows:
        parent = row.get('Parent_Entity', '').strip()
        child = row.get('Child_Entity', '').strip()
        rel_type = row.get('Relationship_Type', 'references').strip().lower()
        if not parent or not child:
            continue
        relationships.append({
            'parentId': entity_id_map.get(parent, 0),
            'parentName': parent,
            'childId': entity_id_map.get(child, 0),
            'childName': child,
            'type': rel_type or 'references',
        })
    return relationships


# ============================
# 3. Attributes
# ============================
def convert_attributes(repo, entity_id_map):
    path = repo / 'fsdm_output' / 'fsdm_data_dictionary.csv'
    if not path.exists():
        print(f"  WARNING: {path} not found")
        return [], {}
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} attribute rows, columns: {headers}")

    attributes = []
    attr_counts = defaultdict(int)
    for row in rows:
        entity_name = row.get('Entity', '').strip()
        attr_name = row.get('Attribute', '').strip()
        if not entity_name or not attr_name:
            continue
        data_type = row.get('Teradata_Type', row.get('Classword_Type', 'VARCHAR')).strip()
        required = row.get('Required', 'N').strip().upper()
        nullable = required != 'Y'
        attr_counts[entity_name] += 1
        attributes.append({
            'entityId': entity_id_map.get(entity_name, 0),
            'entityName': entity_name,
            'name': attr_name,
            'dataType': data_type,
            'nullable': nullable,
        })
    return attributes, dict(attr_counts)


# ============================
# 4. Domains
# ============================
def convert_domains(repo, entities):
    domain_counts = defaultdict(int)
    for e in entities:
        domain_counts[e['domain']] += 1

    DEFAULT_DESCS = {
        'Party Management': 'Entities representing parties, individuals, organizations, and their roles, relationships, and contact information.',
        'Agreement & Account': 'Entities covering agreements, accounts, terms, conditions, and contractual arrangements between parties.',
        'Campaign & Marketing': 'Entities for marketing campaigns, offers, promotions, and customer engagement programs.',
        'Product Management': 'Entities describing banking products, services, features, pricing, and product lifecycle management.',
        'Investment Management': 'Entities representing securities, derivatives, fixed income instruments, and other financial assets.',
        'Channel Management': 'Entities representing distribution channels including branch, digital, ATM, and contact center.',
        'Claims Management': 'Entities for insurance claims, claim processing, settlements, and claims lifecycle management.',
        'Financial Transaction': 'Entities covering financial transactions, payments, transfers, and settlement processing.',
        'Risk Management': 'Entities for credit risk, market risk, operational risk, and risk assessment frameworks.',
        'Event Management': 'Entities for business events, calendars, schedules, holidays, and temporal reference data.',
        'Web Analytics': 'Entities for digital analytics, web tracking, online behavior, and digital channel metrics.',
        'Asset Management': 'Entities for market prices, rates, indices, benchmarks, and asset portfolio data.',
        'Reference Data': 'Reference data entities including codes, classifications, taxonomies, and standardized lookup values.',
        'Human Resources': 'Entities for employee records, organizational structure, roles, and HR management.',
        'Location & Geography': 'Entities for addresses, regions, countries, branches, and geographic hierarchies.',
        'Other': 'Cross-cutting and miscellaneous entities spanning multiple subject areas.',
    }

    domains = []
    for i, (name, count) in enumerate(sorted(domain_counts.items(), key=lambda x: -x[1]), start=1):
        domains.append({
            'id': i,
            'name': name,
            'description': DEFAULT_DESCS.get(name, f'Entities in the {name} domain.'),
            'entityCount': count,
            'color': DOMAIN_COLORS.get(name, '#A0CBE8'),
        })
    return domains


# ============================
# 5. Inheritance Tree
# ============================
def convert_inheritance(repo):
    path = repo / 'fsdm_output' / 'fsdm_inheritance_tree.json'
    if not path.exists():
        print(f"  WARNING: {path} not found")
        return []

    tree = read_json(path)
    entries = []
    entry_id = [0]

    def walk(node, parent_name=None, depth=0):
        name = node.get('name', '')
        if not name:
            return
        if parent_name:
            entry_id[0] += 1
            entries.append({
                'parentId': entry_id[0],
                'parentName': parent_name,
                'childId': entry_id[0] + 1,
                'childName': name,
                'depth': depth,
            })
        for child in node.get('children', []):
            walk(child, name, depth + 1)

    if isinstance(tree, list):
        for node in tree:
            walk(node)
    elif isinstance(tree, dict):
        walk(tree)

    return entries


# ============================
# 6. Capabilities
# ============================
def convert_capabilities(repo):
    path = repo / 'bvf_fsdm_output' / 'bvf_capability_summary.csv'
    if not path.exists():
        print(f"  WARNING: {path} not found")
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} capability rows, columns: {headers}")

    capabilities = []
    theme_ids = {}
    group_ids = {}
    for i, row in enumerate(rows, start=1):
        theme = row.get('Theme', '').strip()
        group = row.get('Capability_Group', '').strip()
        sub = row.get('Sub_Capability', '').strip()
        data_req_count = int(row.get('Data_Requirements_Count', 0) or 0)
        if not sub:
            continue
        if theme not in theme_ids:
            theme_ids[theme] = slugify(theme)
        if group not in group_ids:
            group_ids[group] = slugify(group)

        theme_color = 'slate'
        for pattern, color in THEME_COLORS.items():
            if pattern.lower() in theme.lower():
                theme_color = color
                break

        capabilities.append({
            'id': i,
            'name': sub,
            'themeId': theme_ids[theme],
            'themeName': theme,
            'groupId': group_ids[group],
            'groupName': group,
            'description': f"{sub} capability within {group} ({theme}).",
            'dataReqCount': data_req_count,
            'phase': ((i - 1) % 4) + 1,
        })
    return capabilities


# ============================
# 7. Data Requirements
# ============================
def convert_data_requirements(repo, cap_name_to_id):
    path = repo / 'bvf_fsdm_output' / 'bvf_data_requirements.csv'
    if not path.exists():
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} data requirement rows, columns: {headers}")

    reqs = []
    for i, row in enumerate(rows, start=1):
        name = row.get('Data_Requirement', '').strip()
        subject_area = row.get('FSDM_Subject_Area', '').strip()
        cap_count = int(row.get('Capability_Count', 0) or 0)
        caps_str = row.get('Capabilities_Using', '')
        cap_ids = []
        if caps_str:
            for cn in caps_str.split(';'):
                cn = cn.strip()
                if cn in cap_name_to_id:
                    cap_ids.append(cap_name_to_id[cn])
        primary_cap = cap_ids[0] if cap_ids else i
        reqs.append({
            'id': slugify(name),
            'capabilityId': primary_cap,
            'description': name,
            'fsdmSubjectArea': norm_domain(subject_area),
            'priority': 'HIGH' if cap_count > 40 else 'MEDIUM' if cap_count > 20 else 'LOW',
        })
    return reqs


# ============================
# 8. Mappings
# ============================
def convert_mappings(repo):
    path = repo / 'bvf_fsdm_output' / 'bvf_to_fsdm_entity_mapping.csv'
    if not path.exists():
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} mapping rows")
    return [{
        'requirement': row.get('Data_Requirement', '').strip(),
        'entity': row.get('FSDM_Entity', '').strip(),
        'domain': norm_domain(row.get('FSDM_Domain', row.get('FSDM_Subject_Area', '')).strip()),
        'confidence': row.get('Mapping_Confidence', 'Medium').strip(),
        'notes': row.get('Notes', '').strip(),
    } for row in rows]


# ============================
# 9. Dependencies
# ============================
def convert_dependencies(repo, entity_id_map, cap_name_to_id):
    path = repo / 'bvf_fsdm_output' / 'capability_fsdm_dependencies.csv'
    if not path.exists():
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} dependency rows, columns: {headers}")

    dependencies = []
    for row in rows:
        sub_cap = row.get('Sub_Capability', '').strip()
        cap_id = cap_name_to_id.get(sub_cap, 0)
        entities_str = row.get('FSDM_Entities', '').strip()
        subject_area = row.get('FSDM_Subject_Area', '').strip()
        if not entities_str:
            continue
        for ent_name in entities_str.split(';'):
            ent_name = ent_name.strip()
            if not ent_name:
                continue
            dependencies.append({
                'capabilityId': cap_id,
                'capabilityName': sub_cap,
                'entityId': entity_id_map.get(ent_name, 0),
                'entityName': ent_name,
                'domain': norm_domain(subject_area),
            })
    return dependencies


# ============================
# 10. Reuse Scores
# ============================
def convert_reuse_scores(repo, entity_id_map):
    path = repo / 'bvf_fsdm_output' / 'fsdm_entity_reuse_scores.csv'
    if not path.exists():
        return []
    rows, headers = read_csv(path)
    print(f"  Read {len(rows)} reuse score rows")

    scores = []
    for row in rows:
        entity = row.get('FSDM_Entity', '').strip()
        domain = norm_domain(row.get('Subject_Area', '').strip())
        caps = int(row.get('Capabilities_Supported', 0) or 0)
        tier_raw = row.get('Priority_Tier', 'P4').strip()
        tier = tier_raw.split('-')[0].strip() if '-' in tier_raw else tier_raw
        scores.append({
            'entityId': entity_id_map.get(entity, 0),
            'entityName': entity,
            'domain': domain,
            'reuseCount': caps,
            'tier': tier,
        })
    scores.sort(key=lambda s: -s['reuseCount'])
    return scores


# ============================
# 10b. Reuse Matrix
# ============================
def convert_reuse_matrix(repo, cap_name_to_id):
    path = repo / 'bvf_fsdm_output' / 'bvf_reuse_matrix.csv'
    if not path.exists():
        return []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        header = next(reader)
        cap_names = [h.strip() for h in header[3:] if h.strip() and not h.startswith('%')]

    rows, _ = read_csv(path)
    pairs = []
    for row in rows:
        sub_cap = row.get('Sub_Capability', '').strip()
        cap1_id = cap_name_to_id.get(sub_cap, 0)
        if not cap1_id:
            continue
        for cap2_name in cap_names:
            val = row.get(cap2_name, '').strip()
            if not val:
                continue
            try:
                sim = float(val)
            except ValueError:
                continue
            if sim <= 0 or sim >= 1.0:
                continue
            cap2_id = cap_name_to_id.get(cap2_name.strip(), 0)
            if cap2_id and cap1_id < cap2_id:
                pairs.append({
                    'cap1Id': cap1_id,
                    'cap2Id': cap2_id,
                    'similarity': round(sim, 3),
                })
    pairs.sort(key=lambda p: -p['similarity'])
    return pairs[:500]


# ============================
# 11. Data Lineage
# ============================
def convert_lineage(repo):
    path = repo / 'bvf_fsdm_output' / 'data_lineage.json'
    if not path.exists():
        return []
    data = read_json(path)
    entries = []

    # Handle various possible structures in the source JSON
    for key in data:
        if key == 'metadata':
            continue
        section = data[key]
        if isinstance(section, list):
            for item in section:
                if isinstance(item, dict):
                    entries.append({
                        'id': str(len(entries) + 1),
                        'sourceEntity': str(item.get('source_entities', item.get('source_entity', item.get('fsdm_entities', '')))),
                        'sourceColumn': item.get('source_column', item.get('source_attribute', '')),
                        'targetEntity': item.get('target_table', item.get('target_entity', '')),
                        'targetColumn': item.get('target_column', item.get('column', '')),
                        'transformation': item.get('calculation', item.get('transformation', '')),
                    })
    return entries


# ============================
# 12. Star Schema (SQL)
# ============================
def parse_sql_tables(filepath):
    if not filepath.exists():
        return []
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    tables = []
    # Handle Teradata syntax: CREATE MULTISET TABLE name, FALLBACK, ...
    parts = re.split(r'(?=CREATE\s+(?:MULTISET\s+|SET\s+)?(?:TABLE|VIEW))', content, flags=re.IGNORECASE)

    for part in parts:
        match = re.search(r'CREATE\s+(?:MULTISET\s+|SET\s+)?(?:TABLE|VIEW)\s+(\w+)', part, re.IGNORECASE)
        if not match:
            continue
        table_name = match.group(1).strip()
        is_view = bool(re.search(r'CREATE\s+(?:MULTISET\s+|SET\s+)?VIEW', part, re.IGNORECASE))

        if table_name.startswith('FACT_') or table_name.startswith('FCT_'):
            ttype = 'fact'
        elif table_name.startswith('DIM_'):
            ttype = 'dimension'
        elif table_name.startswith('AGG_'):
            ttype = 'aggregate'
        elif table_name.startswith('VW_') or is_view:
            ttype = 'view'
        else:
            ttype = 'table'

        columns = []
        if not is_view:
            paren_start = part.find('(')
            if paren_start > -1:
                depth = 0
                paren_end = paren_start
                for j in range(paren_start, len(part)):
                    if part[j] == '(':
                        depth += 1
                    elif part[j] == ')':
                        depth -= 1
                        if depth == 0:
                            paren_end = j
                            break
                col_block = part[paren_start + 1:paren_end]
                # Strip inline comments before splitting by comma
                col_block = re.sub(r'--[^\n]*', '', col_block)
                for line in col_block.split(','):
                    line = line.strip()
                    if not line or line.upper().startswith(('PRIMARY', 'FOREIGN', 'CONSTRAINT', 'INDEX', 'UNIQUE', 'CHECK')):
                        continue
                    col_match = re.match(r'(\w+)\s+(\w[\w().,\s]*?)(?:\s+(NOT\s+NULL|NULL|DEFAULT|PRIMARY|COMMENT).*)?$', line, re.IGNORECASE)
                    if col_match:
                        col_name = col_match.group(1)
                        col_type = col_match.group(2).strip().rstrip(',')
                        is_pk_specific = any(kw in col_name.lower() for kw in ['pakistan', 'sbp', 'pkr', 'islamic', 'zakat', 'kibor'])
                        col_data = {'name': col_name, 'dataType': col_type}
                        if is_pk_specific:
                            col_data['isPakistanSpecific'] = True
                        columns.append(col_data)

        tables.append({'name': table_name, 'type': ttype, 'columns': columns})
    return tables


def convert_star_schema(repo):
    tables = parse_sql_tables(repo / 'bvf_fsdm_output' / 'profitability_star_schema_enhanced.sql')
    fact = None
    dims, aggs, views = [], [], []
    for t in tables:
        if t['type'] == 'fact' and fact is None:
            fact = t
        elif t['type'] == 'dimension':
            dims.append(t)
        elif t['type'] == 'aggregate':
            aggs.append(t)
        elif t['type'] == 'view':
            views.append(t)
    return {
        'factTable': fact or {'name': 'FACT_CUSTOMER_PROFITABILITY', 'type': 'fact', 'columns': []},
        'dimensions': dims, 'aggregates': aggs, 'views': views,
    }


# ============================
# 13. Gap Extensions (SQL)
# ============================
def convert_gap_extensions(repo):
    tables = parse_sql_tables(repo / 'bvf_fsdm_output' / 'fsdm_gap_extensions.sql')
    module_patterns = {
        'abc': (['COST_', 'ACTIVITY', 'ABC_'], 'Activity Based Costing'),
        'clv': (['CLV_', 'CUSTOMER_LIFETIME', 'CUSTOMER_VALUE'], 'Customer Lifetime Value'),
        'budget': (['BUDGET', 'FORECAST', 'KPI_TARGET'], 'Budgets & Forecasts'),
        'bpm': (['BUSINESS_PROCESS', 'PROCESS_'], 'Business Process Management'),
        'ops': (['OPERATIONAL_METRIC', 'CHANNEL_OPERATIONAL', 'BRANCH_OPERATIONAL'], 'Operational Metrics'),
    }
    modules = {}
    for t in tables:
        assigned = False
        for mod_id, (prefixes, mod_name) in module_patterns.items():
            if any(t['name'].upper().startswith(p) for p in prefixes):
                if mod_id not in modules:
                    modules[mod_id] = {'id': mod_id, 'name': mod_name, 'description': f'{mod_name} gap extension tables.', 'tables': []}
                modules[mod_id]['tables'].append({'name': t['name'], 'columns': t['columns']})
                assigned = True
                break
        if not assigned:
            if 'other_ext' not in modules:
                modules['other_ext'] = {'id': 'other_ext', 'name': 'Other Extensions', 'description': 'Additional gap extension tables.', 'tables': []}
            modules['other_ext']['tables'].append({'name': t['name'], 'columns': t['columns']})
    return list(modules.values())


# ============================
# 14. BACR Questions (Excel)
# ============================
def convert_bacr_questions(repo):
    paths = [
        repo / 'BACR - INTERVIEW MASTER - DA004462.xlsm',
        repo / 'BACR_-_INTERVIEW_MASTER_-_DA004462.xlsm',
        repo / 'bacr_output' / 'bacr_assessment_template.xlsx',
    ]
    path = None
    for p in paths:
        if p.exists():
            path = p
            break
    if not path:
        print(f"  WARNING: BACR Excel not found")
        return []

    try:
        import openpyxl
    except ImportError:
        print(f"  WARNING: openpyxl not installed. Run: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target_sheet = None
    for name in ['All Questions', 'Questions', 'Assessment', 'Sheet1']:
        if name in wb.sheetnames:
            target_sheet = wb[name]
            break
    if not target_sheet:
        target_sheet = wb[wb.sheetnames[0]]
    print(f"  Reading BACR from sheet: {target_sheet.title}")

    questions = []
    current_category = ''
    current_subcategory = ''
    valid_categories = {'Business', 'Culture', 'Governance', 'Information', 'Applications', 'Systems', 'Agility', 'Outcomes', 'Data'}
    q_id = 0

    for row in target_sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        cat = str(row[0] or '').strip()
        section = str(row[1] or '').strip()
        text = str(row[2] or '').strip()

        if cat and cat in valid_categories:
            current_category = cat
        if section:
            current_subcategory = section
        if not text or text == 'None' or len(text) < 10:
            continue
        if not current_category:
            continue

        q_id += 1
        questions.append({
            'id': f"{slugify(current_category)}_{slugify(current_subcategory)}_{q_id:03d}",
            'category': current_category,
            'subcategory': current_subcategory,
            'text': text,
            'weight': 1,
        })

    wb.close()
    return questions


# ============================
# 16. Pakistan Context
# ============================
def convert_pakistan_context(repo):
    path = repo / 'bvf_fsdm_output' / 'pakistan_banking_context.md'
    if not path.exists():
        return None
    with open(path, 'r', encoding='utf-8') as f:
        content = f.read()
    print(f"  Read pakistan_banking_context.md ({len(content)} chars)")
    print(f"  NOTE: pakistanContext.json is manually curated. Not overwriting.")
    return None


# ============================
# Main
# ============================
def main():
    args = parse_args()
    repo = Path(args.repo)
    output = Path(args.output)

    if not repo.exists():
        print(f"ERROR: Repo path {repo} does not exist")
        sys.exit(1)

    output.mkdir(parents=True, exist_ok=True)
    report = {'files_converted': 0, 'warnings': [], 'counts': {}}

    print('=' * 60)
    print('BAIW Data Conversion')
    print('=' * 60)
    print(f'Source: {repo}')
    print(f'Output: {output}')
    print()

    # Build entity→domain map from fsdm_domain_map.json
    entity_domain_map = {}
    dm_path = repo / 'fsdm_output' / 'fsdm_domain_map.json'
    if dm_path.exists():
        dm = read_json(dm_path)
        if 'domains' in dm:
            for raw_name, ent_list in dm['domains'].items():
                norm = norm_domain(raw_name)
                for ent in ent_list:
                    entity_domain_map[ent.strip()] = norm

    # 1. Entities
    print('1. Converting entities...')
    entities = convert_entities(repo, entity_domain_map)
    entity_id_map = {e['name']: e['id'] for e in entities}
    report['counts']['entities'] = len(entities)
    print(f'   -> {len(entities)} entities')

    # 3. Attributes (before rels to update attributeCount)
    print('3. Converting attributes...')
    attributes, attr_counts = convert_attributes(repo, entity_id_map)
    report['counts']['attributes'] = len(attributes)
    print(f'   -> {len(attributes)} attributes')
    for e in entities:
        if e['name'] in attr_counts:
            e['attributeCount'] = attr_counts[e['name']]

    if not args.dry_run:
        write_json(output / 'entities.json', entities)
        write_json(output / 'attributes.json', attributes)
        report['files_converted'] += 2

    # 2. Relationships
    print('2. Converting relationships...')
    relationships = convert_relationships(repo, entity_id_map)
    report['counts']['relationships'] = len(relationships)
    print(f'   -> {len(relationships)} relationships')
    if not args.dry_run:
        write_json(output / 'relationships.json', relationships)
        report['files_converted'] += 1

    # 4. Domains
    print('4. Building domains...')
    domains = convert_domains(repo, entities)
    report['counts']['domains'] = len(domains)
    print(f'   -> {len(domains)} domains')
    if not args.dry_run:
        write_json(output / 'domains.json', domains)
        report['files_converted'] += 1

    # 5. Inheritance
    print('5. Converting inheritance tree...')
    inheritance = convert_inheritance(repo)
    report['counts']['inheritance'] = len(inheritance)
    print(f'   -> {len(inheritance)} entries')
    if not args.dry_run:
        write_json(output / 'inheritanceTree.json', inheritance)
        report['files_converted'] += 1

    # 6. Capabilities
    print('6. Converting capabilities...')
    capabilities = convert_capabilities(repo)
    cap_name_to_id = {c['name']: c['id'] for c in capabilities}
    report['counts']['capabilities'] = len(capabilities)
    print(f'   -> {len(capabilities)} capabilities')
    if not args.dry_run:
        write_json(output / 'capabilities.json', capabilities)
        report['files_converted'] += 1

    # 7. Data Requirements
    print('7. Converting data requirements...')
    data_reqs = convert_data_requirements(repo, cap_name_to_id)
    report['counts']['dataRequirements'] = len(data_reqs)
    print(f'   -> {len(data_reqs)} data requirements')
    if not args.dry_run:
        write_json(output / 'dataRequirements.json', data_reqs)
        report['files_converted'] += 1

    # 8. Mappings
    print('8. Converting BVF->FSDM mappings...')
    mappings = convert_mappings(repo)
    report['counts']['mappings'] = len(mappings)
    print(f'   -> {len(mappings)} mappings')
    if not args.dry_run:
        write_json(output / 'mappings.json', mappings)
        report['files_converted'] += 1

    # 9. Dependencies
    print('9. Converting dependencies...')
    deps = convert_dependencies(repo, entity_id_map, cap_name_to_id)
    report['counts']['dependencies'] = len(deps)
    print(f'   -> {len(deps)} dependencies')
    if not args.dry_run:
        write_json(output / 'dependencies.json', deps)
        report['files_converted'] += 1

    # 10. Reuse Scores
    print('10. Converting reuse scores...')
    reuse_scores = convert_reuse_scores(repo, entity_id_map)
    report['counts']['reuseScores'] = len(reuse_scores)
    print(f'   -> {len(reuse_scores)} scores')
    if not args.dry_run:
        write_json(output / 'reuseScores.json', reuse_scores)
        report['files_converted'] += 1

    # 10b. Reuse Matrix
    print('10b. Converting reuse matrix...')
    reuse_matrix = convert_reuse_matrix(repo, cap_name_to_id)
    report['counts']['reuseMatrix'] = len(reuse_matrix)
    print(f'   -> {len(reuse_matrix)} pairs')
    if not args.dry_run:
        write_json(output / 'reuseMatrix.json', reuse_matrix)
        report['files_converted'] += 1

    # 11. Lineage
    print('11. Converting data lineage...')
    lineage = convert_lineage(repo)
    report['counts']['lineage'] = len(lineage)
    print(f'   -> {len(lineage)} entries')
    if not args.dry_run:
        write_json(output / 'lineage.json', lineage)
        report['files_converted'] += 1

    # 12. Star Schema
    print('12. Parsing star schema SQL...')
    star_schema = convert_star_schema(repo)
    total_tables = 1 + len(star_schema['dimensions']) + len(star_schema['aggregates']) + len(star_schema['views'])
    report['counts']['starSchema'] = total_tables
    print(f'   -> {total_tables} objects')
    if not args.dry_run:
        write_json(output / 'starSchema.json', star_schema)
        report['files_converted'] += 1

    # 13. Gap Extensions
    print('13. Parsing gap extensions SQL...')
    gap_ext = convert_gap_extensions(repo)
    gap_count = sum(len(m['tables']) for m in gap_ext)
    report['counts']['gapExtensions'] = gap_count
    print(f'   -> {gap_count} tables across {len(gap_ext)} modules')
    if not args.dry_run:
        write_json(output / 'gapExtensions.json', gap_ext)
        report['files_converted'] += 1

    # 14. BACR Questions
    print('14. Converting BACR questions...')
    bacr = convert_bacr_questions(repo)
    report['counts']['bacrQuestions'] = len(bacr)
    print(f'   -> {len(bacr)} questions')
    if not args.dry_run and bacr:
        write_json(output / 'bacrQuestions.json', bacr)
        report['files_converted'] += 1

    # 15. Enrichment (validate)
    print('15. Validating enrichment.json...')
    enr_path = output / 'enrichment.json'
    if enr_path.exists():
        enr = read_json(enr_path)
        print(f'   -> EXISTS ({len(enr.get("capabilities", {}))} caps enriched) - not overwritten')
    else:
        print(f'   -> WARNING: not found')
        report['warnings'].append('enrichment.json not found')

    # 16. Pakistan Context
    print('16. Pakistan context...')
    convert_pakistan_context(repo)
    pk_path = output / 'pakistanContext.json'
    if pk_path.exists():
        print(f'   -> EXISTS - not overwritten')
    else:
        report['warnings'].append('pakistanContext.json not found')

    # --- Report ---
    print()
    print('=' * 60)
    print('Conversion Report')
    print('=' * 60)
    expected = {
        'entities': 3917, 'relationships': 5636, 'domains': 16,
        'capabilities': 112, 'dataRequirements': 113, 'mappings': 360,
        'reuseScores': 219, 'lineage': 23,
    }
    for key in ['entities', 'relationships', 'attributes', 'domains', 'inheritance',
                'capabilities', 'dataRequirements', 'mappings', 'dependencies',
                'reuseScores', 'reuseMatrix', 'lineage', 'starSchema', 'gapExtensions', 'bacrQuestions']:
        actual = report['counts'].get(key, 0)
        exp = expected.get(key)
        status = f'(expected ~{exp})' if exp and actual != exp else ''
        print(f'  {key:25s}: {actual:>6,d} {status}')

    if report['warnings']:
        print(f"\nWarnings: {report['warnings']}")

    # Write report + regenerate domain splits
    if not args.dry_run:
        write_json(output / 'conversion_report.json', report)

        print('\nRegenerating domain-split files...')
        for subdir in ['entities', 'attributes', 'relationships']:
            (output / subdir).mkdir(exist_ok=True)

        entity_index = [{'id': e['id'], 'name': e['name'], 'domain': e['domain']} for e in entities]
        write_json(output / 'entityIndex.json', entity_index)

        entity_to_slug = {e['id']: domain_slug(e['domain']) for e in entities}

        ents_by_domain = defaultdict(list)
        for e in entities:
            ents_by_domain[domain_slug(e['domain'])].append(e)
        for slug, ents in ents_by_domain.items():
            write_json(output / 'entities' / f'{slug}.json', ents)

        attrs_by_domain = defaultdict(list)
        for a in attributes:
            slug = entity_to_slug.get(a['entityId'], 'other')
            attrs_by_domain[slug].append(a)
        for slug, attrs in attrs_by_domain.items():
            write_json(output / 'attributes' / f'{slug}.json', attrs)

        rels_by_domain = defaultdict(list)
        for r in relationships:
            slug = entity_to_slug.get(r['parentId'], entity_to_slug.get(r['childId'], 'other'))
            rels_by_domain[slug].append(r)
        for slug, rels in rels_by_domain.items():
            write_json(output / 'relationships' / f'{slug}.json', rels)

        print(f'  Split: {len(ents_by_domain)} domain files each')
        print(f'\nTotal: {report["files_converted"]} files written to {output}')
    else:
        print(f'\nDry run - no files written.')


if __name__ == '__main__':
    main()

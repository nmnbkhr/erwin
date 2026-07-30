#!/usr/bin/env python3
"""Phase 7a: Generate BACR Assessment Template Excel"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = '/mnt/e/erwin/bacr_output'

# Load all data
with open(os.path.join(OUTPUT_DIR, 'bacr_all_questions.json')) as f:
    bacr_questions = json.load(f)
with open(os.path.join(OUTPUT_DIR, 'bacr_to_bvf_mapping.json')) as f:
    bacr_bvf = json.load(f)
with open(os.path.join(OUTPUT_DIR, 'bacr_information_to_fsdm.json')) as f:
    info_fsdm = json.load(f)
with open(os.path.join(OUTPUT_DIR, 'profitability_maturity_profile.json')) as f:
    prof_profile = json.load(f)['profitability_maturity']
with open(os.path.join(OUTPUT_DIR, 'cross_category_dependencies.json')) as f:
    cross_cat = json.load(f)

# Styles
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
PROF_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
GAP_RED = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
GAP_YELLOW = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
GAP_GREEN = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WRAP = Alignment(wrap_text=True, vertical='top')

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
        cell.border = THIN_BORDER

def auto_width(ws, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

wb = Workbook()

# ============================================================
# Sheet 1: Assessment Dashboard
# ============================================================
print("Building Sheet 1: Assessment Dashboard...")
ws1 = wb.active
ws1.title = "Assessment Dashboard"

ws1['A1'] = "BACR Maturity Assessment - United Bank Limited"
ws1['A1'].font = Font(bold=True, size=16, color='2F5496')
ws1['A2'] = "Industry: Financial | Review Type: Data Warehouse Maturity Review"
ws1['A2'].font = Font(italic=True, size=11)
ws1['A3'] = "Assessment Date: TBD | Assessor: TBD"

# Category radar chart data
ws1['A5'] = "Category Maturity Summary"
ws1['A5'].font = Font(bold=True, size=13)

headers = ['Category', 'Questions', 'Financial Qs', 'Current Avg', 'Desired Avg', 'Gap']
for i, h in enumerate(headers):
    ws1.cell(row=6, column=i+1, value=h)
style_header_row(ws1, 6, len(headers))

row = 7
cat_maturity = cross_cat.get('category_maturity_summary', {})
for cat_name in ['Outcomes', 'Information', 'Systems', 'Governance', 'Applications', 'Business', 'Culture', 'Agility']:
    data = cat_maturity.get(cat_name, {})
    current = data.get('avg_current_maturity', 2.0)
    desired = 4.0  # default target
    gap = desired - current

    ws1.cell(row=row, column=1, value=cat_name)
    ws1.cell(row=row, column=2, value=data.get('total_questions', 0))
    fin_count = sum(1 for q in bacr_questions if q['category'] == cat_name and q['is_financial_industry'])
    ws1.cell(row=row, column=3, value=fin_count)
    ws1.cell(row=row, column=4, value=round(current, 1))
    ws1.cell(row=row, column=5, value=round(desired, 1))
    ws1.cell(row=row, column=6, value=round(gap, 1))

    # Color gap
    gap_cell = ws1.cell(row=row, column=6)
    if gap >= 2.5:
        gap_cell.fill = GAP_RED
    elif gap >= 1.5:
        gap_cell.fill = GAP_YELLOW
    else:
        gap_cell.fill = GAP_GREEN

    for col in range(1, 7):
        ws1.cell(row=row, column=col).border = THIN_BORDER
    row += 1

# Profitability readiness
row += 1
ws1.cell(row=row, column=1, value="Profitability Readiness Score").font = Font(bold=True, size=13)
row += 1
ws1.cell(row=row, column=1, value="Overall Current")
ws1.cell(row=row, column=2, value=prof_profile['overall_score'])
row += 1
ws1.cell(row=row, column=1, value="Overall Target")
ws1.cell(row=row, column=2, value=prof_profile['overall_target'])
row += 1
ws1.cell(row=row, column=1, value="Overall Gap")
ws1.cell(row=row, column=2, value=prof_profile['overall_gap'])
ws1.cell(row=row, column=2).fill = GAP_YELLOW

# Component scores
row += 2
ws1.cell(row=row, column=1, value="Profitability Component Scores").font = Font(bold=True, size=13)
row += 1
comp_headers = ['Component', 'Current', 'Target', 'Gap']
for i, h in enumerate(comp_headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, len(comp_headers))
row += 1
for comp_key, comp_data in prof_profile['components'].items():
    ws1.cell(row=row, column=1, value=comp_data['label'])
    ws1.cell(row=row, column=2, value=comp_data['current_avg'])
    ws1.cell(row=row, column=3, value=comp_data['target_avg'])
    ws1.cell(row=row, column=4, value=comp_data['gap'])
    gap_cell = ws1.cell(row=row, column=4)
    if comp_data['gap'] >= 2.5:
        gap_cell.fill = GAP_RED
    elif comp_data['gap'] >= 1.5:
        gap_cell.fill = GAP_YELLOW
    for col in range(1, 5):
        ws1.cell(row=row, column=col).border = THIN_BORDER
    row += 1

# Top 10 gaps
row += 1
ws1.cell(row=row, column=1, value="Top 10 Critical Gaps").font = Font(bold=True, size=13)
row += 1
gap_headers = ['Capability', 'Current', 'Desired', 'Gap', 'Component']
for i, h in enumerate(gap_headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, len(gap_headers))
row += 1
for g in prof_profile['critical_gaps'][:10]:
    ws1.cell(row=row, column=1, value=g['capability'])
    ws1.cell(row=row, column=2, value=g['current'])
    ws1.cell(row=row, column=3, value=g['desired'])
    ws1.cell(row=row, column=4, value=g['gap'])
    ws1.cell(row=row, column=4).fill = GAP_RED if g['gap'] >= 3 else GAP_YELLOW
    ws1.cell(row=row, column=5, value=g['profitability_component'])
    for col in range(1, 6):
        ws1.cell(row=row, column=col).border = THIN_BORDER
    row += 1

auto_width(ws1)

# ============================================================
# Sheet 2: Profitability Questions
# ============================================================
print("Building Sheet 2: Profitability Questions...")
ws2 = wb.create_sheet("Profitability Questions")

prof_questions = [r for r in bacr_bvf if r['is_profitability_critical']]
headers2 = ['ID', 'Question', 'BVF Capability', 'BVF Area',
            'Current State (1-5)', 'Desired State (1-5)', 'Gap', 'Notes']

for i, h in enumerate(headers2):
    ws2.cell(row=1, column=i+1, value=h)
style_header_row(ws2, 1, len(headers2))

# UBL defaults for pre-population
from bacr_phase4_priority import UBL_ESTIMATED_DEFAULTS

row = 2
for r in prof_questions:
    cap = r['bvf_capability']
    defaults = UBL_ESTIMATED_DEFAULTS.get('Outcomes', {}).get(cap, UBL_ESTIMATED_DEFAULTS['Outcomes']['_default'])

    ws2.cell(row=row, column=1, value=r['bacr_question_id'])
    ws2.cell(row=row, column=2, value=r['bacr_question_short'])
    ws2.cell(row=row, column=2).alignment = WRAP
    ws2.cell(row=row, column=3, value=cap)
    ws2.cell(row=row, column=4, value=r['bvf_area'])
    ws2.cell(row=row, column=5, value=defaults['current'])
    ws2.cell(row=row, column=6, value=defaults['desired'])
    gap = defaults['desired'] - defaults['current']
    ws2.cell(row=row, column=7, value=gap)

    gap_cell = ws2.cell(row=row, column=7)
    if gap >= 3:
        gap_cell.fill = GAP_RED
    elif gap >= 2:
        gap_cell.fill = GAP_YELLOW
    else:
        gap_cell.fill = GAP_GREEN

    # Highlight row
    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = THIN_BORDER
    ws2.cell(row=row, column=1).fill = PROF_FILL

    row += 1

auto_width(ws2)
ws2.column_dimensions['B'].width = 60

# ============================================================
# Sheet 3: All Financial Questions
# ============================================================
print("Building Sheet 3: All Financial Questions...")
ws3 = wb.create_sheet("All Financial Questions")

financial_qs = [q for q in bacr_questions if q['is_financial_industry']]
headers3 = ['ID', 'Category', 'Section', 'Question', 'Emerging', 'Developing',
            'Practicing', 'Innovating', 'Leading', 'Current (1-5)', 'Desired (1-5)', 'Notes']

for i, h in enumerate(headers3):
    ws3.cell(row=1, column=i+1, value=h)
style_header_row(ws3, 1, len(headers3))

row = 2
prev_cat = None
for q in sorted(financial_qs, key=lambda x: (x['category'], x['section'], x['id'])):
    # Add category separator
    if q['category'] != prev_cat:
        ws3.cell(row=row, column=1, value=q['category']).font = Font(bold=True, size=12)
        ws3.cell(row=row, column=1).fill = SUBHEADER_FILL
        for col in range(1, 13):
            ws3.cell(row=row, column=col).fill = SUBHEADER_FILL
        row += 1
        prev_cat = q['category']

    ws3.cell(row=row, column=1, value=q['id'])
    ws3.cell(row=row, column=2, value=q['category'])
    ws3.cell(row=row, column=3, value=q['section'])
    ws3.cell(row=row, column=4, value=q['question'][:200])
    ws3.cell(row=row, column=4).alignment = WRAP
    ws3.cell(row=row, column=5, value=q['maturity_descriptors']['1_emerging'][:100] if q['maturity_descriptors']['1_emerging'] else '')
    ws3.cell(row=row, column=6, value=q['maturity_descriptors']['2_developing'][:100] if q['maturity_descriptors']['2_developing'] else '')
    ws3.cell(row=row, column=7, value=q['maturity_descriptors']['3_practicing'][:100] if q['maturity_descriptors']['3_practicing'] else '')
    ws3.cell(row=row, column=8, value=q['maturity_descriptors']['4_innovating'][:100] if q['maturity_descriptors']['4_innovating'] else '')
    ws3.cell(row=row, column=9, value=q['maturity_descriptors']['5_leading'][:100] if q['maturity_descriptors']['5_leading'] else '')

    for col in range(1, 13):
        ws3.cell(row=row, column=col).border = THIN_BORDER

    row += 1

auto_width(ws3)
ws3.column_dimensions['D'].width = 60
for col in 'EFGHI':
    ws3.column_dimensions[col].width = 30

# ============================================================
# Sheet 4: Maturity Mapping (traceability)
# ============================================================
print("Building Sheet 4: Maturity Mapping...")
ws4 = wb.create_sheet("Maturity Mapping")

headers4 = ['BACR Question ID', 'BACR Question', 'BVF Capability', 'BVF Area',
            'Match Score', 'Is Profitability Critical', 'FSDM Impact']

for i, h in enumerate(headers4):
    ws4.cell(row=1, column=i+1, value=h)
style_header_row(ws4, 1, len(headers4))

row = 2
for r in bacr_bvf:
    ws4.cell(row=row, column=1, value=r['bacr_question_id'])
    ws4.cell(row=row, column=2, value=r['bacr_question_short'])
    ws4.cell(row=row, column=3, value=r['bvf_capability'])
    ws4.cell(row=row, column=4, value=r['bvf_area'])
    ws4.cell(row=row, column=5, value=r['match_score'])
    ws4.cell(row=row, column=6, value='Yes' if r['is_profitability_critical'] else 'No')
    ws4.cell(row=row, column=7, value='See FSDM mapping' if r['is_profitability_critical'] else '')

    if r['is_profitability_critical']:
        ws4.cell(row=row, column=6).fill = PROF_FILL

    for col in range(1, 8):
        ws4.cell(row=row, column=col).border = THIN_BORDER
    row += 1

auto_width(ws4)

# ============================================================
# Sheet 5: Implementation Roadmap
# ============================================================
print("Building Sheet 5: Implementation Roadmap...")
ws5 = wb.create_sheet("Implementation Roadmap")

headers5 = ['Phase', 'Phase Name', 'Capabilities', 'Description',
            'Enables', 'FSDM Entities', 'Effort', 'Prerequisites']

for i, h in enumerate(headers5):
    ws5.cell(row=1, column=i+1, value=h)
style_header_row(ws5, 1, len(headers5))

row = 2
for p in prof_profile['implementation_sequence']:
    ws5.cell(row=row, column=1, value=p['phase'])
    ws5.cell(row=row, column=2, value=p['name'])
    ws5.cell(row=row, column=3, value='\n'.join(p['capabilities']))
    ws5.cell(row=row, column=3).alignment = WRAP
    ws5.cell(row=row, column=4, value=p['description'])
    ws5.cell(row=row, column=4).alignment = WRAP
    ws5.cell(row=row, column=5, value=p['enables'])
    ws5.cell(row=row, column=5).alignment = WRAP
    ws5.cell(row=row, column=6, value='\n'.join(p['fsdm_entities']))
    ws5.cell(row=row, column=6).alignment = WRAP
    ws5.cell(row=row, column=7, value=p['estimated_effort'])
    ws5.cell(row=row, column=8, value=', '.join(str(x) for x in p['prerequisite_phases']) if p['prerequisite_phases'] else 'None')

    for col in range(1, 9):
        ws5.cell(row=row, column=col).border = THIN_BORDER
    row += 1

auto_width(ws5)
ws5.column_dimensions['C'].width = 40
ws5.column_dimensions['D'].width = 50
ws5.column_dimensions['E'].width = 50
ws5.column_dimensions['F'].width = 40

# Save
output_path = os.path.join(OUTPUT_DIR, 'bacr_assessment_template.xlsx')
print(f"Saving {output_path}...")
wb.save(output_path)
print(f"Excel template saved with 5 sheets, {row} total rows across sheets")

#!/usr/bin/env python3
"""Phase 1: Parse the BVF XLSM File"""

import json
import os
from openpyxl import load_workbook

OUTPUT_DIR = "/mnt/e/erwin/bvf_output"
XLSM_FILE = "/mnt/e/erwin/Banking Business Value Framework Data Mappings 1.2.xlsm"

os.makedirs(OUTPUT_DIR, exist_ok=True)

def parse_capability_to_data_sheet(wb):
    """Parse 'Banking BVF Capability to Data' sheet.
    Row 3: Subject areas (col 5+), col 4 = 'FSDM Subject Area' header
    Row 4: Data requirement names (col 5+), col 4 = 'Data Requirement' header
    Rows 5-116: Capability rows
      Col 1: BVF Area (forward-fill)
      Col 2: Sub-area (forward-fill)
      Col 4: Capability Name
      Cols 5-117: Binary mapping
      Col 118: Total count
    """
    ws = wb["Banking BVF Capability to Data"]

    # Row 3 (cols 5-117): Subject area per column
    subject_areas = {}
    for col in range(5, 118):
        val = ws.cell(row=3, column=col).value
        if val:
            subject_areas[col] = str(val).strip()

    # Row 4 (cols 5-117): Data requirement names
    data_req_names = {}
    for col in range(5, 118):
        val = ws.cell(row=4, column=col).value
        if val:
            data_req_names[col] = str(val).strip()

    print(f"  Found {len(data_req_names)} data requirements")
    print(f"  Subject areas: {sorted(set(subject_areas.values()))}")

    # Map each data req column to its subject area
    data_req_subject_areas = {}
    for col in sorted(data_req_names.keys()):
        if col in subject_areas:
            data_req_subject_areas[col] = subject_areas[col]

    # Rows 5-116: Capabilities
    capabilities = []
    current_area = None
    current_sub_area = None

    for row in range(5, 117):
        area_val = ws.cell(row=row, column=1).value
        if area_val:
            current_area = str(area_val).strip()
        sub_val = ws.cell(row=row, column=2).value
        if sub_val:
            current_sub_area = str(sub_val).strip()

        cap_name = ws.cell(row=row, column=4).value
        if not cap_name:
            continue
        cap_name = str(cap_name).strip()
        if not cap_name:
            continue

        data_reqs = []
        output_data = []
        for col in sorted(data_req_names.keys()):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None:
                sv = str(cell_val).strip()
                if sv.upper() == "OUTPUT":
                    output_data.append(data_req_names[col])
                elif sv == "1" or cell_val == 1:
                    data_reqs.append(data_req_names[col])

        total_val = ws.cell(row=row, column=118).value

        capabilities.append({
            "name": cap_name,
            "bvf_area": current_area,
            "sub_area": current_sub_area,
            "data_requirements": data_reqs,
            "output_data": output_data,
            "data_req_count": len(data_reqs) + len(output_data),
            "stated_count": int(total_val) if total_val else 0
        })

    print(f"  Parsed {len(capabilities)} capabilities")
    return capabilities, data_req_names, data_req_subject_areas


def parse_data_to_capability_sheet(wb):
    """Parse 'Banking BVF Data to Capability' sheet.
    Row 4: Capability names as column headers (cols 5-116)
    Rows 5+: Data requirement rows
      Col 1: Data Requirement name
      Col 2: FSDM Subject Area
      Col 3: Subject Area Sort order
      Col 4: Line Sort order
      Cols 5-116: Binary mapping or 'OUTPUT'
    """
    ws = wb["Banking BVF Data to Capability"]

    cap_headers = {}
    for col in range(5, 117):
        val = ws.cell(row=4, column=col).value
        if val:
            cap_headers[col] = str(val).strip()

    print(f"  Found {len(cap_headers)} capability columns")

    data_requirements = []
    for row in range(5, 118):
        dr_name = ws.cell(row=row, column=1).value
        if not dr_name:
            continue
        dr_name = str(dr_name).strip()
        if not dr_name:
            continue

        fsdm_sa = ws.cell(row=row, column=2).value
        sort_order = ws.cell(row=row, column=3).value
        line_sort = ws.cell(row=row, column=4).value

        caps_using = []
        output_of = []
        for col in sorted(cap_headers.keys()):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None:
                sv = str(cell_val).strip()
                if sv.upper() == "OUTPUT":
                    output_of.append(cap_headers[col])
                elif sv == "1" or cell_val == 1:
                    caps_using.append(cap_headers[col])

        data_requirements.append({
            "name": dr_name,
            "fsdm_subject_area": str(fsdm_sa).strip() if fsdm_sa else "",
            "sort_order": int(sort_order) if sort_order else 0,
            "line_sort_order": int(line_sort) if line_sort else 0,
            "capability_count": len(caps_using) + len(output_of),
            "capabilities": caps_using,
            "output_of": output_of
        })

    print(f"  Parsed {len(data_requirements)} data requirements")
    return data_requirements


def parse_reuse_matrix(wb):
    """Parse 'Banking BVF Data Reuse Matrix' sheet.
    Row 4 col 4: first capability header, then cols 5+
    Row 5+: matrix rows with capability name in col 4 (or col 1 for area)
    """
    ws = wb["Banking BVF Data Reuse Matrix"]

    # Col headers: row 4, cols 5-116
    col_caps = {}
    for col in range(4, 117):
        val = ws.cell(row=4, column=col).value
        if val:
            col_caps[col] = str(val).strip()

    print(f"  Found {len(col_caps)} column headers in reuse matrix")

    # Row headers: rows 5+, capability name in col 4
    reuse_matrix = {}
    for row in range(5, 125):
        row_cap = ws.cell(row=row, column=4).value
        if not row_cap:
            continue
        row_cap = str(row_cap).strip()
        if not row_cap:
            continue

        reuse_matrix[row_cap] = {}
        for col, col_cap in col_caps.items():
            val = ws.cell(row=row, column=col).value
            if val is not None:
                try:
                    coeff = float(val)
                    if coeff > 0 and row_cap != col_cap:
                        reuse_matrix[row_cap][col_cap] = round(coeff, 4)
                except (ValueError, TypeError):
                    pass

    reuse_matrix = {k: v for k, v in reuse_matrix.items() if v}
    print(f"  Parsed reuse matrix for {len(reuse_matrix)} capabilities")
    return reuse_matrix


def main():
    print("=" * 70)
    print("Phase 1: Parse BVF XLSM File")
    print("=" * 70)

    print(f"\nLoading {XLSM_FILE}...")
    wb = load_workbook(XLSM_FILE, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    print("\n--- Parsing 'Banking BVF Capability to Data' ---")
    capabilities, data_req_names, data_req_sas = parse_capability_to_data_sheet(wb)

    print("\n--- Parsing 'Banking BVF Data to Capability' ---")
    data_requirements = parse_data_to_capability_sheet(wb)

    print("\n--- Parsing 'Banking BVF Data Reuse Matrix' ---")
    reuse_matrix = parse_reuse_matrix(wb)

    wb.close()

    # Summary
    print("\n" + "=" * 70)
    print("BVF PARSING SUMMARY")
    print("=" * 70)

    areas = {}
    for cap in capabilities:
        area = cap["bvf_area"] or "Unknown"
        sub = cap["sub_area"] or "Unknown"
        if area not in areas:
            areas[area] = {}
        if sub not in areas[area]:
            areas[area][sub] = []
        areas[area][sub].append(cap["name"])

    for area, subs in areas.items():
        total = sum(len(v) for v in subs.values())
        print(f"\n  {area} ({total} capabilities)")
        for sub, caps in subs.items():
            print(f"    {sub}: {len(caps)} capabilities")

    sa_counts = {}
    for dr in data_requirements:
        sa = dr["fsdm_subject_area"] or "Unknown"
        sa_counts[sa] = sa_counts.get(sa, 0) + 1
    print(f"\n  Data Requirements by FSDM Subject Area:")
    for sa, count in sorted(sa_counts.items()):
        print(f"    {sa}: {count}")

    total_mappings = sum(cap["data_req_count"] for cap in capabilities)
    avg_dr = total_mappings / len(capabilities) if capabilities else 0
    avg_cap = total_mappings / len(data_requirements) if data_requirements else 0
    print(f"\n  Total capability-data mappings: {total_mappings}")
    print(f"  Avg data reqs per capability: {avg_dr:.1f}")
    print(f"  Avg capabilities per data req: {avg_cap:.1f}")

    # Save outputs
    print("\n--- Saving outputs ---")
    out1 = os.path.join(OUTPUT_DIR, "bvf_parsed_capabilities.json")
    with open(out1, "w") as f:
        json.dump(capabilities, f, indent=2)
    print(f"  [1/3] {out1} ({len(capabilities)} capabilities)")

    out2 = os.path.join(OUTPUT_DIR, "bvf_parsed_data_requirements.json")
    with open(out2, "w") as f:
        json.dump(data_requirements, f, indent=2)
    print(f"  [2/3] {out2} ({len(data_requirements)} data requirements)")

    out3 = os.path.join(OUTPUT_DIR, "bvf_reuse_matrix.json")
    with open(out3, "w") as f:
        json.dump(reuse_matrix, f, indent=2)
    print(f"  [3/3] {out3} ({len(reuse_matrix)} capability entries)")

    print("\nPhase 1 complete!")


if __name__ == "__main__":
    main()

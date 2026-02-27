#!/usr/bin/env python3
"""Phase 1: Parse BACR Interview Master - All Questions + Control Sheet"""

import openpyxl
import json
import csv
import os

OUTPUT_DIR = '/mnt/e/erwin/bacr_output'
BACR_FILE = '/mnt/e/erwin/BACR - INTERVIEW MASTER - DA004462.xlsm'
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Column mappings
COL_CATEGORY = 0
COL_SECTION = 1
COL_QUESTION = 2
COL_CURRENT_STATE = 4
COL_DESIRED_STATE = 5
COL_EMERGING = 7
COL_DEVELOPING = 8
COL_PRACTICING = 9
COL_INNOVATING = 10
COL_LEADING = 11

ROLE_COLS = {
    13: "Business Sponsor",
    14: "Business Analyst",
    15: "Data Analyst/Modeller/Scientist",
    16: "Architect",
    17: "Data Engineer",
    18: "IT Management",
    19: "IT Operations",
    20: "Program & Project Management",
    21: "Power User",
    22: "End User",
    23: "C-Level"
}

INDUSTRY_COLS = {
    25: "Retail",
    26: "Financial",
    27: "Communications",
    28: "Manufacturing",
    29: "Travel & Transport",
    30: "Energy & Natural Resources",
    31: "Healthcare"
}

BUS_FUNC_COLS = {
    33: "Customer",
    34: "Product",
    35: "Financial",
    36: "Operations",
    37: "Supply Chain",
    38: "Analytics Operations",
    39: "Asset Lifecycle Mgmt",
    40: "Human Resources",
    41: "Marketing",
    42: "Risks and Threats"
}

# Separator row patterns to skip
SEPARATOR_PATTERNS = [
    'category', 'Category -', 'Category-',
    'Information Category', 'Systems Category', 'Business Category',
    'Outcomes Category', 'Governance Category', 'Culture Category',
    'Agility Category', 'Applications Category',
    'Does your organisation have:'
]

def is_separator_row(text):
    if not text:
        return True
    t = text.strip()
    if not t:
        return True
    for pat in SEPARATOR_PATTERNS:
        if pat.lower() in t.lower() and len(t) < 80:
            return True
    return False

def cell_val(row, idx):
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ''

def is_yes(row, idx):
    v = cell_val(row, idx).lower()
    return v in ('y', '1', 'yes')

# ============================================================
# 1.1 Parse "All Questions" Sheet
# ============================================================
print("Opening BACR Interview Master...")
wb = openpyxl.load_workbook(BACR_FILE, read_only=True, data_only=True, keep_vba=False)
ws = wb['All Questions']

bacr_questions = []
current_cat = None
current_sec = None
question_id = 0

print("Parsing All Questions sheet...")
for i, row in enumerate(ws.iter_rows(min_row=7, values_only=True)):
    # Update category (forward-fill)
    if row[COL_CATEGORY]:
        cat = str(row[COL_CATEGORY]).strip()
        if cat and cat not in ('Category',):
            current_cat = cat

    # Update section (forward-fill)
    if row[COL_SECTION]:
        sec = str(row[COL_SECTION]).strip()
        if sec and sec not in ('Section',):
            current_sec = sec

    # Get question text
    q_text = cell_val(row, COL_QUESTION)
    if not q_text or is_separator_row(q_text):
        continue

    if not current_cat:
        continue

    question_id += 1

    # Maturity descriptors
    emerging = cell_val(row, COL_EMERGING)
    developing = cell_val(row, COL_DEVELOPING)
    practicing = cell_val(row, COL_PRACTICING)
    innovating = cell_val(row, COL_INNOVATING)
    leading = cell_val(row, COL_LEADING)

    is_tba = ('TBA' in q_text) or (not emerging and not developing and not practicing)

    # Role filters
    applicable_roles = [name for col, name in ROLE_COLS.items() if is_yes(row, col)]

    # Industry filters
    applicable_industries = [name for col, name in INDUSTRY_COLS.items() if is_yes(row, col)]

    # Business function filters
    applicable_bus_functions = [name for col, name in BUS_FUNC_COLS.items() if is_yes(row, col)]

    is_financial = 'Financial' in applicable_industries

    bacr_questions.append({
        'id': question_id,
        'category': current_cat,
        'section': current_sec or '',
        'question': q_text,
        'maturity_descriptors': {
            '1_emerging': emerging,
            '2_developing': developing,
            '3_practicing': practicing,
            '4_innovating': innovating,
            '5_leading': leading
        },
        'applicable_roles': applicable_roles,
        'applicable_industries': applicable_industries,
        'applicable_bus_functions': applicable_bus_functions,
        'is_financial_industry': is_financial,
        'is_tba': is_tba
    })

print(f"  Parsed {len(bacr_questions)} questions")

# ============================================================
# 1.3 Parse "Control Sheet" - Review Types
# ============================================================
print("Parsing Control Sheet for review types...")
ws_ctrl = wb['Control Sheet']

REVIEW_TYPE_COLS = {
    6: "Big Data Capability Review",
    7: "Data Warehouse Maturity Review",
    8: "Ecosystem Architecture Review",
    9: "AI and Machine Learning Review",
    10: "Custom Review"
}

review_configs = {}  # (category, section) -> {review_type: Include/Exclude}
current_ctrl_cat = None

for i, row in enumerate(ws_ctrl.iter_rows(min_row=29, max_row=100, values_only=True)):
    row_num = i + 29
    # Category in col 2 (sometimes col 1)
    cat_val = cell_val(row, 2) if len(row) > 2 else ''
    sec_val = cell_val(row, 3) if len(row) > 3 else ''
    # Also check cols 12/13 for the actual category/section used in filtering
    filter_cat = cell_val(row, 12) if len(row) > 12 else ''
    filter_sec = cell_val(row, 13) if len(row) > 13 else ''

    if cat_val and cat_val not in ('Category', 'Module Name'):
        current_ctrl_cat = cat_val
    if not filter_cat and not filter_sec:
        continue
    if filter_cat:
        current_ctrl_cat = filter_cat

    key = (current_ctrl_cat or '', filter_sec)
    if not key[1]:
        continue

    review_status = {}
    for col_idx, review_name in REVIEW_TYPE_COLS.items():
        val = cell_val(row, col_idx) if len(row) > col_idx else ''
        if val in ('Include', 'Exclude'):
            review_status[review_name] = val

    review_configs[key] = review_status

wb.close()

# Assign review types to questions
section_mapping = {}
for (cat, sec), statuses in review_configs.items():
    section_mapping[(cat, sec)] = statuses

for q in bacr_questions:
    review_types = []
    key = (q['category'], q['section'])
    if key in section_mapping:
        for rt, status in section_mapping[key].items():
            if status == 'Include':
                review_types.append(rt)
    q['review_types'] = review_types

# ============================================================
# 1.4 Outputs
# ============================================================

# bacr_all_questions.json
print("Writing bacr_all_questions.json...")
with open(os.path.join(OUTPUT_DIR, 'bacr_all_questions.json'), 'w') as f:
    json.dump(bacr_questions, f, indent=2)

# bacr_all_questions.csv
print("Writing bacr_all_questions.csv...")
with open(os.path.join(OUTPUT_DIR, 'bacr_all_questions.csv'), 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow([
        'Question_ID', 'Category', 'Section', 'Question_Text',
        'Emerging', 'Developing', 'Practicing', 'Innovating', 'Leading',
        'Roles', 'Industries', 'Business_Functions',
        'Is_Financial', 'Is_TBA', 'Review_Types'
    ])
    for q in bacr_questions:
        writer.writerow([
            q['id'], q['category'], q['section'], q['question'],
            q['maturity_descriptors']['1_emerging'],
            q['maturity_descriptors']['2_developing'],
            q['maturity_descriptors']['3_practicing'],
            q['maturity_descriptors']['4_innovating'],
            q['maturity_descriptors']['5_leading'],
            '; '.join(q['applicable_roles']),
            '; '.join(q['applicable_industries']),
            '; '.join(q['applicable_bus_functions']),
            q['is_financial_industry'],
            q['is_tba'],
            '; '.join(q['review_types'])
        ])

# bacr_category_summary.json
print("Building category summary...")
from collections import defaultdict
cat_data = defaultdict(lambda: {
    'question_count': 0, 'sections': set(), 'financial_question_count': 0,
    'tba_count': 0, 'business_functions': defaultdict(int),
    'review_types': defaultdict(int)
})

for q in bacr_questions:
    c = cat_data[q['category']]
    c['question_count'] += 1
    c['sections'].add(q['section'])
    if q['is_financial_industry']:
        c['financial_question_count'] += 1
    if q['is_tba']:
        c['tba_count'] += 1
    for bf in q['applicable_bus_functions']:
        c['business_functions'][bf] += 1
    for rt in q['review_types']:
        c['review_types'][rt] += 1

# Calculate totals
total_q = len(bacr_questions)
total_financial = sum(1 for q in bacr_questions if q['is_financial_industry'])
total_tba = sum(1 for q in bacr_questions if q['is_tba'])
review_type_counts = defaultdict(int)
for q in bacr_questions:
    for rt in q['review_types']:
        review_type_counts[rt] += 1

summary = {
    'categories': {},
    'totals': {
        'total_questions': total_q,
        'financial_industry_questions': total_financial,
        'tba_questions': total_tba,
        'questions_by_review_type': dict(review_type_counts)
    }
}

for cat_name in sorted(cat_data.keys()):
    c = cat_data[cat_name]
    summary['categories'][cat_name] = {
        'question_count': c['question_count'],
        'sections': sorted(c['sections']),
        'financial_question_count': c['financial_question_count'],
        'tba_count': c['tba_count'],
        'business_functions': dict(c['business_functions']),
        'review_types': dict(c['review_types'])
    }

print("Writing bacr_category_summary.json...")
with open(os.path.join(OUTPUT_DIR, 'bacr_category_summary.json'), 'w') as f:
    json.dump(summary, f, indent=2)

# Print summary
print("\n=== BACR PARSING COMPLETE ===")
print(f"Total questions: {total_q}")
print(f"Financial industry questions: {total_financial}")
print(f"TBA questions: {total_tba}")
print(f"\nCategories:")
for cat_name, data in sorted(summary['categories'].items()):
    print(f"  {cat_name}: {data['question_count']} questions, {data['financial_question_count']} financial, sections: {data['sections']}")
print(f"\nReview types:")
for rt, count in sorted(review_type_counts.items()):
    print(f"  {rt}: {count} questions")
print(f"\nOutputs written to {OUTPUT_DIR}/")

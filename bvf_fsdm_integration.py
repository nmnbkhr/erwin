#!/usr/bin/env python3
"""
BVF-FSDM Integration & Profitability Engine
============================================
Integrates Teradata Banking Business Value Framework (BVF) v1.2
with Financial Services Data Model (FSDM) v16.00.00 XSD Schema
to build a Customer Profitability Engine for Pakistani banking.

Usage: python3 bvf_fsdm_integration.py
"""

import csv
import json
import os
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl

# ── Configuration ─────────────────────────────────────────────────────────
BASE_DIR = "/mnt/e/erwin"
BVF_FILE = os.path.join(BASE_DIR, "Banking Business Value Framework Data Mappings 1.2.xlsm")
FSDM_ENTITY_CATALOG = os.path.join(BASE_DIR, "fsdm_output", "fsdm_entity_catalog.csv")
FSDM_RELATIONSHIPS = os.path.join(BASE_DIR, "fsdm_output", "fsdm_relationships.csv")
FSDM_STATS = os.path.join(BASE_DIR, "fsdm_output", "fsdm_stats.json")
FSDM_DOMAIN_MAP = os.path.join(BASE_DIR, "fsdm_output", "fsdm_domain_map.json")
OUTPUT_DIR = os.path.join(BASE_DIR, "bvf_fsdm_output")

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════
# SECTION 1: BVF EXCEL PARSER
# ══════════════════════════════════════════════════════════════════════════

def parse_bvf_excel():
    """Parse all sheets from the BVF Excel workbook."""
    print("=" * 60)
    print("SECTION 1: Parsing BVF Excel")
    print("=" * 60)

    wb = openpyxl.load_workbook(BVF_FILE, read_only=True, data_only=True)

    # 1a. Parse Capability-to-Data sheet
    cap_to_data, subject_areas, data_requirements = parse_capability_to_data(wb)

    # 1b. Parse Data Reuse Analysis sheet
    reuse_analysis = parse_reuse_analysis(wb)

    # 1c. Parse Data Reuse Matrix sheet
    reuse_matrix = parse_reuse_matrix(wb)

    # 1d. Parse Data-to-Capability sheet (reverse mapping)
    data_to_cap = parse_data_to_capability(wb)

    wb.close()

    # Export clean CSVs
    export_bvf_csvs(cap_to_data, subject_areas, data_requirements,
                     reuse_analysis, reuse_matrix, data_to_cap)

    return {
        "cap_to_data": cap_to_data,
        "subject_areas": subject_areas,
        "data_requirements": data_requirements,
        "reuse_analysis": reuse_analysis,
        "reuse_matrix": reuse_matrix,
        "data_to_cap": data_to_cap,
    }


def parse_capability_to_data(wb):
    """Parse 'Banking BVF Capability to Data' sheet."""
    print("  Parsing 'Banking BVF Capability to Data'...")
    ws = wb["Banking BVF Capability to Data"]

    # Row 3: FSDM Subject Areas (cols 4-116, 0-indexed)
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    # Row 4: Data Requirements
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    # Build subject area + data requirement lists for columns 4..116
    subject_areas = {}  # col_idx -> subject area name
    data_requirements = {}  # col_idx -> data requirement name

    for col_idx in range(4, min(len(row3), 118)):
        sa = row3[col_idx]
        dr = row4[col_idx] if col_idx < len(row4) else None
        if sa:
            subject_areas[col_idx] = str(sa).strip()
        if dr:
            data_requirements[col_idx] = str(dr).strip()

    # Parse capability rows (5+)
    current_theme = None
    current_cap = None
    cap_to_data = []

    for row_idx in range(5, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

        if row[0]:
            current_theme = str(row[0]).strip()
        if row[1]:
            current_cap = str(row[1]).strip()

        sub_cap = row[3] if len(row) > 3 and row[3] else None
        if not sub_cap:
            continue
        sub_cap = str(sub_cap).strip()

        # Collect which data requirements this sub-capability needs
        mappings = {}
        for col_idx in range(4, min(len(row), 118)):
            val = row[col_idx]
            if val == 1 or val == "1":
                if col_idx in data_requirements:
                    dr_name = data_requirements[col_idx]
                    sa_name = subject_areas.get(col_idx, "Unknown")
                    mappings[dr_name] = sa_name

        cap_to_data.append({
            "theme": current_theme,
            "capability": current_cap,
            "sub_capability": sub_cap,
            "data_mappings": mappings,
            "data_count": len(mappings),
        })

    print(f"    Found {len(cap_to_data)} sub-capabilities")
    print(f"    Found {len(data_requirements)} data requirement columns")
    print(f"    Found {len(set(subject_areas.values()))} unique FSDM subject areas")
    return cap_to_data, subject_areas, data_requirements


def parse_reuse_analysis(wb):
    """Parse 'Banking BVF Data Reuse Analysis' sheet."""
    print("  Parsing 'Banking BVF Data Reuse Analysis'...")
    ws = wb["Banking BVF Data Reuse Analysis"]

    # Row 4: Data requirement headers (cols 7+)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    dr_cols = {}
    for col_idx in range(7, len(row4)):
        if row4[col_idx]:
            dr_cols[col_idx] = str(row4[col_idx]).strip()

    # Row 6+: capability reuse data
    # Row 6 has the base capability data
    # Row 7 has overrides
    reuse_data = []
    for row_idx in range(6, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        cap_name = row[6] if len(row) > 6 and row[6] else None
        if not cap_name:
            continue
        cap_name = str(cap_name).strip()

        scores = {}
        for col_idx, dr_name in dr_cols.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    scores[dr_name] = float(row[col_idx])
                except (ValueError, TypeError):
                    pass

        if scores:
            reuse_data.append({
                "capability": cap_name,
                "data_scores": scores,
            })

    print(f"    Found {len(reuse_data)} reuse analysis rows")
    return reuse_data


def parse_reuse_matrix(wb):
    """Parse 'Banking BVF Data Reuse Matrix' sheet."""
    print("  Parsing 'Banking BVF Data Reuse Matrix'...")
    ws = wb["Banking BVF Data Reuse Matrix"]

    # Row 4: sub-capability names as column headers (cols 4+)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    col_caps = {}
    for col_idx in range(4, len(row4)):
        if row4[col_idx]:
            col_caps[col_idx] = str(row4[col_idx]).strip()

    # Rows 5+: matrix data
    current_theme = None
    current_cap = None
    matrix = []

    for row_idx in range(5, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        if row[0]:
            current_theme = str(row[0]).strip()
        if row[1]:
            current_cap = str(row[1]).strip()

        sub_cap = row[3] if len(row) > 3 and row[3] else None
        if not sub_cap:
            continue
        sub_cap = str(sub_cap).strip()

        coefficients = {}
        for col_idx, target_cap in col_caps.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = float(row[col_idx])
                    coefficients[target_cap] = val
                except (ValueError, TypeError):
                    pass

        matrix.append({
            "theme": current_theme,
            "capability": current_cap,
            "sub_capability": sub_cap,
            "reuse_coefficients": coefficients,
        })

    print(f"    Found {len(matrix)} rows in reuse matrix")
    return matrix


def parse_data_to_capability(wb):
    """Parse 'Banking BVF Data to Capability' sheet (reverse mapping)."""
    print("  Parsing 'Banking BVF Data to Capability'...")
    ws = wb["Banking BVF Data to Capability"]

    # Row 4: sub-capability names as column headers (cols 4+)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    col_caps = {}
    for col_idx in range(4, len(row4)):
        if row4[col_idx]:
            col_caps[col_idx] = str(row4[col_idx]).strip()

    # Rows 5+: data requirement → capabilities
    data_to_cap = []
    for row_idx in range(5, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        dr_name = row[0] if row[0] else None
        sa_name = row[1] if len(row) > 1 and row[1] else None
        if not dr_name:
            continue

        dr_name = str(dr_name).strip()
        sa_name = str(sa_name).strip() if sa_name else "Unknown"

        capabilities = []
        for col_idx, cap_name in col_caps.items():
            if col_idx < len(row) and (row[col_idx] == 1 or row[col_idx] == "1"):
                capabilities.append(cap_name)

        data_to_cap.append({
            "data_requirement": dr_name,
            "fsdm_subject_area": sa_name,
            "capabilities": capabilities,
            "capability_count": len(capabilities),
        })

    print(f"    Found {len(data_to_cap)} data requirements with capability mappings")
    return data_to_cap


def export_bvf_csvs(cap_to_data, subject_areas, data_requirements,
                     reuse_analysis, reuse_matrix, data_to_cap):
    """Export parsed BVF data as clean CSVs."""
    print("  Exporting BVF CSVs...")

    # 1. Capability summary
    path = os.path.join(OUTPUT_DIR, "bvf_capability_summary.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Theme", "Capability_Group", "Sub_Capability", "Data_Requirements_Count"])
        for c in cap_to_data:
            w.writerow([c["theme"], c["capability"], c["sub_capability"], c["data_count"]])
    print(f"    -> {path}")

    # 2. Data requirements summary
    path = os.path.join(OUTPUT_DIR, "bvf_data_requirements.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Data_Requirement", "FSDM_Subject_Area", "Capabilities_Using", "Capability_Count"])
        for d in data_to_cap:
            w.writerow([
                d["data_requirement"],
                d["fsdm_subject_area"],
                "; ".join(d["capabilities"]),
                d["capability_count"],
            ])
    print(f"    -> {path}")

    # 3. Reuse matrix CSV
    path = os.path.join(OUTPUT_DIR, "bvf_reuse_matrix.csv")
    # Collect all sub-capability names for columns
    all_sub_caps = [m["sub_capability"] for m in reuse_matrix]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Theme", "Capability", "Sub_Capability"] + all_sub_caps)
        for m in reuse_matrix:
            row = [m["theme"], m["capability"], m["sub_capability"]]
            for sc in all_sub_caps:
                row.append(m["reuse_coefficients"].get(sc, ""))
            w.writerow(row)
    print(f"    -> {path}")


# ══════════════════════════════════════════════════════════════════════════
# SECTION 2: LOAD FSDM DATA
# ══════════════════════════════════════════════════════════════════════════

def load_fsdm_data():
    """Load FSDM entity catalog and relationships from previous analysis."""
    print("\n" + "=" * 60)
    print("SECTION 2: Loading FSDM Data")
    print("=" * 60)

    # Load entity catalog
    entities = {}
    with open(FSDM_ENTITY_CATALOG, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entities[row["Entity"]] = {
                "domain": row["Domain"],
                "column_count": int(row["Column_Count"]),
                "relationship_count": int(row["Relationship_Count"]),
                "parent_entity": row.get("Parent_Entity", ""),
                "subtype_count": int(row.get("Subtype_Count", 0)),
                "description": row.get("Description", ""),
            }
    print(f"  Loaded {len(entities)} FSDM entities")

    # Load relationships
    relationships = []
    with open(FSDM_RELATIONSHIPS, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            relationships.append(row)
    print(f"  Loaded {len(relationships)} FSDM relationships")

    # Load stats
    with open(FSDM_STATS, "r", encoding="utf-8") as f:
        stats = json.load(f)
    print(f"  Loaded FSDM stats (v{stats['model_info']['version']})")

    # Load domain map if available
    domain_map = {}
    if os.path.exists(FSDM_DOMAIN_MAP):
        with open(FSDM_DOMAIN_MAP, "r", encoding="utf-8") as f:
            domain_map = json.load(f)
        print(f"  Loaded domain map ({len(domain_map)} domains)")

    return entities, relationships, stats, domain_map


# ══════════════════════════════════════════════════════════════════════════
# SECTION 3: BVF-TO-FSDM ENTITY MAPPING
# ══════════════════════════════════════════════════════════════════════════

# Knowledge-based mapping: BVF Data Requirements → FSDM Entities
# This is the core intellectual mapping that connects business requirements
# to the canonical data model.
BVF_TO_FSDM_MAPPING = {
    "Single Customer View (Master Record)": {
        "entities": ["PARTY", "INDIVIDUAL", "ORGANIZATION", "HOUSEHOLD",
                      "PARTY_RELATIONSHIP", "PARTY_IDENTIFICATION"],
        "confidence": "High",
        "notes": "Core party/customer master record entities"
    },
    "KYC Data": {
        "entities": ["PARTY_IDENTIFICATION", "PARTY_CLASSIFICATION",
                      "PARTY_IDENTITY_DOCUMENT", "PARTY_COMPLIANCE_VERIFICATION"],
        "confidence": "High",
        "notes": "Know-Your-Customer regulatory data"
    },
    "Customer Demographics": {
        "entities": ["INDIVIDUAL", "PARTY_DEMOGRAPHIC",
                      "PARTY_CLASSIFICATION", "PARTY_PREFERENCE"],
        "confidence": "High",
        "notes": "Individual/demographic attributes"
    },
    "Customer Attitudinal Data": {
        "entities": ["PARTY_PREFERENCE", "PARTY_CLASSIFICATION",
                      "PARTY_SURVEY_RESPONSE"],
        "confidence": "Medium",
        "notes": "Attitudinal/behavioral data - partial FSDM coverage"
    },
    "Customer Segments": {
        "entities": ["PARTY_CLASSIFICATION", "PARTY_SEGMENT",
                      "PARTY_CLASSIFICATION_VALUE"],
        "confidence": "High",
        "notes": "Customer segmentation classifications"
    },
    "Customer Contact Preferences": {
        "entities": ["PARTY_PREFERENCE", "PARTY_CONTACT_POINT",
                      "ELECTRONIC_ADDRESS", "POSTAL_ADDRESS"],
        "confidence": "High",
        "notes": "Contact channel preferences"
    },
    "Customer Marketing Preferences": {
        "entities": ["PARTY_PREFERENCE", "PARTY_MARKETING_PREFERENCE",
                      "PARTY_CONSENT"],
        "confidence": "Medium",
        "notes": "Marketing opt-in/opt-out preferences"
    },
    "Customer Satisfaction Measurement (E.g. NPS)": {
        "entities": ["PARTY_SURVEY_RESPONSE", "SURVEY", "SURVEY_QUESTION_RESPONSE"],
        "confidence": "Medium",
        "notes": "NPS/CSAT via survey entities"
    },
    "Economic Forecast Data": {
        "entities": ["ECONOMIC_INDEX", "ECONOMIC_INDEX_VALUE",
                      "MARKET_INDEX"],
        "confidence": "Medium",
        "notes": "External economic indicators - may need extension"
    },
    "Legal Actions": {
        "entities": ["PARTY_LEGAL_ACTION", "LEGAL_ACTION",
                      "PARTY_LITIGATION"],
        "confidence": "Medium",
        "notes": "Legal proceedings involving parties"
    },
    "External Prospect Data / Lists": {
        "entities": ["PARTY", "PROSPECT", "PARTY_LIST",
                      "PARTY_LIST_MEMBERSHIP"],
        "confidence": "Medium",
        "notes": "External prospect/lead data"
    },
    "External Data - Social Media Data": {
        "entities": ["SOCIAL_MEDIA_ACCOUNT", "SOCIAL_MEDIA_TRANSFER",
                      "SOCIAL_MEDIA_POST"],
        "confidence": "High",
        "notes": "Social media interaction entities"
    },
    "External Data - Bureau data": {
        "entities": ["PARTY_CREDIT_BUREAU_REPORT", "CREDIT_BUREAU_SCORE",
                      "PARTY_EXTERNAL_RATING"],
        "confidence": "High",
        "notes": "Credit bureau reports and scores"
    },
    "External Data - Credit Rating": {
        "entities": ["PARTY_EXTERNAL_RATING", "PARTY_LIABILITY_CREDIT_RATING",
                      "CREDIT_BUREAU_SCORE"],
        "confidence": "High",
        "notes": "External credit ratings"
    },
    "External Economic Data (E.g. Property Prices, Salary)": {
        "entities": ["ECONOMIC_INDEX", "ECONOMIC_INDEX_VALUE",
                      "MARKET_INDEX", "MARKET_INDEX_VALUE"],
        "confidence": "Medium",
        "notes": "Economic indicators and indices"
    },
    "Master & Reference Data - staff / colleague": {
        "entities": ["INDIVIDUAL", "ASSOCIATE", "ASSOCIATE_ASSIGNMENT",
                      "INTERNAL_ORGANIZATION_UNIT"],
        "confidence": "High",
        "notes": "Employee/staff reference data"
    },
    "Sales and Performance Forecasts, Targets & KPIs": {
        "entities": ["PERFORMANCE_METRIC", "PERFORMANCE_TARGET",
                      "ORGANIZATION_PERFORMANCE"],
        "confidence": "Low",
        "notes": "Performance management - limited FSDM coverage, needs extension"
    },
    "HR Performance Data": {
        "entities": ["ASSOCIATE", "ASSOCIATE_PERFORMANCE_REVIEW",
                      "ASSOCIATE_ASSIGNMENT"],
        "confidence": "Medium",
        "notes": "HR performance review data"
    },
    "Customer Core Contact Details": {
        "entities": ["PARTY_CONTACT_POINT", "ELECTRONIC_ADDRESS",
                      "POSTAL_ADDRESS", "TELEPHONE_NUMBER"],
        "confidence": "High",
        "notes": "Core contact information"
    },
    "Master & Reference Data - geography / location": {
        "entities": ["GEOGRAPHICAL_AREA", "GEOGRAPHICAL_UNIT",
                      "SITE", "POSTAL_ADDRESS"],
        "confidence": "High",
        "notes": "Geography/location reference data"
    },
    "Master & Reference Data - product ": {
        "entities": ["PRODUCT", "PRODUCT_GROUP", "PRODUCT_TYPE",
                      "PRODUCT_FEATURE", "PRODUCT_CLASSIFICATION"],
        "confidence": "High",
        "notes": "Product master data hierarchy"
    },
    "Master & Reference Data - FTP rates": {
        "entities": ["INTEREST_RATE", "INTEREST_RATE_INDEX",
                      "PRODUCT_INTEREST_RATE"],
        "confidence": "High",
        "notes": "Funds Transfer Pricing rate curves"
    },
    "Master & Reference Data - exchange rates": {
        "entities": ["CURRENCY", "CURRENCY_EXCHANGE_RATE",
                      "CURRENCY_PAIR"],
        "confidence": "High",
        "notes": "FX exchange rate data"
    },
    "Product Revenue, Cost & Margin": {
        "entities": ["PRODUCT", "PRODUCT_PRICING", "PRODUCT_COST",
                      "AGREEMENT_SUMMARY"],
        "confidence": "High",
        "notes": "Product-level P&L components"
    },
    "Product NPV's": {
        "entities": ["PRODUCT", "PRODUCT_VALUATION",
                      "AGREEMENT_SUMMARY"],
        "confidence": "Medium",
        "notes": "Product net present values - may need extension"
    },
    "External Data - market interest rates": {
        "entities": ["INTEREST_RATE_INDEX", "INTEREST_RATE",
                      "MARKET_INDEX", "MARKET_INDEX_VALUE"],
        "confidence": "High",
        "notes": "KIBOR/market rate benchmarks"
    },
    "External Data - market prices (E.g. equities, bonds, commodities)": {
        "entities": ["MARKET_INDEX", "MARKET_INDEX_VALUE",
                      "INVESTMENT_PRODUCT", "INVESTMENT_PRODUCT_PRICE"],
        "confidence": "High",
        "notes": "Market price data for traded instruments"
    },
    "External Data - Competitor Product & Pricing": {
        "entities": ["PRODUCT", "MARKET_INDEX"],
        "confidence": "Low",
        "notes": "Competitor data - needs custom extension to FSDM"
    },
    "Account / Policy / Service Holdings": {
        "entities": ["AGREEMENT", "PARTY_AGREEMENT_ROLE",
                      "AGREEMENT_PARTY_ROLE", "AGREEMENT_RELATIONSHIP"],
        "confidence": "High",
        "notes": "Customer account/product holdings"
    },
    "Account / Policy / Service Detail": {
        "entities": ["AGREEMENT", "LOAN_TERM_AGREEMENT",
                      "DEPOSIT_AGREEMENT", "INSURANCE_AGREEMENT",
                      "CREDIT_CARD_AGREEMENT"],
        "confidence": "High",
        "notes": "Account detail by agreement subtype"
    },
    "Account Status (open closed active / domant /inactive / closure reason )": {
        "entities": ["AGREEMENT", "AGREEMENT_STATUS",
                      "AGREEMENT_STATUS_HISTORY"],
        "confidence": "High",
        "notes": "Account lifecycle status tracking"
    },
    "Account Terms & Conditions": {
        "entities": ["AGREEMENT", "AGREEMENT_CONDITION",
                      "AGREEMENT_TERM"],
        "confidence": "High",
        "notes": "Contractual terms and conditions"
    },
    "Account Feature Usage": {
        "entities": ["AGREEMENT_FEATURE", "FEATURE",
                      "DESCRIPTIVE_FEATURE"],
        "confidence": "High",
        "notes": "Product feature opt-in and usage"
    },
    "Application Data (New Customer, New Product)": {
        "entities": ["APPLICATION", "APPLICATION_STATUS",
                      "APPLICATION_PARTY_ROLE"],
        "confidence": "High",
        "notes": "Loan/product applications"
    },
    "Coverage (insurance)": {
        "entities": ["INSURANCE_COVERAGE", "INSURANCE_AGREEMENT",
                      "COVERAGE_TYPE"],
        "confidence": "High",
        "notes": "Insurance coverage details"
    },
    "Account Balances": {
        "entities": ["AGREEMENT_SUMMARY", "BALANCE_SUMMARY",
                      "AGREEMENT_BALANCE"],
        "confidence": "High",
        "notes": "Account balance snapshots and summaries"
    },
    "Investment Value": {
        "entities": ["INVESTMENT_ACCOUNT", "INVESTMENT_HOLDING",
                      "INVESTMENT_PRODUCT_PRICE"],
        "confidence": "High",
        "notes": "Investment portfolio valuations"
    },
    "Market / Trading Book Positions": {
        "entities": ["TRADE_ORDER", "TRADE_EXECUTION",
                      "TRADING_POSITION"],
        "confidence": "High",
        "notes": "Trading book positions"
    },
    "Mark to market valuations": {
        "entities": ["INVESTMENT_PRODUCT_PRICE", "MARKET_INDEX_VALUE",
                      "AGREEMENT_VALUATION"],
        "confidence": "Medium",
        "notes": "Mark-to-market valuations"
    },
    "Cashflow Projections": {
        "entities": ["AGREEMENT_CASHFLOW", "CASHFLOW_SCHEDULE",
                      "AGREEMENT_SUMMARY"],
        "confidence": "Medium",
        "notes": "Projected cashflows - may need extension"
    },
    "Account Maturity Dates & Repricing Schedules": {
        "entities": ["AGREEMENT", "AGREEMENT_TERM",
                      "AGREEMENT_REPRICING_SCHEDULE"],
        "confidence": "High",
        "notes": "Maturity and repricing dates"
    },
    "Account Repayment Schedules": {
        "entities": ["AGREEMENT_REPAYMENT_SCHEDULE",
                      "LOAN_TERM_AGREEMENT"],
        "confidence": "High",
        "notes": "Loan repayment schedules"
    },
    "Amortisation Schedules": {
        "entities": ["AGREEMENT_AMORTIZATION_SCHEDULE",
                      "LOAN_TERM_AGREEMENT"],
        "confidence": "Medium",
        "notes": "Amortization schedules"
    },
    "Customer Illustrations": {
        "entities": ["DOCUMENT", "PARTY_DOCUMENT"],
        "confidence": "Low",
        "notes": "Customer illustration documents - generic doc entities"
    },
    "Customer Quotes": {
        "entities": ["APPLICATION", "QUOTE", "DOCUMENT"],
        "confidence": "Low",
        "notes": "Customer quotes - limited FSDM coverage"
    },
    "Collateral Agreements": {
        "entities": ["COLLATERAL_AGREEMENT", "COLLATERAL",
                      "COLLATERAL_VALUATION", "PARTY_ASSET"],
        "confidence": "High",
        "notes": "Collateral and security agreements"
    },
    "Risk Exposures": {
        "entities": ["AGREEMENT_RISK_METRIC", "RISK_EXPOSURE",
                      "CREDIT_RISK_SUMMARY"],
        "confidence": "High",
        "notes": "Credit/market/operational risk exposures"
    },
    "Risk Weighted Assets & Capital Results": {
        "entities": ["AGREEMENT_RISK_METRIC", "BANK_CAPITAL_SUMMARY",
                      "CAPITAL_REQUIREMENTS_SUMMARY"],
        "confidence": "High",
        "notes": "Basel RWA and capital calculations"
    },
    "Profitability Results": {
        "entities": ["AGREEMENT_SUMMARY", "AGREEMENT_PROFITABILITY",
                      "PRODUCT_PROFITABILITY"],
        "confidence": "High",
        "notes": "Profitability metrics and results - core fact source"
    },
    "Account Limits": {
        "entities": ["AGREEMENT_LIMIT", "LIMIT",
                      "PARTY_LIMIT"],
        "confidence": "High",
        "notes": "Credit limits and facility limits"
    },
    "Predictive Models - propensity to buy": {
        "entities": ["ANALYTICAL_MODEL", "MODEL_RUN",
                      "MODEL_SCORE", "PARTY_MODEL_SCORE"],
        "confidence": "High",
        "notes": "Propensity model scores"
    },
    "Predictive Models - attrition": {
        "entities": ["ANALYTICAL_MODEL", "MODEL_RUN",
                      "MODEL_SCORE", "PARTY_MODEL_SCORE"],
        "confidence": "High",
        "notes": "Attrition/churn model scores"
    },
    "Predictive Models - price elasticity": {
        "entities": ["ANALYTICAL_MODEL", "MODEL_RUN",
                      "MODEL_SCORE"],
        "confidence": "Medium",
        "notes": "Price elasticity models"
    },
    "Attribution Models - response": {
        "entities": ["ANALYTICAL_MODEL", "MODEL_RUN",
                      "CAMPAIGN_RESPONSE"],
        "confidence": "Medium",
        "notes": "Marketing attribution models"
    },
    "Lead Foodchain data": {
        "entities": ["OPPORTUNITY", "OPPORTUNITY_STATUS",
                      "OPPORTUNITY_PARTY_ROLE"],
        "confidence": "High",
        "notes": "Sales lead/opportunity pipeline"
    },
    "Predictive Models - risk": {
        "entities": ["ANALYTICAL_MODEL", "MODEL_RUN",
                      "MODEL_SCORE", "PARTY_LIABILITY_CREDIT_RATING"],
        "confidence": "High",
        "notes": "Risk scoring models (PD, LGD, EAD)"
    },
    "Risk Segments, Cohorts & Groupings": {
        "entities": ["PARTY_CLASSIFICATION", "RISK_GRADE_VALUE",
                      "RISK_GRADE_SCHEME"],
        "confidence": "High",
        "notes": "Risk segmentation and grading"
    },
    "Provisions, Losses & Writeoffs": {
        "entities": ["AGREEMENT_RISK_METRIC", "PROVISION",
                      "WRITE_OFF", "LOSS_EVENT"],
        "confidence": "High",
        "notes": "Credit loss provisions and write-offs (IFRS9/ECL)"
    },
    "Financial Transactions": {
        "entities": ["MONETARY_TRANSACTION", "TRANSACTION_TYPE",
                      "FINANCIAL_TRANSACTION", "PAYMENT"],
        "confidence": "High",
        "notes": "Core financial transaction records"
    },
    "Account Fees, Commissions & Charges": {
        "entities": ["MONETARY_TRANSACTION", "FEE_TRANSACTION",
                      "COMMISSION_TRANSACTION", "AGREEMENT_FEE"],
        "confidence": "High",
        "notes": "Fee and commission income records"
    },
    "Financial Interbank Payments": {
        "entities": ["PAYMENT", "INTERBANK_PAYMENT",
                      "FINANCIAL_TRANSACTION"],
        "confidence": "High",
        "notes": "Interbank payment/settlement records"
    },
    "Non-Financial Transactions": {
        "entities": ["EVENT", "NON_MONETARY_EVENT",
                      "DIRECT_CONTACT_EVENT"],
        "confidence": "Medium",
        "notes": "Non-monetary events and interactions"
    },
    "Service Usage": {
        "entities": ["CHANNEL_USAGE_METRIC", "SERVICE_USAGE_EVENT",
                      "CHANNEL_INSTANCE"],
        "confidence": "Medium",
        "notes": "Service/channel usage tracking"
    },
    "Complaint Data": {
        "entities": ["COMPLAINT", "COMPLAINT_STATUS",
                      "COMPLAINT_RESOLUTION"],
        "confidence": "High",
        "notes": "Customer complaint management"
    },
    "Offline Customer Interactions - Branch": {
        "entities": ["DIRECT_CONTACT_EVENT", "BRANCH_VISIT_EVENT",
                      "CHANNEL_INSTANCE"],
        "confidence": "High",
        "notes": "In-branch customer interactions"
    },
    "Offline Customer Interactions - Call Centre": {
        "entities": ["DIRECT_CONTACT_EVENT", "CALL_EVENT",
                      "CHANNEL_INSTANCE"],
        "confidence": "High",
        "notes": "Call center interactions"
    },
    "Offline Customer Interactions - letters": {
        "entities": ["DIRECT_CONTACT_EVENT", "CORRESPONDENCE_EVENT",
                      "DOCUMENT"],
        "confidence": "Medium",
        "notes": "Letter/mail correspondence"
    },
    "Offline Customer Interactions - sms text": {
        "entities": ["DIRECT_CONTACT_EVENT", "SMS_EVENT",
                      "ELECTRONIC_ADDRESS"],
        "confidence": "Medium",
        "notes": "SMS text message interactions"
    },
    "Online Interactions -  webchat (internal & external)": {
        "entities": ["WEB_VISIT", "WEB_PAGE_VIEW",
                      "ONLINE_INTERACTION_EVENT"],
        "confidence": "High",
        "notes": "Web chat interaction records"
    },
    "Online Interactions colleague intranet": {
        "entities": ["WEB_VISIT", "WEB_PAGE_VIEW"],
        "confidence": "Medium",
        "notes": "Internal intranet usage"
    },
    "Online Interactions customer digital banking site (web log)": {
        "entities": ["WEB_VISIT", "WEB_PAGE_VIEW",
                      "EXTENDED_SESSION_INFORMATION", "ONLINE_APPLICATION_ACCESS"],
        "confidence": "High",
        "notes": "Digital banking web analytics"
    },
    "Online Interactions customer general banking site": {
        "entities": ["WEB_VISIT", "WEB_PAGE_VIEW",
                      "ONLINE_INTERACTION_EVENT"],
        "confidence": "High",
        "notes": "Public website analytics"
    },
    "Online Interactions email - customer & employee": {
        "entities": ["DIRECT_CONTACT_EVENT", "EMAIL_EVENT",
                      "ELECTRONIC_ADDRESS"],
        "confidence": "Medium",
        "notes": "Email interaction tracking"
    },
    "Online Interactions mobile app": {
        "entities": ["WEB_VISIT", "WEB_PAGE_VIEW",
                      "MOBILE_APP_EVENT", "CHANNEL_INSTANCE"],
        "confidence": "High",
        "notes": "Mobile app usage analytics"
    },
    "Online Interactions off-site - ad server": {
        "entities": ["AD", "AD_PLACEMENT", "AD_IMPRESSION",
                      "AD_CAMPAIGN_OBJECTIVE"],
        "confidence": "High",
        "notes": "Ad server/display advertising data"
    },
    "Online Interactions prospect / unidentified customer (web log)": {
        "entities": ["WEB_VISIT", "WEB_PAGE_VIEW",
                      "VISITOR_INTERACTION_EVENT"],
        "confidence": "High",
        "notes": "Anonymous/prospect web analytics"
    },
    "Case History (Fraud, Money Laundering, Complaint)": {
        "entities": ["CASE", "CASE_EVENT", "CASE_PARTY_ROLE",
                      "FRAUD_CASE"],
        "confidence": "High",
        "notes": "Case management for fraud/AML/complaints"
    },
    "Claims": {
        "entities": ["CLAIM", "CLAIM_STATUS", "CLAIM_PARTY_ROLE",
                      "CLAIM_PAYMENT"],
        "confidence": "High",
        "notes": "Insurance claims management"
    },
    "Defaults & Limit Breaches": {
        "entities": ["AGREEMENT_STATUS", "DEFAULT_EVENT",
                      "LIMIT_BREACH_EVENT", "AGREEMENT_RISK_METRIC"],
        "confidence": "Medium",
        "notes": "Default and limit breach events"
    },
    "Closed Loop Outcomes / Channel Responses": {
        "entities": ["CAMPAIGN_RESPONSE", "CAMPAIGN_OFFER_RESPONSE",
                      "DIRECT_CONTACT_EVENT"],
        "confidence": "High",
        "notes": "Campaign closed-loop response tracking"
    },
    "External Data - Corporate Actions": {
        "entities": ["CORPORATE_ACTION", "INVESTMENT_PRODUCT"],
        "confidence": "High",
        "notes": "Corporate action events (dividends, splits)"
    },
    "Internal Process Activity Data": {
        "entities": ["EVENT", "PROCESS_EVENT",
                      "OPERATIONAL_EVENT"],
        "confidence": "Low",
        "notes": "Business process tracking - limited FSDM coverage"
    },
    "Master & Reference Data - merchant / transaction": {
        "entities": ["MERCHANT", "MERCHANT_CATEGORY",
                      "TRANSACTION_TYPE"],
        "confidence": "High",
        "notes": "Merchant and transaction type reference data"
    },
    "Customer interview / fact find": {
        "entities": ["DOCUMENT", "SURVEY", "SURVEY_QUESTION_RESPONSE",
                      "PARTY_DOCUMENT"],
        "confidence": "Medium",
        "notes": "Customer fact-find/interview documents"
    },
    "Physical Assets": {
        "entities": ["PARTY_ASSET", "FIXED_ASSET",
                      "REAL_PROPERTY"],
        "confidence": "High",
        "notes": "Physical/fixed asset records"
    },
    "Physical Asset Valuations": {
        "entities": ["PARTY_ASSET_VALUATION", "COLLATERAL_VALUATION",
                      "REAL_PROPERTY_VALUATION"],
        "confidence": "High",
        "notes": "Asset valuation records"
    },
    "Master & Reference Data - organisation hierarchy": {
        "entities": ["ORGANIZATION", "INTERNAL_ORGANIZATION_UNIT",
                      "ORGANIZATION_UNIT_RELATIONSHIP",
                      "ORGANIZATION_BUSINESS_TYPE"],
        "confidence": "High",
        "notes": "Internal organization/branch hierarchy"
    },
    "Operational metrics (E.g. Wait times, queue length)": {
        "entities": ["CHANNEL_USAGE_METRIC", "OPERATIONAL_EVENT",
                      "SERVICE_LEVEL_METRIC"],
        "confidence": "Low",
        "notes": "Operational metrics - needs extension"
    },
    "Offer Catalogue": {
        "entities": ["OFFER", "OFFER_TYPE", "CAMPAIGN_OFFER",
                      "PRODUCT"],
        "confidence": "High",
        "notes": "Marketing offer catalog"
    },
    "Master & Reference Data - Campaign": {
        "entities": ["CAMPAIGN", "CAMPAIGN_TYPE", "CAMPAIGN_STATUS",
                      "CAMPAIGN_WAVE"],
        "confidence": "High",
        "notes": "Campaign master data"
    },
    "Campaign Production Costs": {
        "entities": ["CAMPAIGN", "CAMPAIGN_COST",
                      "CAMPAIGN_BUDGET"],
        "confidence": "Medium",
        "notes": "Campaign cost tracking"
    },
    "Campaign History Data": {
        "entities": ["CAMPAIGN", "CAMPAIGN_RESPONSE",
                      "CAMPAIGN_OFFER_RESPONSE", "CAMPAIGN_WAVE"],
        "confidence": "High",
        "notes": "Historical campaign execution and responses"
    },
    "Media Asset Details": {
        "entities": ["MEDIA_ASSET", "DOCUMENT",
                      "DIGITAL_CONTENT"],
        "confidence": "Medium",
        "notes": "Media/content assets"
    },
    "Campaign Promotional Calendar": {
        "entities": ["CAMPAIGN", "CAMPAIGN_WAVE",
                      "CAMPAIGN_SCHEDULE"],
        "confidence": "Medium",
        "notes": "Campaign scheduling/calendar"
    },
    "Master & Reference Data - Channel": {
        "entities": ["CHANNEL_TYPE", "CHANNEL_INSTANCE",
                      "CHANNEL_CLASSIFICATION"],
        "confidence": "High",
        "notes": "Channel master data"
    },
    "Finance Systems - General Ledger ": {
        "entities": ["GL_ACCOUNT", "GL_MAIN_ACCOUNT",
                      "GL_ACCOUNT_BALANCE", "CHART_OF_ACCOUNTS"],
        "confidence": "High",
        "notes": "General ledger and chart of accounts"
    },
    "Finance Systems - ERP": {
        "entities": ["GL_ACCOUNT", "GL_MAIN_ACCOUNT",
                      "VENDOR", "VENDOR_INVOICE"],
        "confidence": "Medium",
        "notes": "ERP integration - AP/AR/inventory entities"
    },
    "Finance Systems - Budgets, Plans & Forecasts": {
        "entities": ["GL_ACCOUNT", "BUDGET", "FORECAST",
                      "FINANCIAL_PLAN"],
        "confidence": "Low",
        "notes": "Financial planning - limited FSDM coverage, needs extension"
    },
    "Model Metadata (E.g. defintion, assumptions, metrics)": {
        "entities": ["ANALYTICAL_MODEL", "MODEL_RUN",
                      "MODEL_PARAMETER"],
        "confidence": "High",
        "notes": "Analytical model metadata and governance"
    },
    "Risk Scenario & Stress Testing Results": {
        "entities": ["RISK_CALCULATION_SCENARIO",
                      "RISK_SCENARIO_RESULT", "STRESS_TEST_RESULT"],
        "confidence": "High",
        "notes": "Stress testing and scenario analysis results"
    },
    "Documents - Structured Data Capture (Survey, Application Forms)": {
        "entities": ["DOCUMENT", "DOCUMENT_TYPE",
                      "SURVEY", "APPLICATION"],
        "confidence": "High",
        "notes": "Structured document capture"
    },
    "Document Image ": {
        "entities": ["DOCUMENT", "DOCUMENT_IMAGE",
                      "DIGITAL_CONTENT"],
        "confidence": "Medium",
        "notes": "Document image/scan storage"
    },
    "Text Data": {
        "entities": ["DOCUMENT", "TEXT_CONTENT",
                      "DIGITAL_CONTENT"],
        "confidence": "Medium",
        "notes": "Unstructured text data"
    },
    "External Data - News / Industry Feeds E.g. Reuters": {
        "entities": ["DOCUMENT", "NEWS_FEED",
                      "MARKET_INDEX"],
        "confidence": "Low",
        "notes": "External news/data feeds - needs extension"
    },
    "Master & Reference Data - Risk Policy Rules": {
        "entities": ["RISK_GRADE_SCHEME", "RISK_GRADE_VALUE",
                      "RISK_POLICY", "BUSINESS_RULE"],
        "confidence": "Medium",
        "notes": "Risk policy and rules framework"
    },
    "Customer Research - Primary & Secondary": {
        "entities": ["SURVEY", "SURVEY_QUESTION_RESPONSE",
                      "PARTY_SURVEY_RESPONSE"],
        "confidence": "Medium",
        "notes": "Customer research via survey entities"
    },
    "Rich Media from all sources": {
        "entities": ["DIGITAL_CONTENT", "MEDIA_ASSET",
                      "DOCUMENT"],
        "confidence": "Low",
        "notes": "Rich media content - generic document entities"
    },
    "Voice Recording": {
        "entities": ["DIGITAL_CONTENT", "CALL_RECORDING",
                      "DIRECT_CONTACT_EVENT"],
        "confidence": "Low",
        "notes": "Voice/call recordings - needs extension"
    },
    "External Data - Fraud Watch Lists": {
        "entities": ["WATCH_LIST", "WATCH_LIST_ENTRY",
                      "PARTY_WATCH_LIST_CHECK"],
        "confidence": "High",
        "notes": "Fraud/sanctions watch lists"
    },
    "Compliance Approvals": {
        "entities": ["COMPLIANCE_EVENT", "APPROVAL",
                      "PARTY_COMPLIANCE_VERIFICATION"],
        "confidence": "Medium",
        "notes": "Regulatory compliance approvals"
    },
    "Business Process Definition": {
        "entities": ["EVENT", "PROCESS_EVENT"],
        "confidence": "Low",
        "notes": "Business process definitions - not in FSDM scope"
    },
    "External Data - External Loss Data": {
        "entities": ["LOSS_EVENT", "OPERATIONAL_LOSS_EVENT"],
        "confidence": "Medium",
        "notes": "External loss data for operational risk"
    },
    "System Log Files (E.g. Firewall, Webserver)": {
        "entities": ["WEB_VISIT", "SYSTEM_EVENT"],
        "confidence": "Low",
        "notes": "System/security logs - outside FSDM scope"
    },
    "Data Subject Areas required": {
        "entities": [],
        "confidence": "N/A",
        "notes": "Summary column - not a data requirement"
    },
}


def build_bvf_fsdm_mapping(bvf_data, fsdm_entities):
    """Build and export BVF-to-FSDM entity mapping (Task 2.1)."""
    print("\n" + "=" * 60)
    print("SECTION 3: BVF-to-FSDM Entity Mapping")
    print("=" * 60)

    data_to_cap = bvf_data["data_to_cap"]

    # Validate entity names against actual FSDM catalog
    all_fsdm_names = set(fsdm_entities.keys())
    validated_mapping = []

    for d2c in data_to_cap:
        dr = d2c["data_requirement"]
        sa = d2c["fsdm_subject_area"]

        if dr in BVF_TO_FSDM_MAPPING:
            m = BVF_TO_FSDM_MAPPING[dr]
            for entity in m["entities"]:
                exists = entity in all_fsdm_names
                domain = fsdm_entities[entity]["domain"] if exists else "Not Found"
                desc = fsdm_entities[entity]["description"][:200] if exists else ""
                validated_mapping.append({
                    "Data_Requirement": dr,
                    "FSDM_Subject_Area": sa,
                    "FSDM_Entity": entity,
                    "FSDM_Entity_Description": desc,
                    "Entity_In_XSD": "Yes" if exists else "No",
                    "FSDM_Domain": domain,
                    "Mapping_Confidence": m["confidence"],
                    "Notes": m["notes"],
                })

    # Export
    path = os.path.join(OUTPUT_DIR, "bvf_to_fsdm_entity_mapping.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "Data_Requirement", "FSDM_Subject_Area", "FSDM_Entity",
            "FSDM_Entity_Description", "Entity_In_XSD", "FSDM_Domain",
            "Mapping_Confidence", "Notes"])
        w.writeheader()
        w.writerows(validated_mapping)

    # Stats
    found = sum(1 for r in validated_mapping if r["Entity_In_XSD"] == "Yes")
    total = len(validated_mapping)
    print(f"  Mapped {len(data_to_cap)} data requirements -> {total} entity mappings")
    print(f"  Entities validated in XSD: {found}/{total} ({100*found//total}%)")
    print(f"  -> {path}")

    return validated_mapping


# ══════════════════════════════════════════════════════════════════════════
# SECTION 4: CAPABILITY DEPENDENCIES & REUSE SCORES
# ══════════════════════════════════════════════════════════════════════════

def build_capability_dependencies(bvf_data, fsdm_entities):
    """Build capability-to-FSDM dependency graph (Task 2.2) and reuse scores (Task 2.3)."""
    print("\n" + "=" * 60)
    print("SECTION 4: Capability Dependencies & Reuse Scores")
    print("=" * 60)

    cap_to_data = bvf_data["cap_to_data"]

    # 4a. Capability FSDM Dependencies
    deps = []
    for cap in cap_to_data:
        for dr_name, sa_name in cap["data_mappings"].items():
            entities = BVF_TO_FSDM_MAPPING.get(dr_name, {}).get("entities", [])
            entity_str = "; ".join(entities) if entities else "Not mapped"
            deps.append({
                "Theme": cap["theme"],
                "Capability": cap["capability"],
                "Sub_Capability": cap["sub_capability"],
                "Data_Requirement": dr_name,
                "FSDM_Subject_Area": sa_name,
                "FSDM_Entities": entity_str,
            })

    path = os.path.join(OUTPUT_DIR, "capability_fsdm_dependencies.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "Theme", "Capability", "Sub_Capability",
            "Data_Requirement", "FSDM_Subject_Area", "FSDM_Entities"])
        w.writeheader()
        w.writerows(deps)
    print(f"  -> {path} ({len(deps)} rows)")

    # 4b. FSDM Entity Reuse Scores
    entity_usage = defaultdict(set)  # entity -> set of sub-capabilities
    for cap in cap_to_data:
        for dr_name in cap["data_mappings"]:
            entities = BVF_TO_FSDM_MAPPING.get(dr_name, {}).get("entities", [])
            for ent in entities:
                entity_usage[ent].add(cap["sub_capability"])

    reuse_scores = []
    for entity, caps_set in sorted(entity_usage.items(), key=lambda x: -len(x[1])):
        count = len(caps_set)
        domain = fsdm_entities[entity]["domain"] if entity in fsdm_entities else "Unknown"
        # Priority tier: P1 (>70 caps), P2 (40-70), P3 (20-40), P4 (<20)
        if count >= 70:
            tier = "P1-Critical"
        elif count >= 40:
            tier = "P2-High"
        elif count >= 20:
            tier = "P3-Medium"
        else:
            tier = "P4-Low"

        reuse_scores.append({
            "FSDM_Entity": entity,
            "Subject_Area": domain,
            "Capabilities_Supported": count,
            "Reuse_Score": round(count / 112.0, 3),
            "Priority_Tier": tier,
        })

    path = os.path.join(OUTPUT_DIR, "fsdm_entity_reuse_scores.csv")
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "FSDM_Entity", "Subject_Area", "Capabilities_Supported",
            "Reuse_Score", "Priority_Tier"])
        w.writeheader()
        w.writerows(reuse_scores)
    print(f"  -> {path} ({len(reuse_scores)} entities)")

    p1 = [r for r in reuse_scores if r["Priority_Tier"] == "P1-Critical"]
    p2 = [r for r in reuse_scores if r["Priority_Tier"] == "P2-High"]
    print(f"  P1-Critical entities: {len(p1)}")
    print(f"  P2-High entities: {len(p2)}")

    return deps, reuse_scores


# ══════════════════════════════════════════════════════════════════════════
# SECTION 5: PROFITABILITY STAR SCHEMA (Enhanced with BVF)
# ══════════════════════════════════════════════════════════════════════════

def generate_star_schema():
    """Generate the enhanced profitability star schema DDL (Task 3.2)."""
    print("\n" + "=" * 60)
    print("SECTION 5: Profitability Star Schema DDL")
    print("=" * 60)

    ddl = r"""-- ================================================================
-- CUSTOMER PROFITABILITY ENGINE - Enhanced Star Schema
-- ================================================================
-- Source: Teradata FSDM v16.00.00 + Banking BVF v1.2
-- Target: Pakistani Bank (UBL-style) Customer Profitability
-- Generated: """ + datetime.now().strftime("%Y-%m-%d %H:%M") + r"""
--
-- Dimensions: Customer, Product, Branch, Business Segment,
--             Channel, Time, Agreement, Geography
-- Facts: Customer Profitability (monthly grain)
-- ================================================================

-- ==============================================
-- FACT TABLE: Customer Profitability
-- BVF Capabilities: Profitability Modelling, Activity Based Costing,
--   Funds Transfer Pricing, Performance Management and KPIs
-- FSDM Sources: AGREEMENT_SUMMARY, MONETARY_TRANSACTION,
--   AGREEMENT_RISK_METRIC, GL_MAIN_ACCOUNT
-- ==============================================
CREATE MULTISET TABLE FACT_CUSTOMER_PROFITABILITY, FALLBACK, NO JOURNAL
(
    -- Dimension Keys
    Customer_Party_Id       BIGINT          NOT NULL,   -- FK -> DIM_CUSTOMER (FSDM: PARTY.Party_Id)
    Agreement_Id            BIGINT          NOT NULL,   -- FK -> DIM_AGREEMENT (FSDM: AGREEMENT.Agreement_Id)
    Product_Id              BIGINT          NOT NULL,   -- FK -> DIM_PRODUCT (FSDM: PRODUCT.Product_Id)
    Branch_Org_Id           BIGINT          NOT NULL,   -- FK -> DIM_BRANCH (FSDM: ORGANIZATION.Organization_Party_Id)
    Business_Segment_Cd     VARCHAR(50),                -- FK -> DIM_BUSINESS_SEGMENT
    Channel_Instance_Id     BIGINT,                     -- FK -> DIM_CHANNEL (FSDM: CHANNEL_INSTANCE)
    Time_Period_Key         DATE            NOT NULL,   -- FK -> DIM_TIME (FSDM: TIME_PERIOD_TYPE)
    Geography_Area_Id       BIGINT,                     -- FK -> DIM_GEOGRAPHY (FSDM: GEOGRAPHICAL_AREA)
    Currency_Cd             VARCHAR(3)      DEFAULT 'PKR', -- FSDM: CURRENCY

    -- ── Revenue Measures ──────────────────────────────────────
    -- BVF: "Financial Transactions", "Account Fees, Commissions & Charges"
    -- FSDM: AGREEMENT_SUMMARY, MONETARY_TRANSACTION
    Interest_Income_Amt     DECIMAL(18,2),  -- Markup earned (Islamic: Profit earned)
    Interest_Expense_Amt    DECIMAL(18,2),  -- Markup paid (Islamic: Profit paid)
    Fee_Income_Amt          DECIMAL(18,2),  -- Account fees and charges
    Commission_Income_Amt   DECIMAL(18,2),  -- Commission income (LC, guarantees)
    FX_Income_Amt           DECIMAL(18,2),  -- Foreign exchange gain/loss
    Other_Income_Amt        DECIMAL(18,2),  -- Miscellaneous income
    Total_Revenue_Amt       DECIMAL(18,2),  -- = Sum of all income components

    -- ── Funds Transfer Pricing ────────────────────────────────
    -- BVF: "Funds Transfer Pricing", "Master & Reference Data - FTP rates"
    -- FSDM: INTEREST_RATE, INTEREST_RATE_INDEX, AGREEMENT_SUMMARY
    FTP_Rate_Pct            DECIMAL(10,6),  -- FTP rate applied (KIBOR-based)
    FTP_Credit_Amt          DECIMAL(18,2),  -- FTP credit (deposit side)
    FTP_Debit_Amt           DECIMAL(18,2),  -- FTP charge (lending side)
    Net_Interest_Income_Amt DECIMAL(18,2),  -- = Interest_Income - Interest_Expense - FTP net

    -- ── Cost Measures ─────────────────────────────────────────
    -- BVF: "Activity Based Costing" capability
    -- FSDM: GL_MAIN_ACCOUNT (allocated costs)
    Direct_Cost_Amt         DECIMAL(18,2),  -- Direct transaction costs
    Transaction_Cost_Amt    DECIMAL(18,2),  -- Per-transaction processing cost
    Servicing_Cost_Amt      DECIMAL(18,2),  -- Account servicing cost
    Channel_Cost_Amt        DECIMAL(18,2),  -- Channel delivery cost
    Allocated_Overhead_Amt  DECIMAL(18,2),  -- Overhead allocation (ABC method)
    Total_Cost_Amt          DECIMAL(18,2),  -- = Sum of all cost components

    -- ── Risk & Provisions ─────────────────────────────────────
    -- BVF: "Risk Exposures", "Provisions, Losses & Writeoffs"
    -- FSDM: AGREEMENT_RISK_METRIC, BANK_CAPITAL_SUMMARY
    Provision_Expense_Amt   DECIMAL(18,2),  -- ECL provision (IFRS 9 Stage 1/2/3)
    Expected_Loss_Amt       DECIMAL(18,2),  -- PD x LGD x EAD
    Risk_Weighted_Asset_Amt DECIMAL(18,2),  -- Basel III RWA
    Economic_Capital_Amt    DECIMAL(18,2),  -- Internal capital model
    IFRS9_Stage_Cd          VARCHAR(10),    -- Stage 1/2/3 classification

    -- ── Profitability Results ─────────────────────────────────
    -- BVF: "Profitability Results", "Profitability Analytics and Optimisation"
    -- Calculated/derived measures
    Gross_Profit_Amt        DECIMAL(18,2),  -- = Revenue - Cost
    Net_Profit_Amt          DECIMAL(18,2),  -- = Gross Profit - Provisions
    Pre_Tax_Profit_Amt      DECIMAL(18,2),  -- Before withholding tax
    RAROC_Pct               DECIMAL(10,6),  -- = Net Profit / Economic Capital
    Cost_To_Income_Ratio    DECIMAL(10,6),  -- = Total Cost / Total Revenue
    ROE_Pct                 DECIMAL(10,6),  -- = Net Profit / Allocated Equity
    Net_Interest_Margin_Pct DECIMAL(10,6),  -- = NII / Avg Balance

    -- ── Balance Measures ──────────────────────────────────────
    -- BVF: "Account Balances"
    -- FSDM: AGREEMENT_SUMMARY, BALANCE_SUMMARY
    Average_Balance_Amt     DECIMAL(18,2),  -- Period average balance
    EOP_Balance_Amt         DECIMAL(18,2),  -- End-of-period balance
    Original_Amount_Amt     DECIMAL(18,2),  -- Original sanctioned/deposit amount

    -- ── Pakistan-Specific ─────────────────────────────────────
    WHT_Amount_Amt          DECIMAL(18,2),  -- Withholding tax on profit
    Zakat_Deduction_Amt     DECIMAL(18,2),  -- Zakat deduction (Islamic obligation)
    Is_Islamic_Ind          CHAR(1),        -- Y/N: Islamic banking product

    -- ── Metadata ──────────────────────────────────────────────
    Calculation_Method_Cd   VARCHAR(20),    -- ACTUAL / ALLOCATED / ESTIMATED
    Data_Source_Cd          VARCHAR(50),    -- Source system identifier
    ETL_Batch_Id            BIGINT,
    ETL_Load_Dttm           TIMESTAMP       DEFAULT CURRENT_TIMESTAMP
)
PRIMARY INDEX (Customer_Party_Id, Time_Period_Key)
PARTITION BY RANGE_N(Time_Period_Key BETWEEN DATE '2020-01-01' AND DATE '2030-12-31' EACH INTERVAL '1' MONTH);

-- ==============================================
-- DIM: Customer (BVF: "Single Customer View")
-- FSDM: PARTY -> INDIVIDUAL / ORGANIZATION
-- ==============================================
CREATE MULTISET TABLE DIM_CUSTOMER, FALLBACK, NO JOURNAL
(
    Customer_Party_Id       BIGINT          NOT NULL,
    Customer_Type_Cd        VARCHAR(20),    -- Individual / Organization / Household
    Customer_Name           VARCHAR(200),
    Customer_Name_Urdu      VARCHAR(200),   -- Urdu name for Pakistan

    -- Segmentation (BVF: "Customer Segments")
    Customer_Segment_Cd     VARCHAR(50),    -- Retail/Corporate/SME/Agriculture/Micro
    Customer_SubSegment_Cd  VARCHAR(50),    -- Sub-segment classification
    Risk_Segment_Cd         VARCHAR(50),    -- BVF: "Risk Segments, Cohorts & Groupings"
    Profitability_Tier_Cd   VARCHAR(20),    -- Platinum/Gold/Silver/Bronze/Dormant
    Value_Band_Cd           VARCHAR(20),    -- High/Medium/Low value

    -- Individual attributes (FSDM: INDIVIDUAL)
    Gender_Cd               VARCHAR(10),
    Date_Of_Birth           DATE,
    Age_Band_Cd             VARCHAR(20),
    Marital_Status_Cd       VARCHAR(20),
    Occupation_Cd           VARCHAR(50),
    Income_Band_Cd          VARCHAR(20),
    Education_Level_Cd      VARCHAR(50),
    CNIC_Number             VARCHAR(15),    -- Pakistan National ID

    -- Organization attributes (FSDM: ORGANIZATION)
    Industry_Cd             VARCHAR(50),
    Business_Size_Cd        VARCHAR(20),    -- Large/Medium/Small/Micro
    Annual_Turnover_Band_Cd VARCHAR(20),
    NTN_Number              VARCHAR(20),    -- National Tax Number (Pakistan)

    -- Relationship
    Relationship_Start_Dt   DATE,
    Relationship_Tenure_Months INTEGER,
    Total_Products_Count    INTEGER,        -- Cross-sell depth
    Primary_Branch_Org_Id   BIGINT,         -- Home branch

    -- KYC / Compliance (BVF: "KYC Data")
    KYC_Status_Cd           VARCHAR(20),    -- Complete / Pending / Expired
    KYC_Last_Review_Dt      DATE,
    AML_Risk_Rating_Cd      VARCHAR(20),    -- High / Medium / Low
    PEP_Indicator           CHAR(1),        -- Politically Exposed Person

    -- Geography (FSDM: GEOGRAPHICAL_AREA)
    City_Cd                 VARCHAR(50),
    District_Cd             VARCHAR(50),
    Province_Cd             VARCHAR(50),    -- Punjab/Sindh/KPK/Balochistan/GB/AJK/ICT
    Country_Cd              VARCHAR(10)     DEFAULT 'PK',

    -- SCD Type 2
    Effective_From_Dt       DATE,
    Effective_To_Dt         DATE,
    Is_Current_Ind          CHAR(1)         DEFAULT 'Y'
)
UNIQUE PRIMARY INDEX (Customer_Party_Id);

-- ==============================================
-- DIM: Product (BVF: "Master & Reference Data - product")
-- FSDM: PRODUCT hierarchy
-- ==============================================
CREATE MULTISET TABLE DIM_PRODUCT, FALLBACK, NO JOURNAL
(
    Product_Id              BIGINT          NOT NULL,
    Product_Name            VARCHAR(100),
    Product_Type_Cd         VARCHAR(50),    -- Deposit/Loan/Card/Insurance/Investment/Trade Finance
    Product_SubType_Cd      VARCHAR(50),    -- Current/Savings/Fixed/Running Finance/etc.
    Product_Category_Cd     VARCHAR(50),
    Product_Family_Cd       VARCHAR(50),
    Asset_Liability_Cd      VARCHAR(10),    -- Asset / Liability / Off-Balance-Sheet
    Is_Financial_Product    CHAR(1),

    -- Pakistani Product Types
    Is_Islamic_Product      CHAR(1),        -- Shariah-compliant product
    Islamic_Mode_Cd         VARCHAR(50),    -- Murabaha/Musharakah/Ijarah/Diminishing Musharakah
    SBP_Product_Code        VARCHAR(20),    -- State Bank classification code

    -- Pricing (BVF: "Product Revenue, Cost & Margin", "FTP Rates")
    FTP_Rate_Pct            DECIMAL(10,6),  -- Funds Transfer Price rate
    Benchmark_Rate_Cd       VARCHAR(20),    -- KIBOR-3M / KIBOR-6M / Fixed
    Spread_Bps              DECIMAL(10,2),  -- Spread over benchmark (basis points)
    Standard_Fee_Amt        DECIMAL(18,2),
    Standard_Cost_Amt       DECIMAL(18,2),  -- BVF: "Product Revenue, Cost & Margin"
    Standard_Revenue_Amt    DECIMAL(18,2),
    Product_NPV_Amt         DECIMAL(18,2),  -- BVF: "Product NPV's"

    Product_Start_Dt        DATE,
    Product_End_Dt          DATE
)
UNIQUE PRIMARY INDEX (Product_Id);

-- ==============================================
-- DIM: Branch (BVF: "Master & Reference Data - organisation hierarchy")
-- FSDM: ORGANIZATION -> INTERNAL_ORGANIZATION_UNIT
-- ==============================================
CREATE MULTISET TABLE DIM_BRANCH, FALLBACK, NO JOURNAL
(
    Branch_Org_Id           BIGINT          NOT NULL,
    Branch_Name             VARCHAR(100),
    Branch_Code             VARCHAR(20),
    Branch_Type_Cd          VARCHAR(50),    -- Full Branch / Sub-Branch / Islamic Branch / Booth / Digital
    SBP_Branch_Code         VARCHAR(20),    -- SBP-assigned branch code

    -- Hierarchy (BVF: "Master & Reference Data - organisation hierarchy")
    Region_Name             VARCHAR(100),   -- Regional office
    Zone_Name               VARCHAR(100),   -- Zone/Area office
    Area_Name               VARCHAR(100),
    Group_Name              VARCHAR(100),   -- Group head office

    -- Location
    City_Name               VARCHAR(100),
    District_Name           VARCHAR(100),
    Province_Cd             VARCHAR(50),    -- Punjab/Sindh/KPK/Balochistan/GB/AJK/ICT
    Country_Cd              VARCHAR(10)     DEFAULT 'PK',
    Latitude_Val            DECIMAL(10,6),
    Longitude_Val           DECIMAL(10,6),

    -- Operations
    Branch_Open_Dt          DATE,
    Branch_Manager_Party_Id BIGINT,
    Employee_Count          INTEGER,

    -- Costing (for ABC allocation)
    Cost_Center_Cd          VARCHAR(20),
    Overhead_Pool_Cd        VARCHAR(20),
    Monthly_Operating_Cost  DECIMAL(18,2),

    -- Classification
    Is_Islamic_Branch       CHAR(1),        -- Islamic banking branch
    Is_Digital_Branch       CHAR(1),        -- Branchless/digital
    Rural_Urban_Cd          VARCHAR(10)     -- Rural / Urban / Semi-Urban
)
UNIQUE PRIMARY INDEX (Branch_Org_Id);

-- ==============================================
-- DIM: Business Segment
-- FSDM: ORGANIZATION_BUSINESS_TYPE
-- ==============================================
CREATE MULTISET TABLE DIM_BUSINESS_SEGMENT, FALLBACK, NO JOURNAL
(
    Business_Segment_Cd     VARCHAR(50)     NOT NULL,
    Segment_Name            VARCHAR(100),
    Segment_Group_Cd        VARCHAR(50),

    -- Pakistani banking segments
    -- Retail Banking, Corporate Banking, Commercial Banking,
    -- SME Banking, Agriculture Finance, Islamic Banking,
    -- Microfinance, Treasury, Trade Finance
    Segment_Head_Party_Id   BIGINT,
    Capital_Allocation_Pct  DECIMAL(7,4),
    Target_ROE_Pct          DECIMAL(7,4),
    Target_CIR_Pct          DECIMAL(7,4),
    SBP_Sector_Code         VARCHAR(20)     -- SBP sector classification
)
UNIQUE PRIMARY INDEX (Business_Segment_Cd);

-- ==============================================
-- DIM: Channel (BVF: "Master & Reference Data - Channel")
-- FSDM: CHANNEL_TYPE / CHANNEL_INSTANCE
-- ==============================================
CREATE MULTISET TABLE DIM_CHANNEL, FALLBACK, NO JOURNAL
(
    Channel_Instance_Id     BIGINT          NOT NULL,
    Channel_Type_Cd         VARCHAR(50),
    Channel_Name            VARCHAR(100),
    -- Branch, ATM, Internet Banking, Mobile App, SMS Banking,
    -- Call Center, POS, Agent Banking (branchless),
    -- JazzCash/Easypaisa integration, RAAST (instant payment)
    Channel_Cost_Per_Txn    DECIMAL(18,2),
    Is_Digital_Channel      CHAR(1),
    Is_Branchless_Channel   CHAR(1),        -- Pakistan branchless banking
    Channel_Status_Cd       VARCHAR(20)
)
UNIQUE PRIMARY INDEX (Channel_Instance_Id);

-- ==============================================
-- DIM: Time Period
-- FSDM: TIME_PERIOD_TYPE
-- ==============================================
CREATE MULTISET TABLE DIM_TIME, FALLBACK, NO JOURNAL
(
    Time_Period_Key         DATE            NOT NULL,
    Calendar_Year           INTEGER,
    Calendar_Quarter        INTEGER,
    Calendar_Month          INTEGER,
    Calendar_Week           INTEGER,
    Calendar_Day            INTEGER,
    Day_Of_Week             INTEGER,
    Month_Name              VARCHAR(20),

    -- Pakistan fiscal year (July-June)
    Fiscal_Year             INTEGER,        -- e.g., FY2024 = Jul 2023 - Jun 2024
    Fiscal_Quarter          INTEGER,
    Fiscal_Month            INTEGER,        -- 1=July, 12=June

    -- Banking calendar
    Is_Business_Day         CHAR(1),
    Is_Friday               CHAR(1),        -- Pakistan weekly holiday
    Is_SBP_Reporting_Dt     CHAR(1),        -- State Bank of Pakistan reporting date
    Is_Month_End            CHAR(1),
    Is_Quarter_End          CHAR(1),
    Is_Year_End             CHAR(1),
    Is_Eid_Holiday          CHAR(1),        -- Eid-ul-Fitr / Eid-ul-Adha
    Is_National_Holiday     CHAR(1)
)
UNIQUE PRIMARY INDEX (Time_Period_Key);

-- ==============================================
-- DIM: Agreement/Account Detail
-- FSDM: AGREEMENT + subtypes
-- ==============================================
CREATE MULTISET TABLE DIM_AGREEMENT, FALLBACK, NO JOURNAL
(
    Agreement_Id            BIGINT          NOT NULL,
    Agreement_Type_Cd       VARCHAR(50),    -- Loan/Deposit/Card/Insurance/Investment
    Agreement_Subtype_Cd    VARCHAR(50),    -- Running Finance/Term Loan/Current Account/etc.
    Agreement_Name          VARCHAR(100),
    Host_System_Cd          VARCHAR(20),    -- Source core banking system
    Host_Agreement_Num      VARCHAR(50),    -- Core banking account number

    -- Status (BVF: "Account Status")
    Account_Status_Cd       VARCHAR(20),    -- Active/Dormant/Inactive/Closed
    Agreement_Open_Dt       DATE,
    Agreement_Close_Dt      DATE,
    Last_Activity_Dt        DATE,

    -- Terms (BVF: "Account Terms & Conditions", "Account Maturity Dates")
    Maturity_Dt             DATE,
    Next_Repricing_Dt       DATE,
    Interest_Rate_Pct       DECIMAL(10,6),
    Rate_Type_Cd            VARCHAR(20),    -- Fixed / Floating / Islamic Profit Rate
    Benchmark_Cd            VARCHAR(20),    -- KIBOR-3M/6M/1Y
    Tenure_Months           INTEGER,

    -- Classification
    Asset_Liability_Cd      VARCHAR(10),    -- Asset / Liability / Off-BS
    Product_Id              BIGINT,
    Customer_Party_Id       BIGINT,
    Branch_Org_Id           BIGINT,
    Business_Segment_Cd     VARCHAR(50),

    -- Risk (BVF: "Risk Exposures")
    Risk_Grade_Cd           VARCHAR(20),    -- SBP classification (1-9)
    SBP_Classification_Cd   VARCHAR(50),    -- Normal/OAEM/Substandard/Doubtful/Loss
    IFRS9_Stage_Cd          VARCHAR(10),

    -- Collateral (BVF: "Collateral Agreements")
    Collateral_Ind          CHAR(1),
    Collateral_Value_Amt    DECIMAL(18,2),
    Collateral_Type_Cd      VARCHAR(50),
    LTV_Ratio_Pct           DECIMAL(10,4),

    -- Islamic Banking
    Is_Islamic_Ind          CHAR(1),
    Islamic_Mode_Cd         VARCHAR(50)     -- Murabaha/Musharakah/Ijarah/Salam
)
UNIQUE PRIMARY INDEX (Agreement_Id);

-- ==============================================
-- DIM: Geography
-- FSDM: GEOGRAPHICAL_AREA
-- ==============================================
CREATE MULTISET TABLE DIM_GEOGRAPHY, FALLBACK, NO JOURNAL
(
    Geography_Area_Id       BIGINT          NOT NULL,
    Country_Cd              VARCHAR(10),
    Country_Name            VARCHAR(100),
    Province_Cd             VARCHAR(50),
    Province_Name           VARCHAR(100),
    Division_Name           VARCHAR(100),
    District_Name           VARCHAR(100),
    Tehsil_Name             VARCHAR(100),
    City_Name               VARCHAR(100),
    Area_Type_Cd            VARCHAR(20),    -- Metro / Urban / Semi-Urban / Rural
    SBP_Region_Cd           VARCHAR(20)     -- SBP geographic classification
)
UNIQUE PRIMARY INDEX (Geography_Area_Id);

-- ==============================================
-- AGGREGATE TABLE: Branch Profitability
-- Monthly branch-level rollup
-- ==============================================
CREATE MULTISET TABLE AGG_BRANCH_PROFITABILITY, FALLBACK, NO JOURNAL
(
    Branch_Org_Id           BIGINT          NOT NULL,
    Business_Segment_Cd     VARCHAR(50),
    Time_Period_Key         DATE            NOT NULL,

    -- Aggregated measures
    Customer_Count          INTEGER,
    Agreement_Count         INTEGER,
    Total_Revenue_Amt       DECIMAL(18,2),
    Total_NII_Amt           DECIMAL(18,2),
    Total_Fee_Income_Amt    DECIMAL(18,2),
    Total_Cost_Amt          DECIMAL(18,2),
    Total_Provision_Amt     DECIMAL(18,2),
    Net_Profit_Amt          DECIMAL(18,2),
    Total_Balance_Amt       DECIMAL(18,2),
    Total_RWA_Amt           DECIMAL(18,2),
    Avg_RAROC_Pct           DECIMAL(10,6),
    Cost_To_Income_Pct      DECIMAL(10,6)
)
PRIMARY INDEX (Branch_Org_Id, Time_Period_Key);

-- ==============================================
-- AGGREGATE TABLE: Segment Profitability
-- Monthly segment-level rollup
-- ==============================================
CREATE MULTISET TABLE AGG_SEGMENT_PROFITABILITY, FALLBACK, NO JOURNAL
(
    Business_Segment_Cd     VARCHAR(50)     NOT NULL,
    Product_Type_Cd         VARCHAR(50),
    Time_Period_Key         DATE            NOT NULL,

    Customer_Count          INTEGER,
    Agreement_Count         INTEGER,
    Total_Revenue_Amt       DECIMAL(18,2),
    Total_NII_Amt           DECIMAL(18,2),
    Total_Fee_Income_Amt    DECIMAL(18,2),
    Total_Cost_Amt          DECIMAL(18,2),
    Total_Provision_Amt     DECIMAL(18,2),
    Net_Profit_Amt          DECIMAL(18,2),
    Total_Balance_Amt       DECIMAL(18,2),
    Total_RWA_Amt           DECIMAL(18,2),
    Avg_RAROC_Pct           DECIMAL(10,6),
    ROE_Pct                 DECIMAL(10,6)
)
PRIMARY INDEX (Business_Segment_Cd, Time_Period_Key);

-- ==============================================
-- VIEWS: Profitability Calculation
-- ==============================================

-- Customer Profitability Summary View
REPLACE VIEW VW_CUSTOMER_PROFITABILITY_SUMMARY AS
SELECT
    c.Customer_Party_Id,
    c.Customer_Name,
    c.Customer_Segment_Cd,
    c.Profitability_Tier_Cd,
    f.Time_Period_Key,
    SUM(f.Total_Revenue_Amt)        AS Total_Revenue,
    SUM(f.Net_Interest_Income_Amt)  AS Total_NII,
    SUM(f.Fee_Income_Amt + f.Commission_Income_Amt) AS Total_Non_Interest_Income,
    SUM(f.Total_Cost_Amt)           AS Total_Cost,
    SUM(f.Provision_Expense_Amt)    AS Total_Provisions,
    SUM(f.Net_Profit_Amt)           AS Net_Profit,
    CASE WHEN SUM(f.Economic_Capital_Amt) > 0
         THEN SUM(f.Net_Profit_Amt) / SUM(f.Economic_Capital_Amt)
         ELSE 0 END                 AS RAROC,
    CASE WHEN SUM(f.Total_Revenue_Amt) > 0
         THEN SUM(f.Total_Cost_Amt) / SUM(f.Total_Revenue_Amt)
         ELSE 0 END                 AS Cost_To_Income_Ratio
FROM FACT_CUSTOMER_PROFITABILITY f
JOIN DIM_CUSTOMER c ON f.Customer_Party_Id = c.Customer_Party_Id
                    AND c.Is_Current_Ind = 'Y'
GROUP BY 1, 2, 3, 4, 5;

-- Product Profitability View
REPLACE VIEW VW_PRODUCT_PROFITABILITY AS
SELECT
    p.Product_Id,
    p.Product_Name,
    p.Product_Type_Cd,
    p.Asset_Liability_Cd,
    p.Is_Islamic_Product,
    f.Time_Period_Key,
    COUNT(DISTINCT f.Customer_Party_Id) AS Customer_Count,
    COUNT(DISTINCT f.Agreement_Id)      AS Account_Count,
    SUM(f.Average_Balance_Amt)          AS Total_Balance,
    SUM(f.Total_Revenue_Amt)            AS Total_Revenue,
    SUM(f.Net_Interest_Income_Amt)      AS Total_NII,
    SUM(f.Total_Cost_Amt)               AS Total_Cost,
    SUM(f.Net_Profit_Amt)               AS Net_Profit,
    CASE WHEN SUM(f.Average_Balance_Amt) > 0
         THEN SUM(f.Net_Interest_Income_Amt) / SUM(f.Average_Balance_Amt) * 100
         ELSE 0 END                     AS NIM_Pct
FROM FACT_CUSTOMER_PROFITABILITY f
JOIN DIM_PRODUCT p ON f.Product_Id = p.Product_Id
GROUP BY 1, 2, 3, 4, 5, 6;

-- Islamic vs Conventional Comparison View
REPLACE VIEW VW_ISLAMIC_VS_CONVENTIONAL AS
SELECT
    f.Time_Period_Key,
    f.Is_Islamic_Ind,
    CASE f.Is_Islamic_Ind WHEN 'Y' THEN 'Islamic Banking'
                          ELSE 'Conventional Banking' END AS Banking_Type,
    COUNT(DISTINCT f.Customer_Party_Id)     AS Customer_Count,
    SUM(f.Average_Balance_Amt)              AS Total_Balance,
    SUM(f.Total_Revenue_Amt)                AS Total_Revenue,
    SUM(f.Net_Profit_Amt)                   AS Net_Profit,
    CASE WHEN SUM(f.Total_Revenue_Amt) > 0
         THEN SUM(f.Total_Cost_Amt) / SUM(f.Total_Revenue_Amt)
         ELSE 0 END                         AS Cost_To_Income_Ratio
FROM FACT_CUSTOMER_PROFITABILITY f
GROUP BY 1, 2, 3;
"""

    path = os.path.join(OUTPUT_DIR, "profitability_star_schema_enhanced.sql")
    with open(path, "w", encoding="utf-8") as f:
        f.write(ddl)
    print(f"  -> {path}")
    return ddl


# ══════════════════════════════════════════════════════════════════════════
# SECTION 6: DATA LINEAGE
# ══════════════════════════════════════════════════════════════════════════

def generate_data_lineage():
    """Generate data lineage JSON (Task 3.4)."""
    print("\n" + "=" * 60)
    print("SECTION 6: Data Lineage")
    print("=" * 60)

    lineage = {
        "metadata": {
            "generated": datetime.now().isoformat(),
            "model": "FSDM v16.00.00 + BVF v1.2",
            "target": "Customer Profitability Engine (Pakistani Bank)",
        },
        "lineage": [
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Interest_Income_Amt",
                "calculation": "SUM(AGREEMENT_SUMMARY.Interest_Income_Amt) for period",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY"],
                "bvf_data_requirements": ["Financial Transactions", "Account Balances"],
                "bvf_capabilities": ["Profitability Modelling", "Financial Accounting / Accounting Hub"],
                "notes": "Markup earned for conventional; profit earned for Islamic products"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Interest_Expense_Amt",
                "calculation": "SUM(AGREEMENT_SUMMARY.Interest_Expense_Amt) for period",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY"],
                "bvf_data_requirements": ["Financial Transactions", "Account Balances"],
                "bvf_capabilities": ["Profitability Modelling"],
                "notes": "Markup paid on deposits; profit shared for Islamic"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Net_Interest_Income_Amt",
                "calculation": "Interest_Income_Amt - Interest_Expense_Amt + FTP_Credit_Amt - FTP_Debit_Amt",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY", "INTEREST_RATE"],
                "bvf_data_requirements": ["Financial Transactions", "FTP Rates", "Account Balances"],
                "bvf_capabilities": ["Funds Transfer Pricing", "Profitability Modelling"],
                "notes": "Net interest income after FTP adjustment; KIBOR-based FTP curves"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Fee_Income_Amt",
                "calculation": "SUM(MONETARY_TRANSACTION.Transaction_Amt) WHERE type IN ('FEE','CHARGE')",
                "source_fsdm_entities": ["MONETARY_TRANSACTION", "FEE_TRANSACTION"],
                "bvf_data_requirements": ["Account Fees, Commissions & Charges"],
                "bvf_capabilities": ["Profitability Modelling", "Financial Accounting / Accounting Hub"],
                "notes": "Account fees, service charges, penalty charges"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Commission_Income_Amt",
                "calculation": "SUM(MONETARY_TRANSACTION.Transaction_Amt) WHERE type IN ('COMMISSION','LC_FEE','GUARANTEE_FEE')",
                "source_fsdm_entities": ["MONETARY_TRANSACTION", "COMMISSION_TRANSACTION"],
                "bvf_data_requirements": ["Account Fees, Commissions & Charges", "Financial Transactions"],
                "bvf_capabilities": ["Profitability Modelling"],
                "notes": "LC commissions, guarantee fees, trade finance commissions"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "FX_Income_Amt",
                "calculation": "SUM(FX gain/loss from CURRENCY_EXCHANGE_RATE revaluation)",
                "source_fsdm_entities": ["CURRENCY_EXCHANGE_RATE", "MONETARY_TRANSACTION"],
                "bvf_data_requirements": ["Master & Reference Data - exchange rates", "Financial Transactions"],
                "bvf_capabilities": ["FX & Trading Book Management", "Profitability Modelling"],
                "notes": "Foreign exchange gains and losses"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "FTP_Rate_Pct",
                "calculation": "INTEREST_RATE.Rate_Value for product/tenor matched FTP curve",
                "source_fsdm_entities": ["INTEREST_RATE", "INTEREST_RATE_INDEX"],
                "bvf_data_requirements": ["Master & Reference Data - FTP rates"],
                "bvf_capabilities": ["Funds Transfer Pricing"],
                "notes": "KIBOR-based FTP rate curve by tenor bucket"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "FTP_Credit_Amt",
                "calculation": "Average_Balance_Amt * FTP_Rate_Pct * Days/365 (for deposits)",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY", "INTEREST_RATE"],
                "bvf_data_requirements": ["FTP Rates", "Account Balances"],
                "bvf_capabilities": ["Funds Transfer Pricing"],
                "notes": "FTP credit for liability (deposit) side"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "FTP_Debit_Amt",
                "calculation": "Average_Balance_Amt * FTP_Rate_Pct * Days/365 (for loans)",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY", "INTEREST_RATE"],
                "bvf_data_requirements": ["FTP Rates", "Account Balances"],
                "bvf_capabilities": ["Funds Transfer Pricing"],
                "notes": "FTP charge for asset (lending) side"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Direct_Cost_Amt",
                "calculation": "Activity cost drivers * unit cost rates (ABC allocation)",
                "source_fsdm_entities": ["GL_MAIN_ACCOUNT", "CHANNEL_USAGE_METRIC"],
                "bvf_data_requirements": ["Finance Systems - General Ledger ", "Financial Transactions"],
                "bvf_capabilities": ["Activity Based Costing"],
                "notes": "Direct costs allocated via ABC model"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Transaction_Cost_Amt",
                "calculation": "COUNT(transactions) * Channel_Cost_Per_Txn by channel type",
                "source_fsdm_entities": ["MONETARY_TRANSACTION", "CHANNEL_USAGE_METRIC", "CHANNEL_INSTANCE"],
                "bvf_data_requirements": ["Financial Transactions", "Master & Reference Data - Channel"],
                "bvf_capabilities": ["Activity Based Costing"],
                "notes": "Per-transaction costs by channel"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Allocated_Overhead_Amt",
                "calculation": "Branch overhead * allocation_key (balance, revenue, or headcount based)",
                "source_fsdm_entities": ["GL_MAIN_ACCOUNT", "INTERNAL_ORGANIZATION_UNIT"],
                "bvf_data_requirements": ["Finance Systems - General Ledger ", "Master & Reference Data - organisation hierarchy"],
                "bvf_capabilities": ["Activity Based Costing", "GL, AP, HR, Expense Analytics & Optimisation"],
                "notes": "Overhead allocation from GL cost centers to customer"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Provision_Expense_Amt",
                "calculation": "AGREEMENT_RISK_METRIC.Provision_Amt (IFRS9 ECL model)",
                "source_fsdm_entities": ["AGREEMENT_RISK_METRIC"],
                "bvf_data_requirements": ["Provisions, Losses & Writeoffs", "Risk Exposures"],
                "bvf_capabilities": ["Credit Risk Expected Loss Model", "Profitability Modelling"],
                "notes": "Expected credit loss (ECL) under IFRS 9 Stage 1/2/3"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Expected_Loss_Amt",
                "calculation": "PD * LGD * EAD from AGREEMENT_RISK_METRIC",
                "source_fsdm_entities": ["AGREEMENT_RISK_METRIC", "PARTY_LIABILITY_CREDIT_RATING"],
                "bvf_data_requirements": ["Risk Exposures", "Predictive Models - risk"],
                "bvf_capabilities": ["Credit Risk Expected Loss Model"],
                "notes": "PD x LGD x EAD calculation"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Risk_Weighted_Asset_Amt",
                "calculation": "EAD * Risk_Weight from AGREEMENT_RISK_METRIC (Basel III SA/IRB)",
                "source_fsdm_entities": ["AGREEMENT_RISK_METRIC", "BANK_CAPITAL_SUMMARY"],
                "bvf_data_requirements": ["Risk Weighted Assets & Capital Results"],
                "bvf_capabilities": ["Risk Weighted Assets and Regulatory Capital", "Capital Management (Planning and Allocation)"],
                "notes": "Basel III risk-weighted assets for RAROC denominator"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "RAROC_Pct",
                "calculation": "Net_Profit_Amt / Economic_Capital_Amt * 100",
                "source_fsdm_entities": ["Calculated"],
                "bvf_data_requirements": ["Profitability Results", "Risk Weighted Assets & Capital Results"],
                "bvf_capabilities": ["Profitability Modelling", "Profitability Analytics and Optimisation"],
                "notes": "Risk-Adjusted Return on Capital - key profitability KPI"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Cost_To_Income_Ratio",
                "calculation": "Total_Cost_Amt / Total_Revenue_Amt",
                "source_fsdm_entities": ["Calculated"],
                "bvf_data_requirements": ["Profitability Results"],
                "bvf_capabilities": ["Performance Management and KPIs", "Profitability Analytics and Optimisation"],
                "notes": "Cost efficiency ratio - SBP benchmark monitoring"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Average_Balance_Amt",
                "calculation": "AVG(daily AGREEMENT_SUMMARY.Balance_Amt) over period",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY", "BALANCE_SUMMARY"],
                "bvf_data_requirements": ["Account Balances"],
                "bvf_capabilities": ["Profitability Modelling", "Asset & Liability Management"],
                "notes": "Average daily balance for the reporting period"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "WHT_Amount_Amt",
                "calculation": "Profit_Amount * WHT_Rate (currently 15%/35% based on filer status)",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY"],
                "bvf_data_requirements": ["Financial Transactions", "Account Balances"],
                "bvf_capabilities": ["Tax Management & Optimisation"],
                "notes": "Pakistan-specific: Withholding tax on bank profit"
            },
            {
                "target_table": "FACT_CUSTOMER_PROFITABILITY",
                "target_column": "Zakat_Deduction_Amt",
                "calculation": "Balance_Amt * 2.5% (on Zakat valuation date, 1st Ramadan)",
                "source_fsdm_entities": ["AGREEMENT_SUMMARY"],
                "bvf_data_requirements": ["Account Balances"],
                "bvf_capabilities": ["Financial Accounting / Accounting Hub"],
                "notes": "Pakistan-specific: Compulsory Zakat deduction on eligible accounts"
            },
            {
                "target_table": "DIM_CUSTOMER",
                "target_column": "Customer_Segment_Cd",
                "calculation": "PARTY_CLASSIFICATION.Classification_Value where type='SEGMENT'",
                "source_fsdm_entities": ["PARTY_CLASSIFICATION", "PARTY_CLASSIFICATION_VALUE"],
                "bvf_data_requirements": ["Customer Segments"],
                "bvf_capabilities": ["Customer Segmentation", "Insight Led Customer Strategy"],
                "notes": "Customer value segmentation"
            },
            {
                "target_table": "DIM_CUSTOMER",
                "target_column": "Risk_Segment_Cd",
                "calculation": "RISK_GRADE_VALUE.Grade_Cd from customer risk model",
                "source_fsdm_entities": ["RISK_GRADE_VALUE", "PARTY_LIABILITY_CREDIT_RATING"],
                "bvf_data_requirements": ["Risk Segments, Cohorts & Groupings"],
                "bvf_capabilities": ["Credit Risk Scoring (Underwriting Likelihood Model)"],
                "notes": "Customer-level risk grade"
            },
            {
                "target_table": "DIM_PRODUCT",
                "target_column": "FTP_Rate_Pct",
                "calculation": "INTEREST_RATE.Rate_Value matched by product tenor to KIBOR curve",
                "source_fsdm_entities": ["INTEREST_RATE", "INTEREST_RATE_INDEX", "PRODUCT"],
                "bvf_data_requirements": ["Master & Reference Data - FTP rates"],
                "bvf_capabilities": ["Funds Transfer Pricing"],
                "notes": "Product-level FTP rate from KIBOR yield curve"
            },
        ]
    }

    path = os.path.join(OUTPUT_DIR, "data_lineage.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(lineage, f, indent=2)
    print(f"  -> {path} ({len(lineage['lineage'])} lineage entries)")
    return lineage


# ══════════════════════════════════════════════════════════════════════════
# SECTION 7: BVF COVERAGE REPORT
# ══════════════════════════════════════════════════════════════════════════

def generate_bvf_coverage_report(bvf_data):
    """Generate profitability BVF capability coverage report (Task 3.3)."""
    print("\n" + "=" * 60)
    print("SECTION 7: BVF Coverage Report")
    print("=" * 60)

    profitability_caps = [
        "Activity Based Costing",
        "Profitability Modelling",
        "Future / Lifetime Value",
        "Profitability Analytics and Optimisation",
        "Performance Management and KPIs",
        "Pricing Analysis & Optimisation",
        "Financial Budgeting, Planning & Forecasting",
        "GL, AP, HR, Expense Analytics & Optimisation",
        "Business Process Analytics & Optimisation",
        "Funds Transfer Pricing",
        "Asset & Liability Management",
        "Capital Planning and Management",
        "Financial Accounting / Accounting Hub",
        "Functional Profit & Loss Statement (P&L)",
    ]

    # Build lookup
    cap_lookup = {}
    for cap in bvf_data["cap_to_data"]:
        cap_lookup[cap["sub_capability"]] = cap

    report = []
    report.append("# BVF Profitability Capability Coverage Report")
    report.append(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    report.append(f"**Source:** BVF v1.2 + FSDM v16.00.00")
    report.append(f"**Target:** Pakistani Bank Customer Profitability Engine\n")

    report.append("## Executive Summary\n")
    report.append("This report maps each profitability-related BVF capability to:")
    report.append("1. Required data requirements from the BVF")
    report.append("2. FSDM entities that supply that data")
    report.append("3. Star schema tables/columns that implement it")
    report.append("4. Gap analysis: what's missing")
    report.append("5. Pakistan/UBL-specific considerations\n")

    report.append("---\n")

    # Star schema column mapping for each key data requirement
    star_schema_map = {
        "Financial Transactions": "FACT_CUSTOMER_PROFITABILITY: Interest_Income_Amt, Fee_Income_Amt, Commission_Income_Amt, FX_Income_Amt",
        "Account Balances": "FACT_CUSTOMER_PROFITABILITY: Average_Balance_Amt, EOP_Balance_Amt",
        "Account Fees, Commissions & Charges": "FACT_CUSTOMER_PROFITABILITY: Fee_Income_Amt, Commission_Income_Amt",
        "Profitability Results": "FACT_CUSTOMER_PROFITABILITY: Gross_Profit_Amt, Net_Profit_Amt, RAROC_Pct, Cost_To_Income_Ratio, ROE_Pct",
        "Master & Reference Data - FTP rates": "DIM_PRODUCT: FTP_Rate_Pct; FACT: FTP_Rate_Pct, FTP_Credit_Amt, FTP_Debit_Amt",
        "Product Revenue, Cost & Margin": "DIM_PRODUCT: Standard_Cost_Amt, Standard_Revenue_Amt; FACT: Total_Revenue_Amt, Total_Cost_Amt",
        "Product NPV's": "DIM_PRODUCT: Product_NPV_Amt",
        "Risk Exposures": "FACT_CUSTOMER_PROFITABILITY: Risk_Weighted_Asset_Amt, Economic_Capital_Amt",
        "Risk Weighted Assets & Capital Results": "FACT_CUSTOMER_PROFITABILITY: Risk_Weighted_Asset_Amt, RAROC_Pct",
        "Provisions, Losses & Writeoffs": "FACT_CUSTOMER_PROFITABILITY: Provision_Expense_Amt, Expected_Loss_Amt",
        "Finance Systems - General Ledger ": "Allocated costs: Direct_Cost_Amt, Allocated_Overhead_Amt",
        "Finance Systems - ERP": "Indirect cost feeds for ABC allocation",
        "Finance Systems - Budgets, Plans & Forecasts": "AGG tables: Target vs Actual comparison views",
        "Master & Reference Data - organisation hierarchy": "DIM_BRANCH: All hierarchy columns; DIM_BUSINESS_SEGMENT",
        "Customer Segments": "DIM_CUSTOMER: Customer_Segment_Cd, Customer_SubSegment_Cd, Profitability_Tier_Cd",
        "Customer Demographics": "DIM_CUSTOMER: Gender_Cd, Age_Band_Cd, Occupation_Cd, Income_Band_Cd",
        "Single Customer View (Master Record)": "DIM_CUSTOMER: All columns - core customer dimension",
        "Master & Reference Data - product ": "DIM_PRODUCT: All columns - core product dimension",
        "Master & Reference Data - Channel": "DIM_CHANNEL: All columns - channel dimension",
        "Account / Policy / Service Holdings": "DIM_AGREEMENT: Agreement_Type_Cd, Account_Status_Cd",
        "Account / Policy / Service Detail": "DIM_AGREEMENT: Rate, tenor, terms columns",
        "Collateral Agreements": "DIM_AGREEMENT: Collateral_Ind, Collateral_Value_Amt, LTV_Ratio_Pct",
    }

    pakistan_notes = {
        "Funds Transfer Pricing": "Use KIBOR (Karachi Interbank Offered Rate) as benchmark instead of LIBOR. KIBOR tenors: O/N, 1W, 2W, 1M, 3M, 6M, 9M, 1Y. SBP publishes daily KIBOR rates.",
        "Activity Based Costing": "Include Pakistan-specific cost drivers: branch security costs (higher than global avg), cash handling costs (cash-heavy economy), regulatory compliance costs (SBP, FBR, SECP).",
        "Profitability Modelling": "Must handle dual banking (conventional + Islamic) profitability. Islamic products use profit-sharing ratios instead of interest rates. WHT deduction impacts net customer profitability.",
        "Financial Accounting / Accounting Hub": "Pakistan follows IFRS. SBP issues BSD circulars for banking-specific accounting. Chart of Accounts must align with SBP prescribed format.",
        "Capital Planning and Management": "Follow SBP's Basel III implementation timeline. MCR (Minimum Capital Requirement) currently PKR 10B for commercial banks. CAR minimum 11.5% (with CCB).",
        "Performance Management and KPIs": "SBP monitors key ratios: CAR, NPL ratio, provision coverage, ROA, ROE, CIR. Quarterly CAMEL rating framework applies.",
        "Pricing Analysis & Optimisation": "SBP Policy Rate is the key benchmark. Spread caps may apply for certain sectors (agriculture, SME, housing). Minimum deposit rates mandated by SBP.",
        "Financial Budgeting, Planning & Forecasting": "Pakistan fiscal year is July-June. Budget cycles align with SBP monetary policy announcements (6-8 per year).",
        "Asset & Liability Management": "KIBOR-based ALM gap analysis. SBP requires regular ALM reporting. Currency risk management for PKR/USD/EUR/GBP/SAR positions.",
    }

    gaps = {
        "Activity Based Costing": "FSDM lacks dedicated ABC/cost allocation entities. Need custom extension tables for cost drivers, activity pools, and allocation rules.",
        "Future / Lifetime Value": "FSDM has no CLV/LTV calculation entities. Need extension for customer lifetime value models and discount rate assumptions.",
        "Business Process Analytics & Optimisation": "FSDM has minimal business process modeling. Need BPM integration or custom process event tables.",
        "Financial Budgeting, Planning & Forecasting": "FSDM has limited budget/forecast entities. Need custom BUDGET, FORECAST, and FINANCIAL_PLAN tables.",
        "Pricing Analysis & Optimisation": "FSDM covers basic product pricing but lacks price elasticity and optimization model entities.",
    }

    for cap_name in profitability_caps:
        cap_info = cap_lookup.get(cap_name)
        if not cap_info:
            continue

        report.append(f"## {cap_name}")
        report.append(f"**Theme:** {cap_info['theme']}")
        report.append(f"**Capability Group:** {cap_info['capability']}")
        report.append(f"**Data Requirements Count:** {cap_info['data_count']}\n")

        # Data requirements
        report.append("### Data Requirements & FSDM Mapping\n")
        report.append("| Data Requirement | FSDM Subject Area | FSDM Entities | Star Schema |")
        report.append("|-----------------|-------------------|---------------|-------------|")

        for dr_name, sa_name in sorted(cap_info["data_mappings"].items()):
            entities = BVF_TO_FSDM_MAPPING.get(dr_name, {}).get("entities", [])
            entity_str = ", ".join(entities[:4])
            if len(entities) > 4:
                entity_str += "..."
            ss_map = star_schema_map.get(dr_name, "—")
            if len(ss_map) > 60:
                ss_map = ss_map[:57] + "..."
            report.append(f"| {dr_name[:45]} | {sa_name} | {entity_str} | {ss_map} |")

        report.append("")

        # Gap analysis
        if cap_name in gaps:
            report.append(f"### Gap Analysis\n")
            report.append(f"> {gaps[cap_name]}\n")

        # Pakistan notes
        if cap_name in pakistan_notes:
            report.append(f"### Pakistan/UBL Considerations\n")
            report.append(f"> {pakistan_notes[cap_name]}\n")

        report.append("---\n")

    # Summary statistics
    report.append("## Coverage Statistics\n")
    total_drs = set()
    covered_drs = set()
    for cap_name in profitability_caps:
        cap_info = cap_lookup.get(cap_name)
        if cap_info:
            for dr in cap_info["data_mappings"]:
                total_drs.add(dr)
                if dr in BVF_TO_FSDM_MAPPING:
                    covered_drs.add(dr)

    report.append(f"| Metric | Value |")
    report.append(f"|--------|-------|")
    report.append(f"| Profitability capabilities analyzed | {len(profitability_caps)} |")
    report.append(f"| Unique data requirements needed | {len(total_drs)} |")
    report.append(f"| Data requirements mapped to FSDM | {len(covered_drs)} |")
    report.append(f"| Coverage percentage | {100*len(covered_drs)//max(len(total_drs),1)}% |")

    high_conf = sum(1 for dr in covered_drs
                    if BVF_TO_FSDM_MAPPING.get(dr, {}).get("confidence") == "High")
    report.append(f"| High-confidence mappings | {high_conf} |")
    report.append(f"| Medium/Low confidence (need review) | {len(covered_drs) - high_conf} |")

    path = os.path.join(OUTPUT_DIR, "profitability_bvf_coverage.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(report))
    print(f"  -> {path}")


# ══════════════════════════════════════════════════════════════════════════
# SECTION 8: PAKISTAN BANKING CONTEXT
# ══════════════════════════════════════════════════════════════════════════

def generate_pakistan_context():
    """Generate Pakistan banking context document (Task 4)."""
    print("\n" + "=" * 60)
    print("SECTION 8: Pakistan Banking Context")
    print("=" * 60)

    doc = r"""# Pakistan Banking Context for Profitability Engine
**Generated:** """ + datetime.now().strftime("%Y-%m-%d %H:%M") + r"""
**Applicable to:** Pakistani commercial banks (UBL-style implementation)

---

## 1. Regulatory Framework

### State Bank of Pakistan (SBP)
- **Role:** Central bank and primary banking regulator
- **Key Regulations:**
  - Banking Companies Ordinance, 1962
  - SBP BSD (Banking Supervision Department) Circulars
  - Prudential Regulations for Corporate/Commercial Banking
  - Prudential Regulations for Consumer Finance
  - Prudential Regulations for SME Finance
  - Prudential Regulations for Agriculture Finance
  - Prudential Regulations for Microfinance Banks

### Basel III Implementation
- **CAR Requirement:** Minimum 11.5% (including Capital Conservation Buffer of 2.5%)
- **Minimum Capital:** PKR 10 billion for commercial banks
- **Leverage Ratio:** Minimum 3%
- **LCR/NSFR:** Fully implemented
- **Pillar 2 (ICAAP):** Required for all banks
- **Pillar 3 (Disclosure):** Quarterly public disclosure

### IFRS Adoption
- Pakistan has adopted IFRS standards
- **IFRS 9** (Financial Instruments): Fully implemented - ECL model
- **IFRS 16** (Leases): Implemented
- **IFRS 15** (Revenue): Implemented
- Banking-specific guidance via SBP circulars

### Impact on Profitability Engine
- IFRS9_Stage_Cd in fact table (Stage 1/2/3) for ECL provisioning
- SBP_Classification_Cd for regulatory classification (1-9 scale)
- Regulatory capital feeds for RAROC calculation
- Quarterly SBP reporting dates in DIM_TIME

---

## 2. Currency & Rate Benchmarks

### Base Currency
- **PKR (Pakistani Rupee)** as primary reporting currency
- Multi-currency support for:
  - USD (US Dollar) - primary foreign currency
  - EUR (Euro)
  - GBP (British Pound)
  - SAR (Saudi Riyal) - remittance corridor
  - AED (UAE Dirham) - remittance corridor
  - CNY (Chinese Yuan) - CPEC-related trade

### KIBOR (Karachi Interbank Offered Rate)
- **Replaces LIBOR** as the primary FTP benchmark
- Published daily by Financial Markets Association of Pakistan (FMAP)
- **Tenors:** Overnight, 1-Week, 2-Week, 1-Month, 3-Month, 6-Month, 9-Month, 1-Year
- FTP curve construction uses KIBOR interpolation
- **SBP Policy Rate** anchors the short end of the yield curve

### FTP Implementation
```
FTP Curve Construction:
  O/N to 1M  -> SBP Policy Rate + liquidity premium
  1M to 1Y   -> KIBOR interpolated curve
  1Y to 5Y   -> KIBOR 1Y + term premium (swap-implied)
  5Y+         -> Government bond yield curve
```

### Impact on Star Schema
- Currency_Cd defaults to 'PKR' in FACT table
- DIM_PRODUCT.Benchmark_Rate_Cd references KIBOR tenors
- FTP_Rate_Pct derived from KIBOR-based curves

---

## 3. Banking Segments (Pakistani Context)

### Standard Segments
| Segment | Description | SBP Classification |
|---------|-------------|-------------------|
| Retail Banking | Individual customers, personal banking | Consumer Finance |
| Corporate Banking | Large corporates (turnover > PKR 800M) | Corporate/Commercial |
| Commercial Banking | Mid-market companies | Corporate/Commercial |
| SME Banking | Small & Medium Enterprises (turnover < PKR 800M) | SME Finance |
| Agriculture Finance | Farmers, agri-businesses | Agriculture Finance |
| Islamic Banking | Shariah-compliant products | Islamic Banking |
| Microfinance | Low-income, unbanked population | Microfinance |
| Treasury | Interbank, investment, trading | Treasury |
| Trade Finance | Import/export financing, LC, guarantees | Trade Finance |

### SBP Mandatory Targets
- Agriculture lending: 20% of total advances
- SME lending: Specific targets set annually
- Housing finance: Growing priority sector
- Low-cost deposit accounts (Asaan Account): Financial inclusion target

---

## 4. Channel Landscape

### Physical Channels
| Channel | Description | Cost Category |
|---------|-------------|--------------|
| Branch | Full-service branch network | High cost |
| Sub-Branch | Limited service branch | Medium cost |
| Islamic Branch | Dedicated Islamic banking | High cost |
| Booth / Extension Counter | Minimal service point | Low cost |
| ATM | Cash dispensing and basic services | Medium cost |

### Digital Channels
| Channel | Description | Cost Category |
|---------|-------------|--------------|
| Internet Banking | Web-based banking portal | Low cost |
| Mobile App | Bank's proprietary mobile application | Low cost |
| SMS Banking | Text-based banking services | Very low cost |
| USSD Banking | Feature phone banking | Very low cost |

### Branchless Banking / Digital Wallets
| Channel | Description | Notes |
|---------|-------------|-------|
| JazzCash | Mobilink-backed mobile wallet | Largest in Pakistan |
| Easypaisa | Telenor-backed mobile wallet | Pioneer in mobile money |
| Agent Banking | Retail agents for deposits/withdrawals | SBP-regulated agents |
| RAAST | SBP's instant payment system | Launched 2021, growing rapidly |
| 1Link | ATM/POS switch network | Interbank connectivity |

### Impact on Profitability
- Channel_Cost_Per_Txn varies significantly (branch PKR 150-300 vs digital PKR 5-15)
- Branchless banking agents have commission-based cost model
- RAAST transactions have near-zero marginal cost
- Channel migration from branch to digital is a key profitability driver

---

## 5. Tax Considerations

### Withholding Tax (WHT) on Bank Profit
- **Filers:** 15% WHT on profit/interest earned above PKR 500,000/year
- **Non-filers:** 30% WHT (higher rate to encourage tax filing)
- Deducted at source by the bank
- WHT_Amount_Amt column in FACT table captures this

### Zakat Deduction
- Compulsory deduction of 2.5% on eligible accounts on 1st Ramadan
- Applies to savings accounts, fixed deposits, and other eligible instruments
- Exemptions available via CZ-50 form (for Shia Muslims and others)
- Zakat_Deduction_Amt column in FACT table

### Corporate Tax
- Standard corporate tax rate for banking: 39% (super tax applicable)
- Impacts segment-level profitability reporting
- Not deducted at individual customer level in profitability model

---

## 6. Islamic Banking Considerations

### Overview
- Pakistan has a dedicated Islamic Banking framework regulated by SBP
- Banks operate Islamic windows or standalone Islamic banking subsidiaries
- Shariah Board approval required for all Islamic products

### Islamic Product Modes
| Mode | Type | Equivalent |
|------|------|-----------|
| Murabaha | Cost-plus sale | Working capital / trade finance |
| Diminishing Musharakah | Declining partnership | Home finance / auto finance |
| Ijarah | Lease | Equipment / vehicle leasing |
| Musharakah | Partnership | Project finance |
| Salam | Advance purchase | Agriculture finance |
| Istisna | Manufacturing order | Construction finance |
| Wakalah | Agency | Investment / deposits |
| Mudarabah | Silent partnership | Savings deposits |

### Profitability Implications
- **No interest:** Profit-sharing ratios replace interest rates
- **FTP adjustment:** Islamic FTP uses notional KIBOR-equivalent benchmark
- **Revenue recognition:** Profit recognized differently (e.g., Murabaha profit recognized over tenure)
- **Risk sharing:** Musharakah involves actual risk sharing with customer
- **Is_Islamic_Ind** and **Islamic_Mode_Cd** columns enable separate reporting

---

## 7. UBL-Specific Context

### Bank Profile
- **United Bank Limited (UBL):** One of Pakistan's largest banks
- **Ownership:** Bestway Group (majority), Government of Pakistan (minority)
- **Network:** 1,300+ branches across Pakistan + international presence
- **Assets:** Among top 5 Pakistani banks by total assets

### Technology Stack (Known)
- **Core Banking:** CTL-based systems (likely Temenos or similar)
- **Data Warehouse:** Teradata platform
  - Historical: 680 SMP → TD 2850 → IntelliFlex migration path
  - FSDM version at UBL: **v13.00.00** (this analysis uses v16.00.00)
- **Reporting:** Mix of BI tools

### FSDM Version Gap: v13 → v16
| Area | v13 | v16 Enhancement |
|------|-----|----------------|
| Party | Core party entities | Enhanced digital identity, social media |
| Agreement | Standard agreements | Extended risk metrics, IFRS 9 support |
| Risk | Basic risk entities | Full Basel III, stress testing, IFRS 9 ECL |
| Channel | Traditional channels | Digital/mobile, web analytics, omnichannel |
| Campaign | Basic CRM | Full marketing automation support |
| Web Analytics | Limited | Full clickstream, session, visitor tracking |
| Investment | Basic | Enhanced trading, portfolio management |
| Social Media | Not present | Full social media entity set |

### Key Gaps for UBL Migration
1. **IFRS 9 entities** (new in v16): AGREEMENT_RISK_METRIC enhancements for ECL staging
2. **Digital channel entities** (enhanced in v16): Web analytics, mobile app tracking
3. **Enhanced risk entities**: Stress testing scenarios, advanced Basel III support
4. **Social media integration**: New entity group not in v13
5. **Marketing automation**: Campaign entity enhancements

---

## 8. Data Quality Considerations

### Common Challenges in Pakistani Banking
1. **Customer deduplication:** Multiple records for same customer (CNIC-based matching needed)
2. **Address standardization:** Urdu/English address inconsistencies, no standard postal code system
3. **Industry classification:** Inconsistent SIC/ISIC coding across systems
4. **Islamic vs Conventional segregation:** Clear tagging needed for regulatory reporting
5. **Historical data gaps:** Core banking migrations may create data quality breaks
6. **Branch hierarchy changes:** Frequent organizational restructuring

### Recommendations
- CNIC as golden key for individual customer matching
- NTN (National Tax Number) for corporate customer matching
- SBP branch codes as authoritative branch identifiers
- KIBOR rates from FMAP as official rate source
- SBP exchange rates for regulatory reporting (market rates for management reporting)
"""

    path = os.path.join(OUTPUT_DIR, "pakistan_banking_context.md")
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)
    print(f"  -> {path}")


# ══════════════════════════════════════════════════════════════════════════
# MAIN EXECUTION
# ══════════════════════════════════════════════════════════════════════════

def main():
    print("╔══════════════════════════════════════════════════════════════╗")
    print("║  BVF-FSDM Integration & Profitability Engine               ║")
    print("║  FSDM v16.00.00 + Banking BVF v1.2                        ║")
    print("║  Target: Pakistani Bank Customer Profitability             ║")
    print("╚══════════════════════════════════════════════════════════════╝")
    print()

    # Section 1: Parse BVF
    bvf_data = parse_bvf_excel()

    # Section 2: Load FSDM
    fsdm_entities, fsdm_rels, fsdm_stats, fsdm_domain_map = load_fsdm_data()

    # Section 3: BVF-to-FSDM entity mapping
    entity_mapping = build_bvf_fsdm_mapping(bvf_data, fsdm_entities)

    # Section 4: Capability dependencies and reuse scores
    deps, reuse_scores = build_capability_dependencies(bvf_data, fsdm_entities)

    # Section 5: Star schema DDL
    generate_star_schema()

    # Section 6: Data lineage
    generate_data_lineage()

    # Section 7: BVF coverage report
    generate_bvf_coverage_report(bvf_data)

    # Section 8: Pakistan context
    generate_pakistan_context()

    print("\n" + "=" * 60)
    print("COMPLETE - All outputs written to:")
    print(f"  {OUTPUT_DIR}/")
    print("=" * 60)

    # List outputs
    for f in sorted(os.listdir(OUTPUT_DIR)):
        size = os.path.getsize(os.path.join(OUTPUT_DIR, f))
        print(f"  {f:50s} {size:>10,} bytes")


if __name__ == "__main__":
    main()
